#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
probe_xlsx.py — an xlsx region inspector: deep-inspect a rectangular block of
an .xlsx worksheet so its full structure AND formatting can be reproduced /
generalised elsewhere (e.g. a status table that is "made to be looked at").

It dumps, for the given sheet + range:
  1. column widths / row heights
  2. merged-cell ranges
  3. freeze panes
  4. data validations (dropdowns, e.g. Pass / Fail / NA)  -- intersecting the range
  5. conditional formatting rules                        -- intersecting the range
  6. VALUES  (one line per non-empty cell:  C5: "text")
  7. FORMAT CATALOG  (distinct cell formats, de-duplicated -> ids a, b, c, ...)
     each = font / fill / borders / alignment / number-format
  8. FORMAT MAP  (a grid of those ids, so the visual structure is obvious)

Usage:
    python probe_xlsx.py book.xlsx                  # active sheet, used range
    python probe_xlsx.py book.xlsx Sheet1           # named sheet, used range
    python probe_xlsx.py book.xlsx Sheet1 B2:H40    # named sheet, that block

Output is written to "<book>_<sheet>_probe.txt" (UTF-8). Open it, redact any
sensitive VALUES if needed (keep the FORMAT sections intact) and paste it back.

Requires:  pip install openpyxl
"""

import sys
import os

try:
    import openpyxl
    from openpyxl.utils import range_boundaries, get_column_letter
except ImportError:
    print("Missing dependency: pip install openpyxl")
    sys.exit(1)

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

OUT = []
def emit(s=""):
    OUT.append(str(s))


# ---------------------------------------------------------------------------
def color_str(c):
    """openpyxl Color -> short description: rgb / theme+tint / indexed. Uses .type to avoid mis-access."""
    if c is None:
        return "-"
    try:
        t = getattr(c, "type", None)
        if t == "rgb":
            rgb = c.rgb
            return ("#" + rgb) if (isinstance(rgb, str) and rgb != "00000000") else "-"
        if t == "theme":
            tint = getattr(c, "tint", 0) or 0
            return f"theme{c.theme}/tint{round(tint, 3)}"
        if t == "indexed":
            return f"indexed{c.indexed}"
        rgb = getattr(c, "rgb", None)
        if isinstance(rgb, str) and rgb != "00000000":
            return "#" + rgb
    except Exception:
        return "-"
    return "-"


def font_sig(f):
    return (
        f.name, f.size, bool(f.bold), bool(f.italic),
        (f.underline or "none"), color_str(f.color),
    )


def fill_sig(fl):
    try:
        if fl is None or fl.patternType in (None, "none"):
            return ("none", "-", "-")
        return (fl.patternType, color_str(fl.fgColor), color_str(fl.bgColor))
    except Exception:
        return ("none", "-", "-")


def side_sig(s):
    if s is None or s.style is None:
        return "-"
    return f"{s.style}:{color_str(s.color)}"


def border_sig(b):
    return (side_sig(b.top), side_sig(b.bottom), side_sig(b.left), side_sig(b.right))


def align_sig(a):
    return (
        a.horizontal or "-", a.vertical or "-",
        bool(a.wrap_text), int(a.text_rotation or 0), int(a.indent or 0),
    )


def cell_sig(cell):
    return (
        font_sig(cell.font),
        fill_sig(cell.fill),
        border_sig(cell.border),
        align_sig(cell.alignment),
        cell.number_format or "General",
    )


def describe_sig(sig):
    font, fill, border, align, numfmt = sig
    name, size, bold, italic, under, color = font
    fl_pat, fl_fg, fl_bg = fill
    bt, bb, bl, br = border
    h, v, wrap, rot, indent = align
    parts = []
    name = name or "default"; size = size if size is not None else "default"
    parts.append(f"font {name} {size}pt"
                 + ("/bold" if bold else "") + ("/italic" if italic else "")
                 + (f"/underline{under}" if under != "none" else "")
                 + (f"/color{color}" if color != "-" else ""))
    if fl_pat != "none":
        parts.append(f"fill {fl_pat} fg{fl_fg} bg{fl_bg}")
    if any(x != "-" for x in border):
        parts.append(f"border top{bt} bottom{bb} left{bl} right{br}")
    al = f"align H={h} V={v}"
    if wrap: al += " wrap"
    if rot: al += f" rotate{rot}"
    if indent: al += f" indent{indent}"
    parts.append(al)
    if numfmt != "General":
        parts.append(f"number-format [{numfmt}]")
    return "  |  ".join(parts)


# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    if not os.path.isfile(path):
        print("File not found:", path)
        sys.exit(1)

    # data_only=True returns the displayed values (cached formula results)
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif sheet:
        print(f"No such sheet: {sheet}; available: {wb.sheetnames}")
        sys.exit(1)
    else:
        ws = wb.active

    rng = sys.argv[3] if len(sys.argv) > 3 else None
    if rng:
        min_c, min_r, max_c, max_r = range_boundaries(rng)
    else:
        min_c, min_r = 1, 1
        max_c, max_r = ws.max_column, ws.max_row
        rng = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"

    emit("=" * 70)
    emit(f"WORKBOOK: {os.path.basename(path)}")
    emit(f"SHEET:    {ws.title}    RANGE: {rng}    "
         f"({max_c - min_c + 1} cols x {max_r - min_r + 1} rows)")
    emit(f"All sheets: {wb.sheetnames}")
    emit("=" * 70)

    # ---- Column widths / Row heights ----
    emit("\n## Column widths (character units)")
    cw = []
    for c in range(min_c, max_c + 1):
        L = get_column_letter(c)
        dim = ws.column_dimensions.get(L)
        w = round(dim.width, 2) if (dim and dim.width) else "default"
        cw.append(f"{L}={w}")
    emit("  " + "  ".join(cw))
    emit("\n## Row heights (points)")
    rh = []
    for r in range(min_r, max_r + 1):
        dim = ws.row_dimensions.get(r)
        h = round(dim.height, 1) if (dim and dim.height) else "default"
        rh.append(f"{r}={h}")
    emit("  " + "  ".join(rh))

    # ---- Freeze panes ----
    emit(f"\n## Freeze panes: {ws.freeze_panes or 'none'}")

    # ---- Merged cells ----
    emit("\n## Merged cells (intersecting the range)")
    merges = []
    for mc in ws.merged_cells.ranges:
        if not (mc.max_col < min_c or mc.min_col > max_c or
                mc.max_row < min_r or mc.min_row > max_r):
            merges.append(str(mc))
    if merges:
        for m in merges:
            emit("  " + m)
    else:
        emit("  none")

    # ---- Data validation (dropdowns, etc.) ----
    emit("\n## Data validation / dropdowns (intersecting the range)")
    found_dv = False
    try:
        for dv in ws.data_validations.dataValidation:
            sq = str(dv.sqref)
            emit(f"  type={dv.type} sqref={sq} formula1={dv.formula1} "
                 f"allowBlank={dv.allow_blank}")
            found_dv = True
    except Exception as e:
        emit(f"  (read error: {e})")
    if not found_dv:
        emit("  none")

    # ---- Conditional formatting ----
    emit("\n## Conditional formatting (whole sheet; focus on rules intersecting the range)")
    found_cf = False
    try:
        for rng_cf in ws.conditional_formatting:
            sqref = str(getattr(rng_cf, "sqref", rng_cf))
            rules = ws.conditional_formatting[rng_cf]
            for rule in rules:
                dxf_desc = ""
                if getattr(rule, "dxf", None) is not None:
                    dxf = rule.dxf
                    f = color_str(dxf.fill.fgColor) if dxf.fill else "-"
                    fc = color_str(dxf.font.color) if (dxf.font and dxf.font.color) else "-"
                    dxf_desc = f" -> fill{f} fontcolor{fc}"
                formula = getattr(rule, "formula", None)
                emit(f"  range {sqref}: {rule.type} "
                     f"{getattr(rule,'operator','')} {formula or ''}{dxf_desc}")
                found_cf = True
    except Exception as e:
        emit(f"  (read error: {e})")
    if not found_cf:
        emit("  none")

    # ---- Values ----
    emit("\n## Values (non-empty only; merged ranges hold the value only in the top-left cell)")
    anchors = {}
    for m in merges:
        a = m.split(":")[0]
        anchors[a] = m
    nonempty = 0
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            nonempty += 1
            addr = f"{get_column_letter(c)}{r}"
            sv = str(v).replace("\n", "\\n")
            if len(sv) > 80:
                sv = sv[:80] + "…"
            mark = f"  [merged {anchors[addr]}]" if addr in anchors else ""
            emit(f"  {addr}: {sv!r}{mark}")
    emit(f"  ({nonempty} non-empty cells total)")

    # ---- Format catalog + map ----
    emit("\n## Format catalog (de-duplicated; each = font/fill/border/align/number-format)")
    sig_to_id = {}
    id_order = []
    grid = []
    import string
    def next_id(n):
        # a..z, then a1..z1, ...
        base = string.ascii_lowercase
        if n < 26:
            return base[n]
        return base[n % 26] + str(n // 26)
    for r in range(min_r, max_r + 1):
        rowids = []
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            sig = cell_sig(cell)
            if sig not in sig_to_id:
                fid = next_id(len(id_order))
                sig_to_id[sig] = fid
                id_order.append(sig)
            rowids.append(sig_to_id[sig])
        grid.append(rowids)
    for i, sig in enumerate(id_order):
        emit(f"  [{next_id(i)}] {describe_sig(sig)}")

    emit("\n## Format map (row=Excel row number, col=C..; each cell is a format id from above)")
    header = "      " + " ".join(get_column_letter(c).rjust(3) for c in range(min_c, max_c + 1))
    emit(header)
    for idx, r in enumerate(range(min_r, max_r + 1)):
        line = f"  {r:>3} " + " ".join(fid.rjust(3) for fid in grid[idx])
        emit(line)

    text = "\n".join(OUT)
    stem = os.path.splitext(path)[0]
    out_path = f"{stem}_{ws.title}_probe.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"[written to] {out_path}  ({len(OUT)} lines)")
    print("Open in a UTF-8 editor; redact sensitive values if needed, keep the FORMAT sections, paste back.")


if __name__ == "__main__":
    main()
