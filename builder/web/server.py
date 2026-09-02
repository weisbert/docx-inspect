#!/usr/bin/env python3
"""Local HTTP server for the structured document builder.

Standard-library only (http.server / json / urllib / base64 / sqlite3).
Binds to 127.0.0.1 exclusively. Serves the single-page app and the JSON API
defined in CONTRACT.md. All domain-specific content (chapter skeleton, cover
fields, fixed body texts, style numbers, validation rules, UI strings, logo)
lives in an external template config file loaded at runtime; this module stays
neutral and domain-agnostic.

Run:
    python server.py --port 8765 --root <reports_root> --config <template.json>
"""

import argparse
import base64
import contextlib
import copy
import datetime
import hashlib
import io
import json
import os
import re
import shutil
import sys
import tempfile
import threading
import time
import traceback
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

HERE = os.path.dirname(os.path.abspath(__file__))          # builder/web
BUILDER = os.path.dirname(HERE)                            # builder
if HERE not in sys.path:
    sys.path.insert(0, HERE)
if BUILDER not in sys.path:
    sys.path.insert(0, BUILDER)

import buildpath  # noqa: E402,F401  (side effect: registers core/docx_io/store/sync)
import templates_store as tstore  # noqa: E402

# ---------------------------------------------------------------------------
# Server configuration (populated in main()).
# ---------------------------------------------------------------------------


# Version of this tool, reported by GET /api/version. Bumped by hand on a
# release; the endpoint never touches the network, so "available" stays null and
# "needsRestart" false until an updater fills them in.
TOOL_VERSION = "2.0.0"


class Config:
    reports_root = None          # absolute path; project dirs must stay under it
    template_config_path = None  # path to the active template config JSON
    bind = "127.0.0.1"
    port = 8765
    _template_cache = None       # (path, mtime) -> parsed dict


CFG = Config()


# ---------------------------------------------------------------------------
# Helpers: template config loading.
# ---------------------------------------------------------------------------


def _load_json_file(path):
    """Read a JSON file tolerating a leading UTF-8 BOM (a common artefact from
    Windows editors / Excel / Notepad). ``utf-8-sig`` strips the BOM if present
    and decodes identically to ``utf-8`` otherwise, so this is a safe drop-in
    for every project.json / template config read in this module. Raises
    ``ValueError`` (``json.JSONDecodeError`` is one) or ``OSError`` on
    failure; callers turn that into a clean error naming ``path`` instead of a
    raw 500 / uncaught traceback."""
    with open(path, "r", encoding="utf-8-sig") as fh:
        return json.load(fh)


def load_template_config(path=None):
    """Load and cache the template config JSON. Returns a dict or raises."""
    p = path or CFG.template_config_path
    if not p:
        raise FileNotFoundError("no template config path configured")
    p = os.path.abspath(p)
    if not os.path.isfile(p):
        raise FileNotFoundError("template config not found: %s" % p)
    mtime = os.path.getmtime(p)
    cache = CFG._template_cache
    if cache and cache[0] == p and cache[1] == mtime:
        return cache[2]
    data = _load_json_file(p)
    CFG._template_cache = (p, mtime, data)
    return data


def config_slice(template_query=None):
    """Frontend-facing slice of the template config (API #3).

    Returns only what the GUI needs; never leaks engine-only style numbers. When
    a library template id is given and exists, slice THAT template's config;
    otherwise fall back to the global template config.
    """
    tc = None
    if template_query:
        p = tstore.template_config_path(CFG.reports_root, template_query)
        if p:
            tc = load_template_config(p)
    if tc is None:
        tc = load_template_config()
    return {
        "template": tc.get("id"),
        "skeleton": tc.get("skeleton", []),
        "ui_strings": tc.get("ui_strings", {}),
        "cover": tc.get("cover", {}),
        "compliance": _compliance_defaults(tc.get("compliance", {})),
        # fixed_bodies: the standard paragraphs a section can be pinned to
        # (`node.fixed_body` is a key into this map). The GUI needs them to show
        # what such a section holds and to turn it into ordinary editable blocks
        # when the user overrides it; the engine reads the same map, so the two
        # sides agree on what the section says.
        "fixed_bodies": tc.get("fixed_bodies", {}),
        # table_presets: company-defined starter tables (with condition rows
        # baked in). Content lives in the local template config; the engine/app
        # stay neutral -- this just forwards the list to the block picker (WS4).
        "table_presets": tc.get("table_presets", []),
    }


def _compliance_defaults(comp):
    """Expose only the compliance defaults the GUI needs for live preview."""
    return {
        "axis_labels": comp.get("axis_labels", ["MIN", "TYP", "MAX", "NTWC"]),
        "default_limit": comp.get("default_limit", {}),
        "flag_color": comp.get("flag_color", "FF0000"),
        "setting_kinds": comp.get(
            "setting_kinds", ["common_setting", "module_setting", "tb"]
        ),
    }


# ---------------------------------------------------------------------------
# Helpers: path safety.
# ---------------------------------------------------------------------------


def resolve_project_dir(dir_arg, create=False):
    """Resolve a project dir from the `dir` query param.

    `dir` may be absolute or relative to reports_root. The resolved path must
    stay under reports_root (if one is configured). Raises ValueError on any
    traversal outside the allowed root.
    """
    if not dir_arg:
        raise ValueError("missing 'dir' parameter")
    root = CFG.reports_root
    # A6: fail CLOSED when no reports root is configured. Previously a missing
    # root skipped the containment check entirely (fail-open), so any absolute /
    # relative path resolved unchecked. Mirror _api_apply_update's stance.
    if not root:
        raise ValueError("no reports root configured")
    if os.path.isabs(dir_arg):
        target = os.path.abspath(dir_arg)
    else:
        target = os.path.abspath(os.path.join(root, dir_arg))

    root_abs = os.path.abspath(root)
    # containment check that is robust to case/sep on Windows
    try:
        common = os.path.commonpath([root_abs, target])
    except ValueError:
        # different drives
        raise ValueError("path escapes reports_root")
    if os.path.normcase(common) != os.path.normcase(root_abs):
        raise ValueError("path escapes reports_root")

    if create:
        os.makedirs(target, exist_ok=True)
        os.makedirs(os.path.join(target, "images"), exist_ok=True)
    return target


def atomic_write(path, data_bytes):
    """Write bytes to path atomically (temp file + os.replace)."""
    d = os.path.dirname(path)
    os.makedirs(d, exist_ok=True)
    tmp = path + ".tmp.%d" % os.getpid()
    with open(tmp, "wb") as fh:
        fh.write(data_bytes)
        fh.flush()
        os.fsync(fh.fileno())
    os.replace(tmp, path)


# ---------------------------------------------------------------------------
# Helpers: local auto-snapshots (belt-and-suspenders backup of project.json).
#
# Every save -- and every overwrite by apply-update / paste-import -- first drops
# a timestamped copy of project.json into <project_dir>/_autosave/, kept rolling.
# This is INDEPENDENT of the apply-time _backups/ history, so an external update
# can never leave local edits unrecoverable: the last saved state is always
# sitting in _autosave/ and is one click to restore.
# ---------------------------------------------------------------------------

AUTOSAVE_DIRNAME = "_autosave"
AUTOSAVE_KEEP = 300  # rolling retention per project

# Root-level bookkeeping dirs that are never valid as a project folder name.
_RESERVED_ROOT_DIRS = {"_backups", "_updates", "_outbox", "_trash",
                       "_autosave", "__pycache__", "assets"}


# The one reason that names no event: the periodic capture of a saved file.
# Everything else is something that happened to the report, and the timeline is
# the only place it is recorded.
AUTOSAVE_PLAIN_REASON = "save"


def _is_untagged_reason(reason):
    return (_sanitize_name(reason) or AUTOSAVE_PLAIN_REASON) == AUTOSAVE_PLAIN_REASON


def _autosave_dir(project_dir):
    return os.path.join(project_dir, AUTOSAVE_DIRNAME)


def _autosave_paths(project_dir):
    """Snapshot paths in chronological order (oldest first). Sorted by mtime
    (tie-break by name) so same-second collision suffixes can't scramble order."""
    d = _autosave_dir(project_dir)
    if not os.path.isdir(d):
        return []
    entries = []
    for n in os.listdir(d):
        if not n.endswith(".json"):
            continue
        p = os.path.join(d, n)
        try:
            mt = os.path.getmtime(p)
        except OSError:
            mt = 0.0
        entries.append((mt, n, p))
    entries.sort()
    return [p for _, _, p in entries]


def autosave_snapshot(project_dir, reason="save"):
    """Copy the current project.json into _autosave/<ts>__<reason>.json.

    An UNTAGGED snapshot -- the ordinary per-save one -- is skipped when its
    bytes repeat the newest snapshot's, because a save that changed nothing is
    not a state worth keeping.

    A TAGGED one is always written. It is taken immediately before an apply, a
    paste, a restore or a rename, so the per-save snapshot that precedes it
    holds exactly the same bytes, and deduping by content therefore threw away
    the only record that the event happened at all: the timeline read back as an
    unbroken row of automatic snapshots, `Roll back` was never offered, and the
    Restores filter was permanently empty. The bytes may repeat; the event does
    not.

    Prunes to the newest AUTOSAVE_KEEP. Best-effort: never raises -- a snapshot
    must never break a save. Returns the snapshot filename or None.
    """
    try:
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return None
        with open(pj, "rb") as fh:
            data = fh.read()
        existing = _autosave_paths(project_dir)
        if existing and _is_untagged_reason(reason):
            try:
                with open(existing[-1], "rb") as fh:
                    if fh.read() == data:
                        return None  # unchanged since last snapshot
            except OSError:
                pass
        os.makedirs(_autosave_dir(project_dir), exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        tag = _sanitize_name(reason) or "save"
        base = "%s__%s" % (ts, tag)
        name = base + ".json"
        k = 1
        while os.path.exists(os.path.join(_autosave_dir(project_dir), name)):
            name = "%s-%d.json" % (base, k)
            k += 1
        atomic_write(os.path.join(_autosave_dir(project_dir), name), data)
        for p in _autosave_paths(project_dir)[:-AUTOSAVE_KEEP]:
            try:
                os.remove(p)
            except OSError:
                pass
        return name
    except Exception:
        return None


def autosave_all(reports_root, reason="preapply"):
    """Snapshot every project under reports_root (used before an apply)."""
    made = []
    try:
        for name in sorted(os.listdir(reports_root)):
            pdir = os.path.join(reports_root, name)
            if os.path.isfile(os.path.join(pdir, "project.json")):
                snap = autosave_snapshot(pdir, reason)
                if snap:
                    made.append("%s/%s" % (name, snap))
    except OSError:
        pass
    return made


# How far down the timeline a content hint is computed. A hint costs one file
# read per entry (memoised below), the drawer shows a scrolling list, and 300
# rolling snapshots of a real report are megabytes -- so the newest stretch gets
# the detail and the tail keeps the shape it always had.
AUTOSAVE_HINT_DEPTH = 40
AUTOSAVE_HINT_TITLES = 3        # named on the entry; the rest are counted

# A snapshot file never changes once written, so its digest is memoised on
# (path, mtime, size). Bounded, because a long-running server sees every
# snapshot of every report it serves.
_SNAPSHOT_DIGESTS = {}
_SNAPSHOT_DIGESTS_MAX = 4000


def _section_digest(project):
    """{section key -> (title, location, fingerprint)} for one report state.

    The fingerprint covers a section's OWN fields -- children are sections in
    their own right and are keyed separately -- so a change is reported against
    the section it happened in rather than against every ancestor it sits under.
    """
    out = {}

    def walk(nodes, prefix):
        pos = 0
        for node in (nodes or []):
            if not isinstance(node, dict):
                continue
            pos += 1
            loc = "%s%d" % (prefix, pos)
            key = str(node.get("id") or "") or ("@" + loc)
            own = {k: v for k, v in node.items() if k != "children"}
            try:
                blob = json.dumps(own, sort_keys=True, ensure_ascii=False,
                                  default=str)
            except Exception:
                blob = repr(sorted(own.keys()))
            out[key] = (str(node.get("title") or ""), loc,
                        hashlib.sha1(blob.encode("utf-8")).hexdigest()[:16])
            walk(node.get("children"), loc + ".")

    walk((project or {}).get("outline"), "")
    return out


def _snapshot_state(path):
    """(section digest, cover fingerprint) for a snapshot file, or None."""
    try:
        st = os.stat(path)
    except OSError:
        return None
    key = (path, st.st_mtime, st.st_size)
    hit = _SNAPSHOT_DIGESTS.get(key)
    if hit is not None:
        return hit
    project = _read_json_quiet(path)
    if project is None:
        return None
    try:
        meta = json.dumps((project or {}).get("meta"), sort_keys=True,
                          ensure_ascii=False, default=str)
    except Exception:
        meta = ""
    value = (_section_digest(project), meta)
    if len(_SNAPSHOT_DIGESTS) > _SNAPSHOT_DIGESTS_MAX:
        _SNAPSHOT_DIGESTS.clear()
    _SNAPSHOT_DIGESTS[key] = value
    return value


def _snapshot_change(newer, older):
    """What one step of the timeline changed: (titles, count, cover changed).

    Titles are the sections that differ between two states, in document order,
    capped at AUTOSAVE_HINT_TITLES; the count is the whole number. Returns None
    when either state could not be read -- "not known" and "nothing changed"
    are two different answers and only one of them is safe to act on.
    """
    if not newer or not older:
        return None
    a, meta_a = newer
    b, meta_b = older
    changed = []
    for key in set(list(a.keys()) + list(b.keys())):
        one, two = a.get(key), b.get(key)
        if one and two and one[2] == two[2]:
            continue
        title, loc = (one or two)[0], (one or two)[1]
        changed.append((loc, title or loc))
    changed.sort(key=lambda item: [int(n) for n in item[0].split(".")
                                   if n.isdigit()])
    titles = [title for _loc, title in changed[:AUTOSAVE_HINT_TITLES]]
    return titles, len(changed), meta_a != meta_b


def list_autosaves(project_dir):
    """Every snapshot of one report, newest first.

    Each entry carries what it always did -- name, mtime, size and the reason
    tag -- plus, for the newest AUTOSAVE_HINT_DEPTH of them, what changed
    between it and the state before it: ``changed`` (up to three section
    titles), ``changed_count`` and ``changed_meta``. A timestamp and a byte
    count cannot tell two states apart, so finding the one from before a
    mistake meant restoring by trial and error -- and every wrong guess costs
    the work done since.

    The detail keys are absent where they are not known (too far down the list,
    an unreadable snapshot, or the oldest entry, which has nothing before it).
    """
    out = []
    for p in _autosave_paths(project_dir):
        name = os.path.basename(p)
        try:
            st = os.stat(p)
        except OSError:
            continue
        stem = name[:-5]
        reason = stem.split("__", 1)[1].split("-")[0] if "__" in stem else ""
        out.append({"name": name, "mtime": st.st_mtime,
                    "size": st.st_size, "reason": reason, "_path": p})
    out.reverse()  # newest first

    for i, entry in enumerate(out[:AUTOSAVE_HINT_DEPTH]):
        older = out[i + 1] if i + 1 < len(out) else None
        if not older:
            break
        step = _snapshot_change(_snapshot_state(entry["_path"]),
                                _snapshot_state(older["_path"]))
        if step is None:
            continue
        titles, count, meta_changed = step
        entry["changed"] = titles
        entry["changed_count"] = count
        entry["changed_meta"] = meta_changed

    for entry in out:
        entry.pop("_path", None)
    return out


def restore_autosave(project_dir, name):
    """Restore _autosave/<name> to project.json (snapshotting current first)."""
    if not name or "/" in name or "\\" in name or not name.endswith(".json"):
        raise ValueError("bad snapshot name")
    src = os.path.join(_autosave_dir(project_dir), name)
    if not os.path.isfile(src):
        raise FileNotFoundError("snapshot not found: %s" % name)
    with open(src, "rb") as fh:
        data = fh.read()
    json.loads(data.decode("utf-8"))  # validate JSON before installing
    autosave_snapshot(project_dir, reason="prerestore")  # keep the restore undoable
    pj = os.path.join(project_dir, "project.json")
    atomic_write(pj, data)
    return {"ok": True, "restored": name, "saved_at": os.path.getmtime(pj)}


# ---------------------------------------------------------------------------
# Helpers: image saving.
# ---------------------------------------------------------------------------

_NAME_RE = re.compile(r"[^A-Za-z0-9_-]+")


def _sanitize_name(name):
    name = _NAME_RE.sub("_", name).strip("_")
    return name[:64] if name else ""


def next_image_path(project_dir, section, name=None):
    """Compute images/<section>_<seq>.png (or images/<section>-<seq>_<name>.png).

    <seq> is the next free sequence for that section. Returns (rel, abs).
    """
    section = _sanitize_name(str(section)) or "0"
    name = _sanitize_name(name) if name else ""
    images_dir = os.path.join(project_dir, "images")
    os.makedirs(images_dir, exist_ok=True)

    # find existing seqs for this section in either naming form
    seqs = []
    pat = re.compile(
        r"^%s[-_](\d+)(?:_.*)?\.png$" % re.escape(section), re.IGNORECASE
    )
    for fn in os.listdir(images_dir):
        m = pat.match(fn)
        if m:
            try:
                seqs.append(int(m.group(1)))
            except ValueError:
                pass
    seq = (max(seqs) + 1) if seqs else 1

    if name:
        rel = "images/%s-%d_%s.png" % (section, seq, name)
    else:
        rel = "images/%s_%d.png" % (section, seq)
    return rel, os.path.join(project_dir, rel)


# ---------------------------------------------------------------------------
# Helpers: xlsx parsing (free-table grid + compliance data).
# ---------------------------------------------------------------------------


def parse_xlsx_grid(xlsx_bytes):
    """Parse an xlsx into a free-table grid: {rows:[[..]], merges:[..]}.

    Reads the first worksheet's used range. Cell values are stringified;
    None -> "". Merged ranges are reported as {r,c,rs,cs} (0-based, top-left).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0

    rows = []
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            row.append("" if v is None else str(v))
        rows.append(row)

    merges = []
    for rng in ws.merged_cells.ranges:
        merges.append(
            {
                "r": rng.min_row - 1,
                "c": rng.min_col - 1,
                "rs": rng.max_row - rng.min_row + 1,
                "cs": rng.max_col - rng.min_col + 1,
            }
        )
    return {"rows": rows, "merges": merges}


def _to_num_or_str(v):
    """Coerce a cell value: numbers stay numeric, blanks -> None, else str."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def parse_xlsx_compliance(xlsx_bytes):
    """Best-effort parse of an xlsx into the compliance DATA model.

    The layout mirrors the renderer's output: a yellow 3-row header band, then
    Category / Item / [Spec] / <group axis columns> / Unit. This is a seed for
    the GUI editor, not a strict importer; the GUI lets the user fix anything.

    Returns {data: {spec_name, sims, rows}}.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row))
    if not grid:
        return {"recognized": False, "data": {"spec_name": "", "sims": [], "rows": []}}

    ncols = max(len(r) for r in grid)
    for r in grid:
        r.extend([None] * (ncols - len(r)))

    # Locate header band: the first row containing an axis label cell.
    AX = {"MIN", "TYP", "MAX", "NTWC", "NT WC", "NT-WC"}
    header_axis_row = None
    for ri, r in enumerate(grid[:10]):
        if any(isinstance(c, str) and c.strip().upper() in AX for c in r):
            header_axis_row = ri
            break
    if header_axis_row is None:
        # Fall back: treat as plain grid, no compliance structure recognized.
        return {"recognized": False, "data": {"spec_name": "", "sims": [], "rows": []}}

    # Map header columns.
    axis_cols = []  # list of (col_index, axis_label)
    cat_col = item_col = spec_col = unit_col = None
    band = grid[header_axis_row]
    for ci, c in enumerate(band):
        if not isinstance(c, str):
            continue
        cu = c.strip().upper()
        if cu in AX:
            axis_cols.append((ci, "NTWC" if cu.replace(" ", "").replace("-", "") == "NTWC" else cu))

    # Category/Item/Unit/Spec live in the merged label columns above the axes:
    # scan rows 0..header_axis_row for the literal labels.
    for ri in range(0, header_axis_row + 1):
        for ci, c in enumerate(grid[ri]):
            if not isinstance(c, str):
                continue
            cl = c.strip().lower()
            if cl == "category" and cat_col is None:
                cat_col = ci
            elif cl == "item" and item_col is None:
                item_col = ci
            elif cl == "spec" and spec_col is None:
                spec_col = ci
            elif cl == "unit" and unit_col is None:
                unit_col = ci

    # Group axis columns into runs of 3-4 consecutive axis columns.
    spec_name = ""
    sims = []
    spec_axis_cols = None
    groups = []  # (title, stage, [cols])
    if axis_cols:
        # Split into groups. A new group starts when columns are non-adjacent
        # (a spacer between groups) OR when an axis label repeats within the
        # current run (e.g. ...MAX, NTWC, MIN... => the second MIN begins a new
        # group even with no spacer column).
        runs = []
        cur = [axis_cols[0]]
        seen = {axis_cols[0][1]}
        for prev, nxt in zip(axis_cols, axis_cols[1:]):
            adjacent = (nxt[0] - prev[0] == 1)
            repeats = nxt[1] in seen
            if adjacent and not repeats:
                cur.append(nxt)
                seen.add(nxt[1])
            else:
                runs.append(cur)
                cur = [nxt]
                seen = {nxt[1]}
        runs.append(cur)

        # Group titles sit in the row above header_axis_row (the merged band top).
        title_row = grid[header_axis_row - 2] if header_axis_row >= 2 else grid[0]
        stage_row = grid[header_axis_row - 1] if header_axis_row >= 1 else grid[0]
        for gi, run in enumerate(runs):
            c0 = run[0][0]
            title = _first_str(title_row, run) or ""
            stage = _first_str(stage_row, run)
            cols = [rc[0] for rc in run]
            if gi == 0 and "spec" in (title or "").lower():
                spec_name = title
                spec_axis_cols = cols
            elif gi == 0 and stage is None and "spec" in (title or "").lower():
                spec_name = title
                spec_axis_cols = cols
            else:
                groups.append((title, stage, cols))
        # If first run wasn't recognized as spec by name, assume it is spec.
        if spec_axis_cols is None and runs:
            spec_axis_cols = [rc[0] for rc in runs[0]]
            spec_name = _first_str(title_row, runs[0]) or spec_name
            groups = []
            for gi, run in enumerate(runs[1:]):
                title = _first_str(title_row, run) or "Sim%d" % (gi + 1)
                stage = _first_str(stage_row, run)
                groups.append((title, stage, [rc[0] for rc in run]))

    taken = set()
    for gi, (title, stage, cols) in enumerate(groups):
        key = _sanitize_name(title).lower() or ("sim%d" % (gi + 1))
        # Two groups whose titles sanitize to the same word would otherwise be
        # one key, and the per-group values below would collapse into it.
        base, n = key, 2
        while key in taken:
            key = "%s%d" % (base, n)
            n += 1
        taken.add(key)
        sims.append(
            {
                "key": key,
                "title": title or ("Sim%d" % (gi + 1)),
                "stage": stage,
            }
        )

    # Data rows start after the header band.
    rows = []
    last_cat = ""
    for ri in range(header_axis_row + 1, len(grid)):
        r = grid[ri]
        item = _to_num_or_str(r[item_col]) if item_col is not None else None
        cat = _to_num_or_str(r[cat_col]) if cat_col is not None else None
        if cat:
            last_cat = str(cat)
        if item is None or item == "":
            continue  # skip separator / blank rows
        unit = _to_num_or_str(r[unit_col]) if unit_col is not None else None
        spec = _to_num_or_str(r[spec_col]) if spec_col is not None else None

        spec_mtm = [None, None, None]
        spec_ntwc = None
        if spec_axis_cols:
            for ai, ci in enumerate(spec_axis_cols[:4]):
                val = _to_num_or_str(r[ci]) if ci < len(r) else None
                if ai < 3:
                    spec_mtm[ai] = val
                else:
                    spec_ntwc = val

        # EVERY declared group keeps its own values, under its own key, the way
        # the engine reads them (row["sims"][key]). The flat sim_mtm below is
        # the first group as well, kept because older callers read only that;
        # without the per-group map a two-group sheet used to come back with the
        # first group's numbers standing in for both.
        row_sims = {}
        for gi, (_title, _stage, cols) in enumerate(groups):
            mtm = [None, None, None]
            ntwc = None
            for ai, ci in enumerate(cols[:4]):
                val = _to_num_or_str(r[ci]) if ci < len(r) else None
                if ai < 3:
                    mtm[ai] = val
                else:
                    ntwc = val
            row_sims[sims[gi]["key"]] = {"mtm": mtm, "ntwc": ntwc}

        sim_mtm = [None, None, None]
        sim_ntwc = None
        if groups:
            first_cols = groups[0][2]
            for ai, ci in enumerate(first_cols[:4]):
                val = _to_num_or_str(r[ci]) if ci < len(r) else None
                if ai < 3:
                    sim_mtm[ai] = val
                else:
                    sim_ntwc = val

        row_out = {
            "cat": last_cat,
            "item": str(item),
            "unit": "" if unit is None else str(unit),
            "kind": "result",
            "spec": spec,
            "spec_mtm": spec_mtm,
            "sim_mtm": sim_mtm,
            "spec_ntwc": spec_ntwc,
            "sim_ntwc": sim_ntwc,
            "limit": None,
            "sim_span": False,
        }
        if row_sims:
            row_out["sims"] = row_sims
        rows.append(row_out)

    return {"recognized": True, "data": {"spec_name": spec_name, "sims": sims, "rows": rows}}


def _first_str(row, run):
    """Return the first non-empty string within the given run's columns."""
    for ci, _ in run:
        if ci < len(row) and isinstance(row[ci], str) and row[ci].strip():
            return row[ci].strip()
    return None


# ---------------------------------------------------------------------------
# Helpers: compliance validation (reuses engine flag logic; never duplicated).
# ---------------------------------------------------------------------------


def _import_flag_positions():
    """Lazily import the engine's flag_positions. Raises on unavailability."""
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    try:
        from tables import flag_positions  # type: ignore
        return flag_positions
    except Exception:
        # engine module may expose it directly
        from engine import flag_positions  # type: ignore
        return flag_positions


def validate_compliance(data):
    """Run engine flag logic over compliance rows.

    Returns {"flags": {"<rowIndex>": [axisIndex,...]}, "color": flag_color}.
    Axis indices are positions within the sim group's [MIN, TYP, MAX] axes.
    """
    flag_positions = _import_flag_positions()
    rows = (data or {}).get("rows", [])
    flags = {}
    for i, row in enumerate(rows):
        positions = flag_positions(row)
        if positions:
            flags[str(i)] = sorted(int(p) for p in positions)
    try:
        tc = load_template_config()
        color = tc.get("compliance", {}).get("flag_color", "FF0000")
    except Exception:
        color = "FF0000"
    return {"flags": flags, "color": color}


# ---------------------------------------------------------------------------
# Helpers: export (delegates to the engine).
# ---------------------------------------------------------------------------


def _import_engine():
    # Reload on every export so a `git pull` of engine.py/tables.py takes effect
    # without restarting this long-running server (same rationale as the
    # apply_update reload). engine imports tables at module scope, so tables must
    # be reloaded first, otherwise engine keeps binding the stale module.
    import importlib
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import content_lint  # type: ignore  # engine imports this at module scope
    importlib.reload(content_lint)
    import tables  # type: ignore
    importlib.reload(tables)
    import engine  # type: ignore
    importlib.reload(engine)
    return engine


def _import_xlsx_export():
    # Reload so a git pull of xlsx_export.py / tables.py takes effect without a
    # restart. xlsx_export imports tables at module scope -> reload tables first.
    import importlib
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import tables  # type: ignore
    importlib.reload(tables)
    import xlsx_export  # type: ignore
    return importlib.reload(xlsx_export)


# How the renderer spells a figure that carries no file at all: engine's
# _place_picture writes ``fname or "(no file)"`` into the warning's detail, so a
# blank file is the empty string or that literal, and anything else is a real
# file name.
_BLANK_FIGURE_DETAILS = ("", "(no file)")


def _level_blank_figures(warnings):
    """Drop a blank figure's ``missing_image`` to info. Returns ``warnings``.

    Two different situations arrive from the renderer under one type. A figure
    whose file is blank is the deliberate "cleared for this stage" state -- a
    freshly inherited report is full of them by design, and content_lint already
    reports each one as an info-level ``image_placeholder``. A figure that NAMES
    a file which is not in the folder is a picture somebody still has to copy in.
    Only the second needs attention, so only the second stays a warning; sixty
    notes about empty frames otherwise bury the handful that matter.
    """
    for w in warnings or []:
        if not isinstance(w, dict) or w.get("type") != "missing_image":
            continue
        if str(w.get("detail") or "").strip() in _BLANK_FIGURE_DETAILS:
            w["level"] = "info"
    return warnings


def _stale_siblings(fresh_abs, project_dir):
    """Artefacts of an EARLIER export that this one did not rewrite.

    A Word-only export overwrites ``<name>.docx`` and leaves any ``<name>.pdf``
    from a previous export sitting beside it: same folder, same name, one render
    out of date, and nothing in a directory listing to tell them apart. Handing
    that PDF on hands on a report that no longer exists.

    Deleting the file would be the tidier folder and the worse surprise -- it is
    the user's file, and it may be the copy they already sent -- so the export
    REPORTS it instead and the interface says which artefact is behind. Only a
    file genuinely older than what was just written is named.
    """
    stem = os.path.splitext(fresh_abs)[0]
    try:
        fresh_at = os.path.getmtime(fresh_abs)
    except OSError:
        return []
    found = []
    for ext, kind in ((".pdf", "pdf"),):
        sibling = stem + ext
        try:
            older_by = fresh_at - os.path.getmtime(sibling)
        except OSError:
            continue                    # not there, or unreadable: nothing to say
        if older_by <= 0:
            continue                    # as new as what we just wrote
        found.append({
            "out": os.path.relpath(sibling, project_dir).replace("\\", "/"),
            "abs": os.path.abspath(sibling).replace("\\", "/"),
            "fmt": kind,
            "older_by_s": int(round(older_by)),
        })
    return found


def run_export(project_dir, fmt, save_first=False, on_progress=None, on_phase=None):
    """Render the project via the engine. fmt in {docx, pdf}.

    Returns {out, abs, fmt, warnings, stats}. For pdf: render docx then
    COM-export (Export then Close, never a method call after Close). Output
    lands in <dir>/out/.

    ``stale_siblings`` is ADDITIVE and appears only when there is something to
    say: the artefacts of an earlier export that this one left behind -- see
    _stale_siblings.

    ``on_progress(done, total, label)`` (per heading/block) and ``on_phase(label)``
    ('preparing'|'rendering'|'converting') drive a live progress bar; both optional
    and best-effort (their exceptions never break the export).
    """
    def phase(label):
        if on_phase:
            try:
                on_phase(label)
            except Exception:
                pass

    phase("preparing")
    engine = _import_engine()
    out_dir = os.path.join(project_dir, "out")
    os.makedirs(out_dir, exist_ok=True)

    # Load the project document.
    with open(os.path.join(project_dir, "project.json"), "r", encoding="utf-8") as fh:
        project = json.load(fh)

    # Resolve and load the template config (engine resolves the logo path).
    # A project bound to a library template wins over the global --config.
    tid = project.get("template")
    tpl_cfg = tstore.template_config_path(CFG.reports_root, tid) if tid else None
    explicit = tpl_cfg or CFG.template_config_path
    config_path = engine._resolve_config_path(project, project_dir, explicit)
    cfg = engine._load_config(config_path)

    name = os.path.basename(project_dir.rstrip(os.sep)) or "report"
    docx_out = os.path.join(out_dir, "%s.docx" % name)
    # render_report returns {out_path, warnings, stats}; tolerate the legacy
    # bare-string return too. The render manifest (warnings + stats) is surfaced
    # to the frontend as additive keys on the success payload so the user can see
    # missing images / out-of-spec / clip risks after an export.
    phase("rendering")
    render_result = engine.render_report(project, cfg, project_dir, docx_out,
                                         on_progress=on_progress)
    out_path = engine._result_out_path(render_result) or docx_out
    docx_abs = os.path.abspath(out_path)
    warnings = render_result.get("warnings", []) if isinstance(render_result, dict) else []
    stats = render_result.get("stats", {}) if isinstance(render_result, dict) else {}

    # Merge pre-render content-lint findings into the SAME manifest so the export
    # panel surfaces structural issues (missing condition rows, never-flagging
    # rows, sim_span axis problems, bad image paths, ...) next to the render
    # warnings -- all carrying a level. no_caption is split (engine owns
    # datatable/table, lint owns image/imagegrid) so there is no double-report.
    try:
        import content_lint  # already reloaded by _import_engine
        lint_findings = content_lint.lint_project(project, cfg)
    except Exception:
        lint_findings = []
    if lint_findings:
        warnings = list(lint_findings) + list(warnings)
    _level_blank_figures(warnings)
    stats = dict(stats)
    stats["errors"] = sum(1 for w in warnings if w.get("level") == "error")
    stats["warns"] = sum(1 for w in warnings if w.get("level") == "warn")
    stats["infos"] = sum(1 for w in warnings if w.get("level") == "info")

    if fmt == "docx":
        rel = os.path.relpath(docx_abs, project_dir).replace("\\", "/")
        answer = {"out": rel, "abs": docx_abs.replace("\\", "/"), "fmt": "docx",
                  "warnings": warnings, "stats": stats}
        # Only a Word-only export can leave an artefact behind: the PDF export
        # below rewrites both files.
        older = _stale_siblings(docx_abs, project_dir)
        if older:
            answer["stale_siblings"] = older
        return answer

    if fmt == "pdf":
        pdf_abs = os.path.splitext(docx_abs)[0] + ".pdf"
        phase("converting")
        _word_export_pdf(docx_abs, pdf_abs)
        rel = os.path.relpath(pdf_abs, project_dir).replace("\\", "/")
        return {"out": rel, "abs": pdf_abs.replace("\\", "/"), "fmt": "pdf",
                "warnings": warnings, "stats": stats}

    raise ValueError("unknown fmt: %s" % fmt)


def _import_live_preview():
    """The live-preview module, or None when it cannot be imported.

    Its COM / process-identity helpers are reused by the PDF export so there is
    exactly ONE set of rules in this codebase for driving and disposing of a Word
    process. Returning None (rather than raising) keeps the export working on a
    machine where the preview module is unavailable: identification is then
    skipped, which only ever means declining to terminate anything.
    """
    import importlib
    for name in ("web.live_preview", "live_preview"):
        try:
            return importlib.import_module(name)
        except ImportError:
            continue
    return None


def _word_instance_identity(lp, word, spawned_after):
    """``(pid, created)`` of the instance just created, or ``(None, None)``.

    Every check here is live_preview's, for live_preview's reason: a wrong id
    would not mean a broken export, it would mean terminating a process this
    server never started -- and the likeliest such process is the copy of Word
    the user has their own unsaved documents open in. Anything uncertain refuses
    and returns nothing, which leaves the instance to its own ``Quit``.
    """
    if lp is None:
        return None, None
    try:
        pid = lp._identify_instance(word)
        created = lp._process_created(pid) if pid else None
        if pid is None or created is None:
            return None, None
        if created < spawned_after - lp._SPAWN_CLOCK_SLACK:
            return None, None       # older than our own start: cannot be ours
        image = os.path.basename(lp._process_image(pid) or "")
        if image.upper() != lp._WORD_IMAGE_NAME:
            return None, None
        return pid, created
    except Exception:
        return None, None


def _word_export_pdf(docx_abs, pdf_abs):
    """Word COM: Open -> ExportAsFixedFormat -> Close. Export then Close.

    COM apartment state is PER THREAD, and this runs on whichever request thread
    ThreadingHTTPServer handed the export to. pywin32 initialises the apartment
    of the thread that first imports pythoncom and of no other, so the first
    export of a session used to work and every later one -- served by a different
    thread -- died on "CoInitialize has not been called", with only a restart
    clearing it. The apartment is entered here and left again on the way out, the
    same way live_preview's COM worker thread brackets its own loop.

    Process safety follows live_preview's rule exactly: ``DispatchEx`` creates
    OUR OWN Word process (``Dispatch`` / ``GetActiveObject`` would attach to the
    copy of Word the user is typing in, and quitting that would throw away their
    unsaved work), and the only process this function may ever terminate is the
    one it just created, identified by pid AND creation time -- and only after a
    graceful ``Quit`` has been given its grace period and did not take.
    """
    import pythoncom       # type: ignore  # pywin32
    import win32com.client  # type: ignore  # pywin32, present on the work machine

    lp = _import_live_preview()
    # 17 = wdExportFormatPDF; the same constant live_preview exports with.
    pdf_format = getattr(lp, "_WD_EXPORT_FORMAT_PDF", 17)
    quit_grace = getattr(lp, "QUIT_GRACE", 8.0)

    try:
        pythoncom.CoInitialize()
        entered = True
    except Exception:
        # Somebody already put this thread in an apartment; leave it as we
        # found it rather than unbalancing their CoUninitialize.
        entered = False
    try:
        # Read before the process exists: nothing this call creates can be older.
        spawned_after = time.time()
        word = win32com.client.DispatchEx("Word.Application")
        pid = created = None
        doc = None
        try:
            word.Visible = False
            pid, created = _word_instance_identity(lp, word, spawned_after)
            doc = word.Documents.Open(os.path.abspath(docx_abs))
            # Update all fields (e.g. a Table of Contents) so the PDF reflects
            # what the user sees in Word after a field refresh; otherwise
            # TOC/PAGEREF fields render as their unpopulated placeholder and
            # pagination differs.
            try:
                for story in doc.StoryRanges:
                    story.Fields.Update()
                for toc in doc.TablesOfContents:
                    toc.Update()
            except Exception:
                pass  # field update is best-effort; never block the export
            doc.ExportAsFixedFormat(os.path.abspath(pdf_abs), pdf_format)
        finally:
            if doc is not None:
                try:
                    doc.Close(False)  # never call a method on doc after this
                except Exception:
                    pass
            try:
                word.Quit()
            except Exception:
                pass
            del word
            # A Quit that did not take (a modal dialog, say) used to leak the
            # process for the life of the machine. Terminate it -- but only it:
            # _process_matches inside _terminate_recorded re-checks the creation
            # time, so a pid Windows has since recycled is left alone.
            if pid is not None and lp is not None:
                if not lp._wait_for_exit(pid, created, quit_grace):
                    lp._terminate_recorded(pid, created)
    finally:
        if entered:
            pythoncom.CoUninitialize()


# ---------------------------------------------------------------------------
# Helpers: docx import (templates + report seeds).
# ---------------------------------------------------------------------------


def _import_docx_module():
    """Lazily import the docx_import parser (mirrors _import_engine).

    Raises ImportError/ModuleNotFoundError if python-docx is unavailable so the
    route can degrade to a clean 503 rather than failing server import.
    """
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import docx_import  # type: ignore
    return docx_import


@contextlib.contextmanager
def _temp_docx(docx_bytes):
    """Materialize uploaded bytes to a temp .docx (docx_import takes a path).

    The temp file is always removed on exit; any images / logo extracted out of
    it are written to their destination dirs before the context closes.
    """
    fd, path = tempfile.mkstemp(suffix=".docx", prefix="import_docx_")
    try:
        with os.fdopen(fd, "wb") as fh:
            fh.write(docx_bytes)
        yield path
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


def _rewrite_image_paths(outline):
    """Guarantee every image block's "file" is project-relative "images/<base>".

    Keeps "src". Idempotent for already-correct parser output; self-heals if a
    future parser change emits an absolute or bare-basename path.
    """
    def walk(nodes):
        for n in nodes:
            for b in n.get("blocks", []):
                if b.get("type") == "image":
                    src = b.get("file") or b.get("src") or ""
                    base = os.path.basename(src.replace("\\", "/"))
                    if base:
                        b["file"] = "images/" + base
                        b.setdefault("src", b["file"])
            walk(n.get("children", []))

    walk(outline or [])
    return outline


def _instantiate_skeleton(skeleton):
    """Clone a skeleton tree into a project outline.

    FIXED nodes carry their blocks verbatim; FILLABLE nodes get an empty blocks
    list (a clean placeholder for the editor). The ``fixed`` marker is dropped --
    a project node is just ``{title, level, blocks, children}``.
    """
    out = []
    for node in skeleton or []:
        is_fixed = bool(node.get("fixed"))
        out.append({
            "title": node.get("title", ""),
            "level": node.get("level", 1),
            "blocks": copy.deepcopy(node.get("blocks", [])) if is_fixed else [],
            "children": _instantiate_skeleton(node.get("children", [])),
        })
    return out


def _empty_meta():
    """Default empty project meta (mirrors parse_docx_report's meta dict)."""
    return {
        "title": "", "secrecy": "", "doc_no": "", "page_count": "",
        "author": "", "reviewers": [], "approver": "", "revisions": [],
    }


# ---------------------------------------------------------------------------
# Helpers: paste-import structural diff (de-LLM upstream channel, WS2).
# ---------------------------------------------------------------------------
#
# The GUI's Copy-text exports the WHOLE App.project. Re-importing it is a full
# replace, so there is no three-way merge: "user version wins" == overwrite +
# backup + a compact diff the assistant can read (structure, not full text).


def _short(v, n=60):
    s = "" if v is None else str(v)
    s = s.replace("\n", " ")
    return s if len(s) <= n else s[:n] + "..."


def _collect_nodes(outline):
    """Every outline node in document order (depth-first, children included)."""
    out = []

    def walk(nodes):
        for n in nodes or []:
            if isinstance(n, dict):
                out.append(n)
                walk(n.get("children", []))

    walk(outline)
    return out


def _block_types(node):
    from collections import Counter
    c = Counter()
    for b in (node.get("blocks") or []):
        if isinstance(b, dict):
            c[b.get("type", "?")] += 1
    return c


def _para_chars(node):
    total = 0
    for b in (node.get("blocks") or []):
        if isinstance(b, dict) and b.get("type") == "para":
            for r in (b.get("runs") or []):
                if isinstance(r, dict):
                    total += len(r.get("t", "") or "")
    return total


def _fmt_multiset_delta(old_c, new_c):
    parts = []
    for k in sorted(set(old_c) | set(new_c)):
        o, n = old_c.get(k, 0), new_c.get(k, 0)
        if o != n:
            parts.append("%s %d->%d" % (k, o, n))
    return ", ".join(parts)


def _match_nodes(old_nodes, new_nodes):
    """Pair old/new outline nodes by title (id as tie-break, robust to node-id
    drift). Returns (pairs, added, removed)."""
    from collections import defaultdict
    by_title = defaultdict(list)
    for n in old_nodes:
        by_title[n.get("title", "")].append(n)
    consumed, pairs, added = set(), [], []
    for nn in new_nodes:
        cands = by_title.get(nn.get("title", ""), [])
        pick = None
        for c in cands:  # prefer same id, not yet consumed
            if id(c) not in consumed and nn.get("id") and c.get("id") == nn.get("id"):
                pick = c
                break
        if pick is None:
            for c in cands:
                if id(c) not in consumed:
                    pick = c
                    break
        if pick is None:
            added.append(nn)
        else:
            consumed.add(id(pick))
            pairs.append((pick, nn))
    removed = [n for n in old_nodes if id(n) not in consumed]
    return pairs, added, removed


def paste_import_diff(old_project, new_project, dir_name, id_warns=None):
    """Compact structural diff between the on-disk project and the pasted one.

    No full text: meta key changes, top-level key changes (e.g. sim_checklist),
    and per-section block-type multiset + para char-length deltas. Returns a
    markdown string (~300-600 tokens) the assistant reads instead of the file."""
    old = old_project if isinstance(old_project, dict) else {}
    new = new_project if isinstance(new_project, dict) else {}
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = ["# Paste-import diff -- %s" % dir_name, "",
             "_Generated %s. Structural summary only (no full text)._" % ts, ""]
    for w in (id_warns or []):
        lines.append("> WARNING: %s" % w)
    if id_warns:
        lines.append("")

    # meta
    om, nm = old.get("meta") or {}, new.get("meta") or {}
    meta_changes = []
    for k in sorted(set(om) | set(nm)):
        ov, nv = om.get(k), nm.get(k)
        if ov == nv:
            continue
        if isinstance(ov, list) or isinstance(nv, list):
            meta_changes.append("%s %d->%d entries" % (k, len(ov or []), len(nv or [])))
        else:
            meta_changes.append("%s %r->%r" % (k, _short(ov), _short(nv)))
    lines += ["## meta",
              ("- " + "; ".join(meta_changes)) if meta_changes else "- (unchanged)", ""]

    # top-level keys besides meta/outline (template + sim_checklist etc.)
    skip = {"meta", "outline", "schema_version"}
    top_changes = []
    for k in sorted(set(old) | set(new)):
        if k in skip:
            continue
        ov, nv = old.get(k), new.get(k)
        if ov == nv:
            continue
        if isinstance(ov, list) or isinstance(nv, list):
            top_changes.append("%s %d->%d items" % (k, len(ov or []), len(nv or [])))
        else:
            top_changes.append("%s %r->%r" % (k, _short(ov), _short(nv)))
    if top_changes:
        lines += ["## top-level", "- " + "; ".join(top_changes), ""]

    # outline sections
    old_nodes = _collect_nodes(old.get("outline"))
    new_nodes = _collect_nodes(new.get("outline"))
    pairs, added, removed = _match_nodes(old_nodes, new_nodes)
    changed = []
    for on, nn in pairs:
        delta = _fmt_multiset_delta(_block_types(on), _block_types(nn))
        opl, npl = _para_chars(on), _para_chars(nn)
        if delta or opl != npl:
            desc = []
            if delta:
                desc.append("blocks " + delta)
            if opl != npl:
                desc.append("para chars %d->%d (%+d)" % (opl, npl, npl - opl))
            changed.append((nn.get("title", ""), "; ".join(desc)))
    lines += ["## sections",
              "- %d total (was %d); %d changed, %d added, %d removed"
              % (len(new_nodes), len(old_nodes), len(changed), len(added), len(removed)), ""]
    if changed:
        lines.append("### changed")
        lines += ['- "%s": %s' % (t, d) for t, d in changed]
        lines.append("")
    if added:
        lines.append("### added")
        lines += ['- "%s" [%s]' % (n.get("title", ""),
                                   _fmt_multiset_delta({}, _block_types(n)) or "empty")
                  for n in added]
        lines.append("")
    if removed:
        lines.append("### removed")
        lines += ['- "%s"' % n.get("title", "") for n in removed]
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


# ---------------------------------------------------------------------------
# HTTP handler.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Helpers: the report tree (PROJECT / MODULE / STAGE).
#
# Reports used to live flat under the reports root: <root>/<MODULE>/project.json.
# The layout the workbench addresses is <root>/<PROJECT>/<MODULE>/<STAGE>/
# project.json -- two extra levels. Both are read by ONE walk, which classifies
# whatever it finds by its depth:
#   depth 1 -> a legacy flat report: project "unfiled", module = folder name,
#              stage read out of meta.title when the title names one
#   depth 2 -> a report parked directly in a module folder: stage from the title
#   depth 3 -> project / module / stage taken straight from the path
# There is no new addressing scheme: a report is identified by its path relative
# to the reports root ("PROJ/MOD/STAGE"), which resolve_project_dir already
# accepts and already guards for containment.
# ---------------------------------------------------------------------------

UNFILED_PROJECT_ID = "unfiled"
UNFILED_PROJECT_NAME = "Unfiled"       # user-facing: 02_GLOSSARY_EN.md, Home

# Optional, at the project and module level: {"name": ..., "description": ...}.
PROJECT_META_FILE = "project_meta.json"

# Bookkeeping and payload folders that can never be a project, module or stage.
_TREE_SKIP_DIRS = set(_RESERVED_ROOT_DIRS) | {
    "_trash", "_autosave", "_backups", "_updates",
    "images", "out", "data", "templates",
}

_STAGE_RE = re.compile(r"(?:^|[^A-Z])(XDR|PDR|CDR|FDR)(?:[^A-Z]|$)")


def _read_json_quiet(path):
    """Parse a JSON file, or None. Never raises: the tree walk must not die on
    one unreadable report."""
    try:
        return _load_json_file(path)
    except Exception:
        return None


def _is_tree_dir(parent, name):
    if name in _TREE_SKIP_DIRS or name.startswith("_") or name.startswith("."):
        return False
    return os.path.isdir(os.path.join(parent, name))


def _listdir_quiet(path):
    try:
        return sorted(os.listdir(path))
    except OSError:
        return []


def _has_project(path):
    return os.path.isfile(os.path.join(path, "project.json"))


def _stage_from_text(text):
    """The stage code named by a title or a folder name, else an empty string."""
    m = _STAGE_RE.search(str(text or "").upper())
    return m.group(1) if m else ""


def _dir_meta(path):
    """(name, description) from an optional project_meta.json, else empties."""
    if not path:
        return "", ""
    d = _read_json_quiet(os.path.join(path, PROJECT_META_FILE))
    if not isinstance(d, dict):
        return "", ""
    return (str(d.get("name") or "").strip(),
            str(d.get("description") or "").strip())


# Over-spec counts are cached against project.json's mtime: the shelf asks for
# every report at once, and re-flagging thousands of rows on each poll is waste.
_OVERSPEC_CACHE = {}                    # normcased abs path -> (mtime, count)
_OVERSPEC_LOCK = threading.Lock()


def _import_tables():
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import tables  # type: ignore
    return tables


def count_over_spec(project):
    """Number of simulated values that violate their row's limit.

    Uses the SAME code the checklist and the rendered table use -- tables'
    ``_flags_from`` over each simulation group's own values, which is what
    ``flag_positions`` and ``render_datatable`` are both built on -- so the shelf
    and the checklist can never disagree. Rows that declare no per-group values
    fall back to the flat-schema ``flag_positions``.
    """
    try:
        tb = _import_tables()
    except Exception:
        return 0
    total = 0
    for node in _collect_nodes((project or {}).get("outline") or []):
        for block in node.get("blocks") or []:
            if not isinstance(block, dict) or block.get("type") != "datatable":
                continue
            data = block.get("data") or {}
            keys = [g.get("key") for g in (data.get("sims") or [])
                    if isinstance(g, dict) and g.get("key")]
            for row in data.get("rows") or []:
                if not isinstance(row, dict):
                    continue
                try:
                    if keys and hasattr(tb, "_sim_axis_vals"):
                        for k in keys:
                            mtm, ntwc = tb._sim_axis_vals(row, k)
                            total += len(tb._flags_from(row, mtm, ntwc))
                    else:
                        total += len(tb.flag_positions(row))
                except Exception:
                    pass        # one malformed row never hides the whole count
    return total


def over_spec_count(project_dir, project=None):
    """Cached count_over_spec for a project dir, keyed on project.json's mtime."""
    pj = os.path.join(project_dir, "project.json")
    try:
        mtime = os.path.getmtime(pj)
    except OSError:
        return 0
    key = os.path.normcase(os.path.abspath(pj))
    with _OVERSPEC_LOCK:
        hit = _OVERSPEC_CACHE.get(key)
    if hit and hit[0] == mtime:
        return hit[1]
    if project is None:
        project = _read_json_quiet(pj)
    n = count_over_spec(project or {})
    with _OVERSPEC_LOCK:
        _OVERSPEC_CACHE[key] = (mtime, n)
    return n


def _json_read_error(path):
    """Why a JSON file does not parse, in the parser's words, or "" when it does
    (or does not exist). Only called for a file _read_json_quiet gave up on, so
    the second parse is of a broken file and fails fast."""
    try:
        with open(path, "r", encoding="utf-8") as fh:
            json.load(fh)
        return ""
    except FileNotFoundError:
        return ""
    except Exception as exc:
        return str(exc)


def _report_entry(root, parts, stage=""):
    """One report row of the tree. ``parts`` is its path under the reports root.

    ``unreadable`` is true when project.json is there but will not parse, with
    ``error`` carrying the parser's reason -- so the shelf can say so before
    the report is opened. Both keys are additive; every field that was here
    before is unchanged.
    """
    project_dir = os.path.join(root, *parts)
    pj = os.path.join(project_dir, "project.json")
    parsed = _read_json_quiet(pj)
    unreadable = parsed is None and os.path.isfile(pj)
    project = parsed if isinstance(parsed, dict) else {}
    title = str((project.get("meta") or {}).get("title") or "")
    stage = stage or _stage_from_text(title)
    try:
        mtime = os.path.getmtime(pj)
    except OSError:
        mtime = 0.0
    baseline = os.path.join(project_dir, "_baseline.json")
    last = os.path.getmtime(baseline) if os.path.isfile(baseline) else None
    return {
        "dir": "/".join(parts),
        "stage": stage,
        "name": stage or parts[-1],
        "title": title,
        "mtime": mtime,
        "unreadable": unreadable,
        "error": _json_read_error(pj) if unreadable else "",
        "overSpec": over_spec_count(project_dir, project),
        # exchange.sectionsSince is a documented STUB, always 0: counting the
        # sections changed since the last exchange needs the baseline op-diff,
        # which the exchange view computes on demand via /api/copy-diff.
        "exchange": {"last": last, "sectionsSince": 0},
    }


def _project_row(project_dir, parts):
    """One /api/projects entry -- the flat listing the previous interface renders.

    Same four keys it has always had; only "dir" grew, from a folder name to the
    report's path under the reports root. Never raises: one unreadable report
    must not empty the whole list.
    """
    pj = os.path.join(project_dir, "project.json")
    project = _read_json_quiet(pj) or {}
    try:
        mtime = os.path.getmtime(pj)
    except OSError:
        mtime = 0.0
    return {
        "dir": "/".join(parts),
        "title": str((project.get("meta") or {}).get("title") or ""),
        "template": str(project.get("template") or ""),
        "mtime": mtime,
    }


def build_report_tree(root):
    """Walk the reports root (three levels deep) into the /api/tree shape."""
    if not root or not os.path.isdir(root):
        return []

    projects = {}          # id -> {..., "modules": {id -> module}}
    project_order = []

    def project_bucket(pid, path):
        if pid not in projects:
            name, desc = _dir_meta(path)
            if not name:
                name = UNFILED_PROJECT_NAME if pid == UNFILED_PROJECT_ID else pid
            projects[pid] = {"id": pid, "name": name, "description": desc,
                             "modules": {}, "_order": []}
            project_order.append(pid)
        return projects[pid]

    def module_bucket(proj, mid, path):
        if mid not in proj["modules"]:
            name, desc = _dir_meta(path)
            proj["modules"][mid] = {"id": mid, "name": name or mid,
                                    "description": desc, "reports": []}
            proj["_order"].append(mid)
        return proj["modules"][mid]

    for l1 in _listdir_quiet(root):
        if not _is_tree_dir(root, l1):
            continue
        p1 = os.path.join(root, l1)
        if _has_project(p1):
            # depth 1: a flat legacy report, filed under the "unfiled" project.
            proj = project_bucket(UNFILED_PROJECT_ID, None)
            module_bucket(proj, l1, p1)["reports"].append(
                _report_entry(root, [l1]))
            continue
        has_meta_1 = os.path.isfile(os.path.join(p1, PROJECT_META_FILE))
        found_1 = False
        for l2 in _listdir_quiet(p1):
            if not _is_tree_dir(p1, l2):
                continue
            p2 = os.path.join(p1, l2)
            if _has_project(p2):
                # depth 2: a report sitting directly in a module folder.
                proj = project_bucket(l1, p1)
                module_bucket(proj, l2, p2)["reports"].append(
                    _report_entry(root, [l1, l2]))
                found_1 = True
                continue
            found_2 = False
            for l3 in _listdir_quiet(p2):
                if not _is_tree_dir(p2, l3):
                    continue
                if not _has_project(os.path.join(p2, l3)):
                    continue
                # depth 3: the target layout.
                proj = project_bucket(l1, p1)
                module_bucket(proj, l2, p2)["reports"].append(
                    _report_entry(root, [l1, l2, l3],
                                  stage=_stage_from_text(l3) or l3))
                found_1 = found_2 = True
            if not found_2 and os.path.isfile(os.path.join(p2, PROJECT_META_FILE)):
                # a module the user just created, still empty, belongs on the shelf
                module_bucket(project_bucket(l1, p1), l2, p2)
                found_1 = True
        if has_meta_1 and not found_1:
            project_bucket(l1, p1)

    out = []
    for pid in project_order:
        p = projects[pid]
        mods = []
        for mid in p["_order"]:
            m = p["modules"][mid]
            m["reports"].sort(key=lambda r: r["mtime"], reverse=True)
            mods.append(m)
        mods.sort(key=lambda m: m["name"].lower())
        out.append({"id": p["id"], "name": p["name"],
                    "description": p["description"], "modules": mods})
    # named projects first, the legacy "unfiled" bucket last
    out.sort(key=lambda p: (p["id"] == UNFILED_PROJECT_ID, p["name"].lower()))
    return out


# ---------------------------------------------------------------------------
# Helpers: the trash.
#
# A delete never unlinks anything: the folder is MOVED to <root>/_trash/ under a
# timestamped name. That has always been true, but nothing could read it back,
# so a delete was a one-way door behind a confirmation promising it was not.
#
# Putting a folder back needs one fact the timestamped name does not carry: the
# path it came FROM. A report at PROJ/MOD/STAGE trashed as "STAGE-20260830-..."
# is otherwise indistinguishable from a flat report of the same name, and
# guessing wrong files it somewhere it never lived. So the delete writes a
# sidecar next to the folder -- "<entry>.trashinfo.json", {from, deleted} --
# and the restore reads it. The sidecar sits BESIDE the folder rather than
# inside it, so what was trashed comes back byte for byte.
#
# Entries trashed before the sidecar existed are still listed and still
# restorable: their origin is derived from the entry name with its timestamp
# suffix removed, which lands them at the reports root, and the listing says so
# rather than pretending to know more than it does.
# ---------------------------------------------------------------------------

TRASH_DIRNAME = "_trash"
TRASH_INFO_SUFFIX = ".trashinfo.json"
_TRASH_STAMP_RE = re.compile(r"-\d{8}-\d{6}(?:-\d+)?$")


def _trash_root(root):
    return os.path.join(root, TRASH_DIRNAME)


def _trash_entry_id(name):
    """Validate one trash entry id -- a single folder name under _trash.

    Rejected before any path is built: separators (either platform's), the two
    relative names, a Windows drive prefix, and the empty string. Containment is
    still enforced afterwards by resolve_project_dir; this is the cheap guard
    that keeps a crafted id from ever reaching it.
    """
    seg = str(name or "").strip()
    if not seg or seg in (".", "..") or seg != os.path.basename(seg):
        raise ValueError("invalid trash entry")
    if "/" in seg or "\\" in seg or ":" in seg:
        raise ValueError("invalid trash entry")
    return seg


def _origin_from_name(entry_id):
    """Where an entry with no sidecar came from: its name without the stamp."""
    base = _TRASH_STAMP_RE.sub("", entry_id)
    return base or entry_id


def _safe_origin(origin):
    """A restore target as a list of path segments, or ValueError.

    The origin is read off disk, so it is treated as input: every segment must
    be an ordinary folder name. That keeps a restore from writing into the
    bookkeeping folders (_trash, _backups, ...) even if the sidecar is edited by
    hand, and resolve_project_dir then enforces the reports root itself.
    """
    parts = [p for p in str(origin or "").replace("\\", "/").split("/") if p]
    if not parts:
        raise ValueError("the trash entry does not name where it came from")
    for p in parts:
        if p in (".", "..") or p.startswith("_") or p in _RESERVED_ROOT_DIRS:
            raise ValueError("invalid restore path: %r" % origin)
        if ":" in p:
            raise ValueError("invalid restore path: %r" % origin)
    return parts


def _count_reports_under(path, depth=3):
    """How many reports a trashed folder holds (itself included)."""
    if _has_project(path):
        return 1
    if depth <= 0:
        return 0
    total = 0
    for name in _listdir_quiet(path):
        child = os.path.join(path, name)
        if os.path.isdir(child) and not name.startswith("_"):
            total += _count_reports_under(child, depth - 1)
    return total


def _trash_entry(root, entry_id):
    """One row of the trash listing, or None when the entry is not a folder."""
    folder = os.path.join(_trash_root(root), entry_id)
    if not os.path.isdir(folder):
        return None
    info = _read_json_quiet(folder + TRASH_INFO_SUFFIX)
    info = info if isinstance(info, dict) else {}
    origin = str(info.get("from") or "").strip()
    remembered = bool(origin)
    if not origin:
        origin = _origin_from_name(entry_id)
    try:
        deleted = float(info.get("deleted") or os.path.getmtime(folder))
    except (OSError, TypeError, ValueError):
        deleted = 0.0
    project = _read_json_quiet(os.path.join(folder, "project.json")) or {}
    title = str((project.get("meta") or {}).get("title") or "")
    if not title:
        title, _desc = _dir_meta(folder)
    occupied = False
    try:
        occupied = os.path.exists(resolve_project_dir(origin))
    except ValueError:
        occupied = True          # unrestorable path: not a free target either
    return {
        "id": entry_id,
        "name": title or origin.split("/")[-1] or entry_id,
        "dir": origin,
        "knownOrigin": remembered,
        "deleted": deleted,
        "reports": _count_reports_under(folder),
        "occupied": occupied,
    }


def list_trash(root):
    """Every restorable entry in the trash, newest deletion first."""
    if not root:
        return []
    trash = _trash_root(root)
    if not os.path.isdir(trash):
        return []
    rows = []
    for name in _listdir_quiet(trash):
        try:
            entry = _trash_entry(root, _trash_entry_id(name))
        except ValueError:
            entry = None         # a name this server would never have written
        if entry:
            rows.append(entry)
    rows.sort(key=lambda r: r["deleted"], reverse=True)
    return rows


# ---------------------------------------------------------------------------
# Helpers: v2 endpoints whose implementation module is not written yet.
#
# Every v2 endpoint below is wired to a real module function with a fixed
# signature. The import happens INSIDE the handler, so a module that does not
# exist yet can never stop the server from starting. While a module is missing
# the endpoint answers with the documented response shape, empty data and
# {"stub": true}: the frontend is written against these endpoints today and
# needs them to answer. Flip STUB_HTTP_STATUS to 501 once every module below is
# real and a missing one should be a hard error instead. A module that DOES
# exist but fails to import, or lacks the function, is always a clear 501 --
# that is a broken build, not a missing feature.
# ---------------------------------------------------------------------------

STUB_HTTP_STATUS = 200


def _lazy_call(modnames, funcname, args, empty):
    """Import the first available module of ``modnames`` and call ``funcname``.

    Returns (payload, status). Layer directories are on sys.path (buildpath), so
    both the layered name ("core.report_new") and the flat one ("report_new")
    are tried before the module is declared missing.
    """
    import importlib
    mod = None
    for name in modnames:
        try:
            mod = importlib.import_module(name)
            break
        except ModuleNotFoundError:
            continue
        except ImportError as exc:
            return {"error": "%s could not be imported: %s" % (name, exc)}, 501
    if mod is None:
        body = dict(empty)
        body["stub"] = True
        body["detail"] = "%s is not available yet" % modnames[0]
        return body, STUB_HTTP_STATUS
    fn = getattr(mod, funcname, None)
    if not callable(fn):
        return {"error": "%s has no %s()" % (mod.__name__, funcname)}, 501
    result = fn(*args)
    if not isinstance(result, dict):
        result = {"ok": True, "result": result}
    return result, 200


def _project_config(project_dir, project):
    """Resolve a project's template config exactly like the export path does (a
    library template named by the project wins over the global --config)."""
    engine = _import_engine()
    tid = (project or {}).get("template")
    tpl_cfg = tstore.template_config_path(CFG.reports_root, tid) if tid else None
    explicit = tpl_cfg or CFG.template_config_path
    return engine._load_config(
        engine._resolve_config_path(project, project_dir, explicit))


def _baseline_verdict(apply_update, project_dir, package_base, dir_name=None):
    """Ask the shared library whether a package may be applied to one report.

    There is exactly ONE implementation of the baseline rule, and it lives in
    apply_update.check_baseline: it compares the content of this report's
    ``_baseline.json`` -- the state both sides last agreed on -- against the
    ``base_sha`` the package declares it was cut from. The HTTP layer must never
    fingerprint the CURRENT project.json instead: local drift since the last
    exchange is expected and is exactly what an incoming op-diff merges onto, so
    measuring against the live file refuses every report that has been edited
    since it was last exchanged, which in practice is all of them.

    Returns (verdict, refusal_body). verdict is check_baseline's own answer --
    "ok" / "no_baseline" / "no_fingerprint" -- and refusal_body is None unless
    the two sides declare different ancestors, in which case it is the 409 body
    carrying both shas.
    """
    rel = os.path.join(os.path.relpath(project_dir, CFG.reports_root),
                       "project.json")
    try:
        return apply_update.check_baseline(CFG.reports_root, rel, package_base,
                                           dir_name), None
    except apply_update.BaselineMismatch as exc:
        return "mismatch", exc.as_dict()


def _truncation_refusal(payload):
    """409 body for a returned report that looks cut off.

    The merge refuses rather than reading every absent section as a deletion.
    The body has to carry enough for the UI to say WHAT looks missing and to
    offer the override, so it always ends up as
    {"error": "truncated", "truncated": {kind, missing, total, detail}} whatever
    shape the merge library reports it in.
    """
    t = payload.get("truncated") or {}
    total = t.get("total")
    if total is None:
        total = t.get("base_sections")
    return {"error": "truncated",
            "truncated": {"kind": t.get("kind") or "sections",
                          "missing": t.get("missing"),
                          "total": total,
                          "detail": t.get("detail") or payload.get("error") or
                          "the returned report is missing much of what it was "
                          "cut from",
                          "override": "allowBulkDelete"}}


def _open_with_os(path):
    """Hand a path to the desktop's own handler. Returns True on success."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            import subprocess
            subprocess.Popen(["open", path])
        else:
            import subprocess
            subprocess.Popen(["xdg-open", path])
        return True
    except Exception:
        return False


class Handler(BaseHTTPRequestHandler):
    server_version = "DocBuilder/1.0"
    protocol_version = "HTTP/1.1"

    # --- low-level response helpers ---

    def _send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_error_json(self, msg, status=400):
        self._send_json({"error": str(msg)}, status=status)

    def _send_bytes(self, body, content_type, status=200):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    # The body of the request being served, once something has asked for it.
    # None means "not read yet"; see _read_body.
    _body = None

    def _read_body(self):
        """The request body -- read from the socket once, then remembered.

        Reading it twice used to mean reading PAST it, into the first line of
        the next request on a kept-alive connection. Caching is what lets the
        drain guard in do_POST ask for the body unconditionally without having
        to know whether the handler already took it."""
        if self._body is None:
            length = int(self.headers.get("Content-Length") or 0)
            self._body = self.rfile.read(length) if length > 0 else b""
        return self._body

    def _drain_body(self):
        """Take the body off the socket if nothing else did. Never raises: it
        runs on the way out of a request that may already have failed."""
        try:
            self._read_body()
        except Exception:
            pass

    def _read_json(self):
        raw = self._read_body()
        if not raw:
            return {}
        return json.loads(raw.decode("utf-8"))

    def handle_one_request(self):
        # One cached body per REQUEST. The handler instance is per CONNECTION,
        # so without this reset the second request on a kept-alive socket would
        # be handed the first request's body.
        self._body = None
        BaseHTTPRequestHandler.handle_one_request(self)

    def log_message(self, fmt, *args):  # quieter logging
        sys.stderr.write("[server] %s - %s\n" % (self.address_string(), fmt % args))

    # --- routing ---

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)

            # The workbench (v2) is the default UI. The old single-file app
            # stays reachable, byte-identical, at /legacy and at /app.html.
            if path in ("/", "/index.html", "/v2", "/v2/"):
                return self._serve_v2_index()
            if path == "/legacy" or path == "/app.html":
                return self._serve_app_html()
            if path.startswith("/assets/"):
                return self._serve_asset(path)
            if path.startswith("/images/") or path.startswith("images/"):
                return self._serve_project_file(path, qs)
            if path == "/api/config":
                template = (qs.get("template") or [None])[0]
                return self._send_json(config_slice(template))
            if path == "/api/project":
                return self._api_project_get(qs)
            if path == "/api/projects":
                return self._api_projects_list()
            if path == "/api/autosaves":
                return self._api_autosaves(qs)
            if path == "/api/templates":
                return self._api_templates_list()
            if path == "/api/template":
                return self._api_template_get(qs)
            if path == "/api/tree":
                return self._api_tree()
            if path == "/api/trash":
                return self._api_trash()
            if path == "/api/assets":
                return self._api_assets(qs)
            if path == "/api/version":
                return self._api_version()
            if path == "/api/health":
                # `config` says WHICH template config this server was started
                # against -- the launcher can auto-detect it, so the answer is
                # not always the one the reader assumes.
                return self._send_json({"ok": True,
                                        "config": CFG.template_config_path})

            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)

    def do_PUT(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)
            if path == "/api/project":
                return self._api_project_put(qs)
            if path == "/api/template":
                return self._api_template_put(qs)
            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)
        finally:
            self._drain_body()     # see do_POST: an unread body desyncs the socket

    def do_DELETE(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)
            if path == "/api/template":
                return self._api_template_delete(qs)
            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)
        finally:
            self._drain_body()     # see do_POST: an unread body desyncs the socket

    def do_POST(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)

            if path == "/api/image":
                return self._api_image(qs)
            if path == "/api/import-xlsx":
                return self._api_import_xlsx()
            if path == "/api/validate-compliance":
                return self._api_validate_compliance()
            if path == "/api/export":
                return self._api_export(qs)
            if path == "/api/export-stream":
                return self._api_export_stream(qs)
            if path == "/api/export-xlsx":
                return self._api_export_xlsx()
            if path == "/api/import-docx":
                return self._api_import_docx()
            if path == "/api/new-from-template":
                return self._api_new_from_template()
            if path == "/api/paste-import":
                return self._api_paste_import(qs)
            if path == "/api/copy-diff":
                return self._api_copy_diff(qs)
            if path == "/api/apply-update":
                return self._api_apply_update()
            if path == "/api/rollback":
                return self._api_rollback()
            if path == "/api/autosave-restore":
                return self._api_autosave_restore()
            if path == "/api/project-delete":
                return self._api_project_delete()
            if path == "/api/trash-restore":
                return self._api_trash_restore()
            if path == "/api/project-rename":
                return self._api_project_rename()
            if path == "/api/project-copy":
                return self._api_project_copy()
            if path == "/api/report-new":
                return self._api_report_new()
            if path == "/api/refcol":
                return self._api_refcol()
            if path == "/api/preview-section":
                return self._api_preview_section()
            if path == "/api/merge3":
                return self._api_merge3()
            if path == "/api/merge3-apply":
                return self._api_merge3_apply()
            if path == "/api/asset-tag":
                return self._api_asset_tag()
            if path == "/api/asset-rename":
                return self._api_asset_rename()
            if path == "/api/asset-delete":
                return self._api_asset_delete()
            if path == "/api/open-path":
                return self._api_open_path()

            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)
        finally:
            # The body comes off the socket on EVERY route out of here -- the
            # unknown path, the handler that refused before reading, the one
            # that raised half way. The socket is kept alive, so bytes left on
            # it are parsed as the START of the next request: that request is
            # then answered out of the leftovers -- 501 "Unsupported method",
            # or 400 "Bad request syntax" -- and the caller sees an invented
            # failure instead of the answer it asked for.
            # Only a drain that no handler can forget holds that shut, so it
            # lives here rather than in each of the thirty handlers.
            self._drain_body()

    def _handle_exc(self, exc):
        # Never crash the serve loop; always return a JSON error.
        if isinstance(exc, ValueError):
            status = 400
        elif isinstance(exc, FileNotFoundError):
            status = 404
        elif isinstance(exc, json.JSONDecodeError):
            status = 400
            exc = "invalid JSON body: %s" % exc
        else:
            status = 500
            sys.stderr.write(traceback.format_exc())
        try:
            self._send_error_json(exc, status=status)
        except Exception:
            pass

    # --- endpoint implementations ---

    def _serve_v2_index(self):
        """Serve the workbench shell: builder/web/assets/v2/index.html.

        Falls back to the old single-file UI while that file is being written,
        so "/" never 404s. /legacy always serves the old UI.
        """
        path = os.path.join(HERE, "assets", "v2", "index.html")
        if not os.path.isfile(path):
            return self._serve_app_html()
        with open(path, "rb") as fh:
            body = fh.read()
        return self._send_bytes(body, "text/html; charset=utf-8")

    def _serve_app_html(self):
        path = os.path.join(HERE, "app.html")
        if not os.path.isfile(path):
            # app.html is produced by the frontend agent; until then, a stub.
            stub = (
                b"<!doctype html><meta charset=utf-8>"
                b"<title>Document Builder</title>"
                b"<h1>Document Builder</h1>"
                b"<p>app.html is not present yet.</p>"
            )
            return self._send_bytes(stub, "text/html; charset=utf-8")
        with open(path, "rb") as fh:
            body = fh.read()
        return self._send_bytes(body, "text/html; charset=utf-8")

    def _serve_asset(self, path):
        rel = path[len("/assets/"):]
        rel = rel.replace("\\", "/")
        if ".." in rel.split("/"):
            return self._send_error_json("forbidden", status=403)
        full = os.path.join(HERE, "assets", *rel.split("/"))
        if not os.path.isfile(full):
            return self._send_error_json("not found", status=404)
        ctype = _guess_content_type(full)
        with open(full, "rb") as fh:
            body = fh.read()
        return self._send_bytes(body, ctype)

    def _serve_project_file(self, path, qs):
        """Serve a project-relative file (e.g. images/<name>.png) from the
        project folder identified by the `dir` query param.

        The frontend requests image thumbnails as `images/<file>?dir=<project>`.
        Only the project's `images/` subtree is served; traversal is rejected.
        """
        dir_arg = (qs.get("dir") or [None])[0]
        rel = path.lstrip("/")  # e.g. "images/1-1_checklist.png"
        rel = rel.replace("\\", "/")
        parts = rel.split("/")
        if parts[0] != "images" or ".." in parts:
            return self._send_error_json("forbidden", status=403)
        try:
            project_dir = resolve_project_dir(dir_arg, create=False)
        except ValueError as exc:
            return self._send_error_json(exc, status=400)
        full = os.path.join(project_dir, *parts)
        # Containment guard: resolved file must stay under the project dir.
        full_abs = os.path.abspath(full)
        pdir_abs = os.path.abspath(project_dir)
        if os.path.normcase(os.path.commonpath([full_abs, pdir_abs])) != os.path.normcase(pdir_abs):
            return self._send_error_json("forbidden", status=403)
        if not os.path.isfile(full_abs):
            return self._send_error_json("not found", status=404)
        ctype = _guess_content_type(full_abs)
        with open(full_abs, "rb") as fh:
            body = fh.read()
        return self._send_bytes(body, ctype)

    # TODO: extract the project scanning / rename / copy / delete block below
    # into store/projects_store.py. Not in this pass: several agents are editing
    # against server.py's current shape and the split would collide with them.
    def _api_projects_list(self):
        """List every report under the reports root, at any depth up to three.

        Returns {"projects":[{"dir","title","template","mtime"}, ...]} sorted by
        most-recently-modified first. Returns an empty list (not an error) when
        no reports_root is configured.

        Reports used to sit one level below the root; they now sit three, at
        PROJECT/MODULE/STAGE, and both layouts can be present at the same time.
        Scanning one level deep therefore reported nothing at all after a report
        root was reorganised. This walks down to depth three and stops descending
        the moment a folder holds a project.json, so a report is found wherever
        it sits.

        The identifier stays what it always was -- the report's path relative to
        the reports root, now "PROJECT/MODULE/STAGE" instead of "MODULE" -- which
        is exactly what resolve_project_dir accepts and already guards for
        containment, so a client can keep handing an entry's "dir" straight back
        as ?dir=. Folders that can never be a report (bookkeeping dirs, images/,
        out/, templates/) are skipped through the same _is_tree_dir the report
        tree uses, so the two listings cannot disagree about what is a report.
        """
        root = CFG.reports_root
        out = []

        def walk(parent, parts):
            for name in _listdir_quiet(parent):
                if not _is_tree_dir(parent, name):
                    continue
                path = os.path.join(parent, name)
                here = parts + [name]
                if _has_project(path):
                    out.append(_project_row(path, here))
                elif len(here) < 3:
                    walk(path, here)

        if root and os.path.isdir(root):
            walk(root, [])
        out.sort(key=lambda p: p["mtime"], reverse=True)
        return self._send_json({"projects": out})

    def _api_project_get(self, qs):
        dir_arg = (qs.get("dir") or [None])[0]
        project_dir = resolve_project_dir(dir_arg, create=False)
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return self._send_json(
                {"project": None, "meta_info": {"exists": False, "mtime": None}}
            )
        # A file that will not parse is a fact about the FILE, not about the
        # request. Left to _handle_exc it came back as "invalid JSON body", the
        # wording meant for a malformed request, with the parser's own text as
        # the whole message. Say what it is, flag it, and keep the parser's
        # words as a detail somebody debugging the file can use.
        try:
            project = _load_json_file(pj)
        except (ValueError, UnicodeDecodeError) as exc:
            return self._send_json(
                {"error": "this report's project.json is not valid JSON",
                 "unreadable": True, "detail": str(exc),
                 "dir": self._rel_dir(project_dir)}, status=422)
        mtime = os.path.getmtime(pj)
        return self._send_json(
            {"project": project, "meta_info": {"exists": True, "mtime": mtime}}
        )

    def _api_project_put(self, qs):
        """Write project.json. Query: dir, saved_at (the mtime the client last
        saw), overwrite (the explicit way past a conflict).

        A SAVE NEVER BRINGS A REPORT BACK. A PUT that carries a token, or asks
        to overwrite, is by definition a save of a report the client has read;
        if that report is not on disk any more it was deleted meanwhile (moved
        to the trash from the shelf), and writing it would recreate the folder
        with none of its images, baseline or history -- and then block the
        trash restore, which refuses an occupied path. That answers 410 with
        {"gone": true} and writes nothing.

        A PUT with neither token nor overwrite flag to a folder that does not
        exist is the create path the previous interface and the smoke test use
        to seed a new report, and it still creates. /api/report-new is the
        proper door for that; this one stays open only because callers depend
        on it.
        """
        dir_arg = (qs.get("dir") or [None])[0]
        saved_at = (qs.get("saved_at") or [None])[0]
        overwrite = (qs.get("overwrite") or [None])[0] in ("1", "true", "yes")
        project_dir = resolve_project_dir(dir_arg, create=False)
        project = self._read_json()
        if not isinstance(project, dict):
            return self._send_error_json("body must be a JSON object")
        sv = project.get("schema_version")
        if sv is not None and sv != 1:
            return self._send_error_json("unsupported schema_version: %r" % sv)
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            if saved_at or overwrite:
                return self._send_json(
                    {"error": "this report is no longer on disk; nothing was written",
                     "gone": True, "dir": self._rel_dir(project_dir)},
                    status=410)
            os.makedirs(project_dir, exist_ok=True)
            os.makedirs(os.path.join(project_dir, "images"), exist_ok=True)
        # A4 optimistic concurrency: if the client last saw mtime `saved_at` but the
        # file changed on disk since (a second tab, or an applied update bundle),
        # refuse with 409 so a stale autosave cannot clobber / revert it. The client
        # then reloads or explicitly overwrites (by re-PUTting without saved_at).
        if saved_at and not overwrite and os.path.isfile(pj):
            try:
                if abs(os.path.getmtime(pj) - float(saved_at)) > 1e-6:
                    return self._send_json(
                        {"error": "conflict: project.json changed on disk",
                         "conflict": True, "saved_at": os.path.getmtime(pj)},
                        status=409)
            except (ValueError, OSError):
                pass
        # The explicit overwrite replaces a version this client has never seen,
        # so that version is kept first: the snapshot is what makes "keep my
        # version" a choice rather than a loss.
        if overwrite and os.path.isfile(pj):
            autosave_snapshot(project_dir, "overwrite")
        body = json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")
        atomic_write(pj, body)
        autosave_snapshot(project_dir, "save")  # capture every saved state
        return self._send_json(
            {
                "ok": True,
                "saved_at": os.path.getmtime(pj),
                "path": pj.replace("\\", "/"),
            }
        )

    def _api_image(self, qs):
        dir_arg = (qs.get("dir") or [None])[0]
        section = (qs.get("section") or [None])[0]
        name = (qs.get("name") or [None])[0]

        ctype = (self.headers.get("Content-Type") or "").lower()
        if ctype.startswith("image/png"):
            png_bytes = self._read_body()
        else:
            payload = self._read_json()
            b64 = payload.get("png_b64")
            if not b64:
                return self._send_error_json("missing png bytes / png_b64")
            if "," in b64 and b64.lstrip().lower().startswith("data:"):
                b64 = b64.split(",", 1)[1]
            png_bytes = base64.b64decode(b64)
            section = payload.get("section", section)
            name = payload.get("name", name)

        if section is None:
            return self._send_error_json("missing 'section'")
        if not png_bytes:
            return self._send_error_json("empty image payload")
        if png_bytes[:8] != b"\x89PNG\r\n\x1a\n":
            return self._send_error_json("payload is not a PNG")

        project_dir = resolve_project_dir(dir_arg, create=True)
        rel, abs_path = next_image_path(project_dir, section, name)
        atomic_write(abs_path, png_bytes)
        return self._send_json({"file": rel})

    def _api_import_xlsx(self):
        payload = self._read_json()
        b64 = payload.get("xlsx_b64")
        mode = payload.get("mode", "grid")
        if not b64:
            return self._send_error_json("missing 'xlsx_b64'")
        if "," in b64 and b64.lstrip().lower().startswith("data:"):
            b64 = b64.split(",", 1)[1]
        xlsx_bytes = base64.b64decode(b64)
        if mode == "grid":
            return self._send_json(parse_xlsx_grid(xlsx_bytes))
        if mode == "compliance":
            return self._send_json(parse_xlsx_compliance(xlsx_bytes))
        return self._send_error_json("unknown mode: %r" % mode)

    def _api_validate_compliance(self):
        payload = self._read_json()
        data = payload.get("data")
        if data is None:
            return self._send_error_json("missing 'data'")
        try:
            result = validate_compliance(data)
        except (ImportError, ModuleNotFoundError) as exc:
            return self._send_error_json(
                "engine not available: %s" % exc, status=503
            )
        return self._send_json(result)

    def _api_export(self, qs):
        dir_arg = (qs.get("dir") or [None])[0]
        fmt = (qs.get("fmt") or ["docx"])[0]
        if fmt not in ("docx", "pdf"):
            return self._send_error_json("fmt must be docx|pdf")
        project_dir = resolve_project_dir(dir_arg, create=False)
        if not os.path.isfile(os.path.join(project_dir, "project.json")):
            return self._send_error_json("no project.json in dir", status=404)
        payload = {}
        if int(self.headers.get("Content-Length") or 0) > 0:
            payload = self._read_json()
        save_first = bool(payload.get("save_first"))
        try:
            result = run_export(project_dir, fmt, save_first=save_first)
        except (ImportError, ModuleNotFoundError) as exc:
            return self._send_error_json(
                "engine not available: %s" % exc, status=503
            )
        return self._send_json(result)

    def _api_export_stream(self, qs):
        """Export with a LIVE progress feed as newline-delimited JSON. Emits
        {type:'phase',label}, {type:'progress',done,total,label} (throttled to
        ~50 lines), then {type:'done',result} or {type:'error',error}. The body is
        close-delimited (Connection: close, no Content-Length); the GUI reads the
        stream to drive a progress bar. /api/export stays as the plain fallback."""
        dir_arg = (qs.get("dir") or [None])[0]
        fmt = (qs.get("fmt") or ["docx"])[0]
        if fmt not in ("docx", "pdf"):
            return self._send_error_json("fmt must be docx|pdf")
        try:
            project_dir = resolve_project_dir(dir_arg, create=False)
        except ValueError as exc:
            return self._send_error_json(str(exc))
        if not os.path.isfile(os.path.join(project_dir, "project.json")):
            return self._send_error_json("no project.json in dir", status=404)
        payload = {}
        if int(self.headers.get("Content-Length") or 0) > 0:
            payload = self._read_json()
        save_first = bool(payload.get("save_first"))

        self.send_response(200)
        self.send_header("Content-Type", "application/x-ndjson; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("X-Accel-Buffering", "no")   # discourage any proxy buffering
        self.send_header("Connection", "close")
        self.end_headers()
        self.close_connection = True

        def emit(obj):
            try:
                self.wfile.write((json.dumps(obj, ensure_ascii=False) + "\n").encode("utf-8"))
                self.wfile.flush()
            except Exception:
                pass   # client gone: writes are ignored, the render still finishes

        last = {"done": -1}

        def on_progress(done, total, label):
            step = max(1, total // 50)        # ~50 lines max; always send the last
            if done - last["done"] >= step or done >= total:
                last["done"] = done
                emit({"type": "progress", "done": done, "total": total, "label": label or ""})

        try:
            result = run_export(project_dir, fmt, save_first=save_first,
                                on_progress=on_progress,
                                on_phase=lambda label: emit({"type": "phase", "label": label}))
            emit({"type": "done", "result": result})
        except (ImportError, ModuleNotFoundError) as exc:
            emit({"type": "error", "error": "engine not available: %s" % exc})
        except Exception as exc:
            sys.stderr.write(traceback.format_exc())
            emit({"type": "error", "error": str(exc)})

    def _api_export_xlsx(self):
        """Export ONE table/datatable block to a .xlsx that visually mirrors the
        Word table. Body: {dir, block}. The project's template config supplies the
        compliance fills/flags (datatable) or the header fill (free table). Returns
        {ok, filename, xlsx_b64} -- the client turns the base64 into a download."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        block = payload.get("block")
        if not isinstance(block, dict):
            return self._send_error_json("missing 'block'")
        project_dir = resolve_project_dir(payload.get("dir"), create=False)
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return self._send_error_json("no project.json in dir", status=404)
        with open(pj, "r", encoding="utf-8") as fh:
            project = json.load(fh)
        # resolve the template config exactly like run_export (compliance section)
        engine = _import_engine()
        tid = project.get("template")
        tpl_cfg = tstore.template_config_path(CFG.reports_root, tid) if tid else None
        explicit = tpl_cfg or CFG.template_config_path
        cfg = engine._load_config(engine._resolve_config_path(project, project_dir, explicit))
        xe = _import_xlsx_export()
        try:
            data = xe.build_block_xlsx(block, cfg)
        except ValueError as ex:
            return self._send_error_json(str(ex))
        return self._send_json({
            "ok": True,
            "filename": xe.filename_for(block),
            "xlsx_b64": base64.b64encode(data).decode("ascii"),
        })

    # --- template library ---

    def _api_templates_list(self):
        return self._send_json(
            {"templates": tstore.list_templates(CFG.reports_root)}
        )

    def _api_template_get(self, qs):
        tid = (qs.get("id") or [None])[0]
        return self._send_json(tstore.get_template(CFG.reports_root, tid))

    def _api_template_put(self, qs):
        tid = (qs.get("id") or [None])[0]
        body = self._read_json()
        if not isinstance(body, dict):
            return self._send_error_json("body must be a JSON object")
        config = body.get("config")
        if not isinstance(config, dict):
            return self._send_error_json("missing 'config'")
        name = body.get("name") or tid
        skeleton = body.get("skeleton", [])
        rid = tstore.save_template(
            CFG.reports_root, tid, name, config, skeleton, atomic_write
        )
        return self._send_json({"ok": True, "id": rid})

    def _api_template_delete(self, qs):
        tid = (qs.get("id") or [None])[0]
        tstore.delete_template(CFG.reports_root, tid)
        return self._send_json({"ok": True})

    def _api_import_docx(self):
        payload = self._read_json()
        b64 = payload.get("docx_b64")
        mode = payload.get("mode", "report")
        if not b64:
            return self._send_error_json("missing 'docx_b64'")
        if "," in b64 and b64.lstrip().lower().startswith("data:"):
            b64 = b64.split(",", 1)[1]
        docx_bytes = base64.b64decode(b64)
        try:
            di = _import_docx_module()
        except (ImportError, ModuleNotFoundError) as exc:
            return self._send_error_json(
                "docx parser not available: %s" % exc, status=503
            )
        with _temp_docx(docx_bytes) as tmp_path:
            if mode == "template":
                return self._import_docx_template(di, tmp_path, payload)
            if mode == "report":
                return self._import_docx_report(di, tmp_path, payload)
        return self._send_error_json("unknown mode: %r" % mode)

    def _import_docx_template(self, di, tmp_path, payload):
        warnings = []
        logo_dir = tempfile.mkdtemp(prefix="docx_logo_")
        try:
            derived = di.derive_template(
                tmp_path, logo_dir=logo_dir, warn=warnings.append
            )
            tid, rname = tstore.save_derived_template(
                CFG.reports_root, payload.get("name"), derived,
                atomic_write, _sanitize_name,
                logo_dir=logo_dir, warn=warnings.append,
            )
        finally:
            _rmtree_quiet(logo_dir)
        return self._send_json(
            {"id": tid, "name": rname, "warnings": warnings or None}
        )

    def _import_docx_report(self, di, tmp_path, payload):
        warnings = []
        dir_arg = payload.get("dir")
        if not dir_arg:
            return self._send_error_json("missing 'dir' for report import")
        project_dir = resolve_project_dir(dir_arg, create=True)
        images_dir = os.path.join(project_dir, "images")
        parsed = di.parse_docx_report(
            tmp_path, images_dir=images_dir, warn=warnings.append
        )
        meta = parsed.get("meta", {})
        outline = parsed.get("outline", [])
        _rewrite_image_paths(outline)

        # Determine the bound template id -- import MUST NOT dead-end.
        tid = payload.get("template")
        if not (tid and tstore.template_config_path(CFG.reports_root, tid)):
            logo_dir = tempfile.mkdtemp(prefix="docx_logo_")
            try:
                derived = di.derive_template(
                    tmp_path, logo_dir=logo_dir, warn=warnings.append
                )
                name = meta.get("title") or os.path.basename(
                    project_dir.rstrip(os.sep)
                )
                tid, _ = tstore.save_derived_template(
                    CFG.reports_root, name, derived,
                    atomic_write, _sanitize_name,
                    logo_dir=logo_dir, warn=warnings.append,
                )
            finally:
                _rmtree_quiet(logo_dir)

        project_seed = {
            "schema_version": 1, "template": tid,
            "meta": meta, "outline": outline,
        }
        return self._send_json(
            {
                "project": project_seed,
                "template_id": tid,
                "warnings": warnings or None,
            }
        )

    def _api_new_from_template(self):
        payload = self._read_json()
        dir_arg = payload.get("dir")
        tid = payload.get("template")
        force = bool(payload.get("force"))
        if not dir_arg:
            return self._send_error_json("missing 'dir'")
        tpl = tstore.get_template(CFG.reports_root, tid)  # 404 if absent
        outline = _instantiate_skeleton(tpl.get("skeleton", []))
        project = {
            "schema_version": 1, "template": tid,
            "meta": _empty_meta(), "outline": outline,
        }
        project_dir = resolve_project_dir(dir_arg, create=True)
        pj = os.path.join(project_dir, "project.json")
        # Refuse to clobber an existing project unless the caller explicitly
        # opted in; even then keep a .bak so an accidental overwrite is
        # recoverable. (This is the most destructive path in the tool.)
        if os.path.isfile(pj):
            if not force:
                return self._send_error_json(
                    "a project already exists in this folder", status=409
                )
            try:
                os.replace(pj, pj + ".bak")
            except OSError:
                pass
        atomic_write(
            pj, json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")
        )
        return self._send_json({"ok": True, "project": project})

    def _api_paste_import(self, qs):
        """Full-project replace from pasted text (de-LLM upstream channel, WS2).

        Body: the raw pasted project.json text (bytes go browser -> HTTP -> disk,
        never through a model). Validates, backs up the existing project into the
        shared rollback history, writes it, and emits a compact structural diff to
        <project>/_paste_diff.md. A truncated/mangled paste fails to parse and is
        rejected -- the on-disk project is left untouched. Returns
        {ok, backup, warn, diff, diff_file}."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        dir_arg = (qs.get("dir") or [None])[0]
        raw = self._read_body()
        if not raw:
            return self._send_error_json("empty paste body -- nothing changed")
        try:
            project = json.loads(raw.decode("utf-8"))
        except Exception as exc:
            return self._send_error_json(
                "could not parse pasted JSON (truncated?): %s -- nothing changed" % exc)
        if not isinstance(project, dict):
            return self._send_error_json(
                "pasted JSON must be an object -- nothing changed")
        sv = project.get("schema_version")
        if sv is not None and sv != 1:
            return self._send_error_json(
                "unsupported schema_version: %r -- nothing changed" % sv)

        project_dir = resolve_project_dir(dir_arg, create=True)
        pj = os.path.join(project_dir, "project.json")
        old_project = None
        if os.path.isfile(pj):
            try:
                with open(pj, "r", encoding="utf-8") as fh:
                    old_project = json.load(fh)
            except Exception:
                old_project = None

        # identity check: guard against pasting into the wrong project folder.
        warns = []
        if isinstance(old_project, dict):
            ot = (old_project.get("meta") or {}).get("title")
            nt = (project.get("meta") or {}).get("title")
            if ot and nt and ot != nt:
                warns.append('pasted title "%s" != existing "%s"' % (nt, ot))
            otpl, ntpl = old_project.get("template"), project.get("template")
            if otpl and ntpl and otpl != ntpl:
                warns.append('pasted template "%s" != existing "%s"' % (ntpl, otpl))

        name = os.path.basename(project_dir.rstrip(os.sep)) or "project"
        diff_md = paste_import_diff(old_project, project, name, warns)

        body = json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")
        autosave_snapshot(project_dir, "prepaste")  # keep pre-paste state recoverable
        apply_update = _import_apply_update()
        rel = os.path.relpath(pj, CFG.reports_root)
        rec = apply_update.record_replace(CFG.reports_root, rel, body)

        diff_path = os.path.join(project_dir, "_paste_diff.md")
        atomic_write(diff_path, diff_md.encode("utf-8"))

        return self._send_json({
            "ok": True,
            "backup": (rec.get("backup") or "").replace("\\", "/"),
            "warn": warns,
            "diff": diff_md,
            "diff_file": diff_path.replace("\\", "/"),
        })

    def _api_copy_diff(self, qs):
        """Compute the upstream 'Copy diff' delta for a project (WS: incremental
        upstream). Body: the CURRENT project JSON (the editor's in-memory state,
        same payload as Copy text). The server diffs it against the last synced
        <project>/_baseline.json and returns a COMPACT op-diff the user pastes to
        the assistant instead of the whole report -- only edited sections travel.

        Returns {ok, no_baseline, empty, diff_text, summary, diff_chars,
        full_chars}. no_baseline=True when the project has never been synced from
        the assistant yet (the UI falls back to Copy text)."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        dir_arg = (qs.get("dir") or [None])[0]
        raw = self._read_body()
        if not raw:
            return self._send_error_json("empty body -- nothing to diff")
        try:
            current = json.loads(raw.decode("utf-8"))
        except Exception as exc:
            return self._send_error_json("could not parse current project JSON: %s" % exc)
        if not isinstance(current, dict):
            return self._send_error_json("current project must be a JSON object")

        project_dir = resolve_project_dir(dir_arg)
        # The delta names the report it belongs to, and that name is the ONLY
        # address a pasted diff carries -- the CLI and the paste box both route
        # by it when no dir is given alongside. It must therefore be the report's
        # path under the reports root, not its folder's basename: with reports
        # nested as <project>/<module>/<stage>, a basename is the stage name,
        # which is shared by every module and can even name a different report.
        # A delta cut for one report then applies cleanly to another and the
        # edits never reach the one they were meant for.
        name = os.path.relpath(project_dir, CFG.reports_root).replace("\\", "/")
        if name in (".", "", os.curdir):
            name = os.path.basename(project_dir.rstrip(os.sep)) or "project"
        baseline_path = os.path.join(project_dir, "_baseline.json")
        if not os.path.isfile(baseline_path):
            return self._send_json({
                "ok": True, "no_baseline": True,
                "hint": ("No sync baseline yet for this project -- Copy diff starts "
                         "working after the next 'Apply update' from the assistant. "
                         "Use Copy text this once."),
            })
        try:
            with open(baseline_path, "r", encoding="utf-8") as fh:
                baseline = json.load(fh)
        except Exception as exc:
            return self._send_error_json("baseline unreadable: %s" % exc)

        apply_update = _import_apply_update()
        diff = apply_update.make_text_diff(baseline, current, name)
        diff_text = json.dumps(diff, ensure_ascii=False, separators=(",", ":"))
        full_text = json.dumps(current, ensure_ascii=False, separators=(",", ":"))
        return self._send_json({
            "ok": True,
            "no_baseline": False,
            "empty": apply_update.diff_is_empty(diff),
            # Edits spread across a small report (or a datatable-cell change, which
            # resends its whole section) can make the op-diff bigger than the full
            # text; surface that so the UI never silently hands over MORE to paste.
            "smaller": len(diff_text) < len(full_text),
            "diff_text": diff_text,
            "summary": apply_update.diff_summary(diff),
            "diff_chars": len(diff_text),
            "full_chars": len(full_text),
        })

    def _api_apply_update(self):
        """Apply an update package to the reports root.

        Body: {"name": <filename>, "zip_b64": <base64 zip>} for a package, or
        {"dir": <report>, "diff_text" | "diff": <op-diff>} for a returned text
        diff. A package is stored under reports_root/_updates/ and applied via
        the shared apply_update module (full backup of every overwritten file
        first). Returns the apply summary {note, actions:[{verb,rel,warn}],
        backup, logs}.

        BASELINE POLICY. A package carries the fingerprint of the state it was
        cut from. When this machine's report has moved past that state, applying
        would silently overwrite everything changed in between, so a MISMATCH is
        now REFUSED with 409 {"error":"baseline_mismatch", packageBase,
        localBase} -- it used to be only a warning.

        A MISSING baseline is a different case and must NOT be refused: a report
        that has never been exchanged has no _baseline.json at all, and refusing
        it would lock it out of the exchange for good. Those are allowed through,
        take an autosave snapshot first, and get a baseline stamped afterwards
        (run_plan / apply_text_diff do the stamping).

        The rule itself is apply_update.check_baseline's -- see
        _baseline_verdict. It compares the local baseline's CONTENT with the
        package's declared base_sha, never the current project.json.
        """
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        if payload.get("diff") or payload.get("diff_text"):
            return self._apply_update_diff(payload)
        b64 = payload.get("zip_b64") or ""
        if not b64:
            return self._send_error_json("missing 'zip_b64'")
        try:
            raw = base64.b64decode(b64)
        except Exception:
            return self._send_error_json("invalid base64 zip")
        base = payload.get("name") or "update"
        if base.lower().endswith(".zip"):
            base = base[:-4]
        name = (_sanitize_name(base) or "update") + ".zip"
        updates_dir = os.path.join(CFG.reports_root, "_updates")
        os.makedirs(updates_dir, exist_ok=True)
        dest = os.path.join(updates_dir, name)
        atomic_write(dest, raw)
        refusal = self._bundle_baseline_refusal(dest)
        if refusal:
            return self._send_json(refusal, status=409)
        # snapshot every project's current state BEFORE applying, so an update
        # bundle can never silently overwrite local edits beyond recovery.
        autosaved = autosave_all(CFG.reports_root, "preapply")
        apply_update = _import_apply_update()
        summary = apply_update.apply_bundle(CFG.reports_root, dest, dry=False)
        return self._send_json({"ok": True, "autosaved": autosaved, **summary})

    def _bundle_baseline_refusal(self, bundle_path):
        """Refuse a package cut from a state this machine has already moved past.

        Returns the 409 body, or None when the package may be applied. A report
        that has NO baseline yet is allowed through on purpose (and snapshotted
        first) -- see the baseline policy in _api_apply_update. The verdict comes
        from the shared library, never from a fingerprint of the live file.
        """
        apply_update = _import_apply_update()
        try:
            manifest, _actions = apply_update.read_bundle(CFG.reports_root,
                                                          bundle_path)
        except Exception:
            return None       # an unreadable package fails later, with detail
        manifest = manifest or {}
        default_base = manifest.get("base_sha")
        for pdir, spec in (manifest.get("projects") or {}).items():
            base = (spec or {}).get("base_sha") if isinstance(spec, dict) else None
            base = base or default_base
            if not base:
                continue
            try:
                project_dir = resolve_project_dir(pdir, create=False)
            except ValueError:
                continue
            if not os.path.isfile(os.path.join(project_dir, "project.json")):
                continue
            verdict, refusal = _baseline_verdict(apply_update, project_dir,
                                                 base, pdir)
            if refusal:
                return refusal
            if verdict == "no_baseline":
                # Never exchanged: full replace after a snapshot; run_plan
                # stamps the first baseline afterwards.
                autosave_snapshot(project_dir, "prebaseline")
        return None

    def _apply_update_diff(self, payload):
        """Apply a returned op-diff (the counterpart of the GUI's Copy diff) to
        one report. Same baseline policy as the package path."""
        diff = payload.get("diff")
        if diff is None:
            try:
                diff = json.loads(payload.get("diff_text") or "")
            except Exception as exc:
                return self._send_error_json("could not parse diff text: %s" % exc)
        if not isinstance(diff, dict):
            return self._send_error_json("diff must be a JSON object")
        dir_arg = payload.get("dir") or diff.get("dir")
        project_dir = resolve_project_dir(dir_arg, create=False)
        if not os.path.isfile(os.path.join(project_dir, "project.json")):
            return self._send_error_json("no project.json in dir", status=404)
        apply_update = _import_apply_update()
        verdict, refusal = _baseline_verdict(apply_update, project_dir,
                                             diff.get("base_sha"), dir_arg)
        if refusal:
            return self._send_json(refusal, status=409)
        # A missing baseline is allowed through on purpose: snapshot, apply, and
        # let apply_text_diff stamp the baseline the next exchange compares to.
        autosave_snapshot(project_dir,
                          "prebaseline" if verdict == "no_baseline"
                          else "preapply")
        rel_dir = os.path.relpath(project_dir, CFG.reports_root).replace("\\", "/")
        try:
            result = apply_update.apply_text_diff(CFG.reports_root, diff, rel_dir)
        except apply_update.BaselineMismatch as exc:
            # The library re-checks; keep the answer the documented 409 rather
            # than letting it surface as a 500.
            return self._send_json(exc.as_dict(), status=409)
        out = {"ok": True}
        out.update(result)
        return self._send_json(out)

    def _api_rollback(self):
        """Undo the most recent apply / paste-import OF ONE REPORT by restoring
        that report's newest backup. Non-interactive counterpart of the CLI
        --rollback. Body {dir} names the report the sync screen has open.

        The backup history is shared by the whole root, so the newest backup in
        it is whichever report was written last -- undoing that when the screen
        says another report's name reverts a change nobody asked about and
        leaves the one they did ask about untouched. With no ``dir`` (the old
        single-file UI asks for exactly that) the root-wide behaviour stands.

        Returns {ok, dir, restored, from, pre}, or 409 when this report has
        nothing of its own to undo."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        dir_arg = payload.get("dir")
        rel_dir = None
        if dir_arg:
            # Same containment guard as every other dir-taking endpoint: the
            # scope is a path under reports_root or it is nothing.
            project_dir = resolve_project_dir(dir_arg, create=False)
            rel = os.path.relpath(project_dir, CFG.reports_root).replace("\\", "/")
            if rel not in (".", ""):       # the root itself is not a report
                rel_dir = rel
        apply_update = _import_apply_update()
        result = apply_update.rollback_last(CFG.reports_root, dir_name=rel_dir)
        if not result.get("ok"):
            # "nothing of yours to undo" is a state, not a bad request: 409 so
            # the drawer can tell the two apart.
            status = (409 if result.get("reason") in ("out_of_scope", "spans_reports")
                      else 400)
            return self._send_error_json(result.get("error", "rollback failed"),
                                         status=status)
        return self._send_json(result)

    def _api_autosaves(self, qs):
        """List the local auto-snapshots of a project's project.json (newest first)."""
        dir_arg = (qs.get("dir") or [None])[0]
        project_dir = resolve_project_dir(dir_arg)
        return self._send_json({"autosaves": list_autosaves(project_dir)})

    def _api_autosave_restore(self):
        """Restore a named auto-snapshot to project.json (current is snapshotted
        first, so the restore is itself undoable). Body: {dir, name}."""
        payload = self._read_json()
        dir_arg = payload.get("dir")
        name = payload.get("name")
        if not name:
            return self._send_error_json("missing 'name'")
        project_dir = resolve_project_dir(dir_arg)
        return self._send_json(restore_autosave(project_dir, name))

    def _new_project_target(self, name, parent_dir=None):
        """Validate a NEW report folder name -> (segment, abs path). Single
        sanitized segment, contained under reports_root, not reserved, not '_'.

        The new folder is created NEXT TO the report being renamed or copied
        (parent_dir), so a report at PROJ/MOD/STAGE stays inside its module
        instead of being relocated to the reports root. A flat report has the
        root as its parent and therefore behaves exactly as it did before.
        """
        if not name or not isinstance(name, str):
            raise ValueError("missing new project name")
        seg = _sanitize_name(name.strip())
        if not seg or seg.startswith("_") or seg in _RESERVED_ROOT_DIRS:
            raise ValueError("invalid project name: %r" % name)
        rel = seg
        if parent_dir:
            rel_parent = os.path.relpath(parent_dir, CFG.reports_root)
            if rel_parent not in (".", ""):
                rel = os.path.join(rel_parent, seg)
        return seg, resolve_project_dir(rel)  # resolve_project_dir enforces containment

    def _rel_dir(self, path):
        """A project dir as the frontend addresses it: relative, forward slashes."""
        return os.path.relpath(path, CFG.reports_root).replace("\\", "/")

    def _api_project_delete(self):
        """Move a project to reports_root/_trash/<name>-<ts>/ (recoverable, never
        a hard delete). Body: {dir}.

        The move also writes "<entry>.trashinfo.json" beside the folder, holding
        the path it came from and when it went. That is the one fact the entry
        name cannot carry and the one fact a restore needs -- see the trash
        helpers above. The answer keeps every key it has always had and adds the
        entry's id, so a caller can name what it just trashed.
        """
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        project_dir = resolve_project_dir(payload.get("dir"))
        if not os.path.isdir(project_dir):
            return self._send_error_json("no such project", status=404)
        origin = self._rel_dir(project_dir)
        name = os.path.basename(project_dir.rstrip(os.sep)) or "project"
        trash = _trash_root(CFG.reports_root)
        os.makedirs(trash, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        # Two deletes in the same second must not collide: shutil.move onto an
        # existing directory moves the folder INSIDE it, which would bury one
        # entry in another and lose its sidecar.
        entry_id = "%s-%s" % (name, ts)
        n = 1
        while os.path.exists(os.path.join(trash, entry_id)):
            entry_id = "%s-%s-%d" % (name, ts, n)
            n += 1
        dest = os.path.join(trash, entry_id)
        shutil.move(project_dir, dest)
        try:
            atomic_write(dest + TRASH_INFO_SUFFIX, json.dumps(
                {"from": origin, "deleted": time.time()},
                ensure_ascii=False, indent=2).encode("utf-8"))
        except OSError:
            pass   # the entry is still listed and still restorable, by its name
        return self._send_json({"ok": True, "id": entry_id, "from": origin,
                                "trashed_to": dest.replace("\\", "/")})

    def _api_trash(self):
        """Everything in the trash, newest deletion first: {"items":[...]}.

        Each item carries the id the restore takes, the name to show, the path
        it came from, whether that path was recorded or guessed, when it went,
        how many reports it holds, and whether anything occupies that path as
        this listing is read -- which is the state a restore would refuse on,
        read once here rather than guessed at by the caller.
        """
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        return self._send_json({"items": list_trash(CFG.reports_root)})

    def _api_trash_restore(self):
        """Move one trash entry back where it came from. Body: {id}.

        Refuses rather than overwrites: if anything at all occupies the original
        path the answer is 409 with code "occupied" and NOTHING is moved. The
        destination is resolved through resolve_project_dir, so a restore can
        never write outside the reports root, and every segment of the recorded
        origin is checked first -- the sidecar is a file on disk and is treated
        as input, not as this server's own word.
        """
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        entry_id = _trash_entry_id(payload.get("id"))
        src = resolve_project_dir(os.path.join(TRASH_DIRNAME, entry_id))
        if not os.path.isdir(src):
            return self._send_error_json("no such trash entry", status=404)

        info = _read_json_quiet(src + TRASH_INFO_SUFFIX)
        info = info if isinstance(info, dict) else {}
        origin = str(info.get("from") or "").strip() or _origin_from_name(entry_id)
        parts = _safe_origin(origin)
        dest = resolve_project_dir("/".join(parts))
        if os.path.exists(dest):
            return self._send_json(
                {"error": "something already occupies %s" % "/".join(parts),
                 "code": "occupied", "dir": "/".join(parts)}, status=409)

        parent = os.path.dirname(dest)
        if parent:
            os.makedirs(parent, exist_ok=True)
        shutil.move(src, dest)
        try:
            os.remove(src + TRASH_INFO_SUFFIX)
        except OSError:
            pass   # a leftover sidecar names a folder that is no longer there
        return self._send_json({"ok": True, "id": entry_id,
                                "dir": self._rel_dir(dest)})

    def _api_project_rename(self):
        """Rename a project or module FOLDER, or a report's TITLE.
        Body: {dir, new_name, title?} -> {ok, dir, title?, moved}.

        A report's folder is its ADDRESS. `dir` is the only thing an update
        package or an op-diff carries, and the stage leaf of PROJECT/MODULE/STAGE
        is read back as the report's stage. Moving that folder therefore orphans
        every package already cut for the old path and invents a stage nobody
        chose, so this endpoint refuses to do it: renaming a report writes
        meta.title and leaves the folder exactly where it is (`moved: false`).

        The rule is enforced here rather than trusted to the caller, because the
        damage is silent and permanent for anyone holding an old package. A
        folder that is not itself a report -- a project or a module container --
        still moves as before.

        A report asked to rename by `new_name` alone is therefore a request this
        endpoint cannot carry out, and it is refused. It used to be rounded to
        the nearest thing it could do -- write the folder name into meta.title --
        which answered 200 to a caller that had asked for neither: a title it
        never chose, and a folder that never moved.
        """
        payload = self._read_json()
        src = resolve_project_dir(payload.get("dir"))
        if not os.path.isdir(src):
            return self._send_error_json("no such project", status=404)

        if _has_project(src):
            title = payload.get("title")
            if title is None and payload.get("new_name") is not None:
                return self._send_error_json(
                    "a report's folder is its address and is never moved; "
                    "send 'title' to rename the report", status=400)
            written = self._write_project_title(src, title)
            return self._send_json({"ok": True, "dir": self._rel_dir(src),
                                    "title": written, "moved": False})

        seg, dest = self._new_project_target(payload.get("new_name"),
                                             os.path.dirname(src))
        if os.path.normcase(dest) == os.path.normcase(src):
            return self._send_json({"ok": True, "dir": self._rel_dir(dest)})  # no-op
        if os.path.exists(dest):
            return self._send_error_json("a project named '%s' already exists" % seg,
                                         status=409)
        shutil.move(src, dest)
        written = self._write_project_title(dest, payload.get("title"))
        return self._send_json({"ok": True, "dir": self._rel_dir(dest),
                                "title": written, "moved": True})

    def _write_project_title(self, project_dir, title):
        """Set meta.title of the report in ``project_dir``. Returns what was
        written, or None when there was nothing to write (no title asked for, or
        the folder is a container rather than a report). Best effort, exactly as
        it was when this lived inline in the rename handler: a document that
        cannot be parsed keeps the name it has rather than failing the call."""
        if title is None:
            return None
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return None
        try:
            with open(pj, encoding="utf-8") as fh:
                d = json.load(fh)
            d.setdefault("meta", {})["title"] = title
            atomic_write(pj, json.dumps(d, ensure_ascii=False,
                                        indent=2).encode("utf-8"))
            return title
        except Exception:
            return None

    def _api_project_copy(self):
        """Duplicate a project (project.json + images/ only, not the heavy
        _backups/_autosave/_updates) under a new name. Body: {dir, new_name}."""
        payload = self._read_json()
        src = resolve_project_dir(payload.get("dir"))
        if not os.path.isfile(os.path.join(src, "project.json")):
            return self._send_error_json("no such project", status=404)
        seg, dest = self._new_project_target(payload.get("new_name"),
                                             os.path.dirname(src))
        if os.path.exists(dest):
            return self._send_error_json("a project named '%s' already exists" % seg,
                                         status=409)
        os.makedirs(dest)
        shutil.copy2(os.path.join(src, "project.json"),
                     os.path.join(dest, "project.json"))
        img = os.path.join(src, "images")
        if os.path.isdir(img):
            shutil.copytree(img, os.path.join(dest, "images"))
        return self._send_json({"ok": True, "dir": self._rel_dir(dest)})

    # --- v2 workbench endpoints ---

    def _api_tree(self):
        """The three-level report shelf: {"projects":[{modules:[{reports:[...]}]}]}.

        Flat legacy reports and the PROJECT/MODULE/STAGE layout are returned by
        the same walk; see build_report_tree. /api/projects keeps its own flat
        shape for the old UI and must not change.
        """
        return self._send_json({"projects": build_report_tree(CFG.reports_root)})

    def _api_version(self):
        """Local tool version. No network call is ever made from here: an
        updater fills 'available'/'needsRestart' in when one exists."""
        return self._send_json({"local": TOOL_VERSION, "available": None,
                                "needsRestart": False})

    def _api_report_new(self):
        """Create a report (or a project / module shell).

        Body {project, module, stage, name, mode:'inherit'|'template'|'docx',
        from, template, clearValues} -> {"dir": "PROJ/MOD/STAGE"}.
        """
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        body = self._read_json()
        payload, status = _lazy_call(("core.report_new", "report_new"),
                                     "create_report", (CFG.reports_root, body),
                                     {"dir": ""})
        return self._send_json(payload, status=status)

    def _api_refcol(self):
        """Build a reference column pulled from another report of this module.

        Body {dir, srcReport, targetBlock, group, axis, title} ->
        {"matches":[...], "column":{...}}. A snapshot, never a live link.
        """
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        body = self._read_json()
        payload, status = _lazy_call(("core.report_new", "report_new"),
                                     "build_reference_column",
                                     (CFG.reports_root, body),
                                     {"matches": [], "column": None})
        return self._send_json(payload, status=status)

    def _api_preview_section(self):
        """Render ONE section to page images. Body {dir, node} ->
        {"pages":[{"png_b64","w","h"}], "ms": int}."""
        body = self._read_json()
        project_dir = resolve_project_dir(body.get("dir"), create=False)
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return self._send_error_json("no project.json in dir", status=404)
        node_id = body.get("node") or body.get("node_id")
        cfg = None
        try:
            cfg = _project_config(project_dir, _read_json_quiet(pj) or {})
        except Exception:
            cfg = None      # the renderer may resolve its own config
        payload, status = _lazy_call(("web.live_preview", "live_preview"),
                                     "render_section",
                                     (project_dir, cfg, node_id),
                                     {"pages": [], "ms": 0})
        return self._send_json(payload, status=status)

    def _api_merge3(self):
        """Three-way compare of an incoming report against the baseline.

        Body {dir, incoming, allowBulkDelete?} -> 200 {merged, conflicts, auto,
        pending, deletions, token, truncated: null}. Nothing is written: this
        only computes what applying would do. ``token`` fingerprints the
        project.json this merge was computed from and must be handed back to
        /api/merge3-apply, which refuses to write over a report that moved in
        the meantime.

        A package missing so much of the common ancestor that it looks cut off
        is refused with 409 {"error": "truncated", truncated: {...}} -- the body
        carries what looks missing so the UI can say so and offer the override.
        Re-post with allowBulkDelete: true to merge a genuine bulk deletion.
        """
        body = self._read_json()
        project_dir = resolve_project_dir(body.get("dir"), create=False)
        mine = _read_json_quiet(os.path.join(project_dir, "project.json")) or {}
        base = _read_json_quiet(os.path.join(project_dir, "_baseline.json"))
        if base is None:
            base = {}       # never exchanged: every difference is a new change
        theirs = body.get("incoming")
        if not isinstance(theirs, dict):
            return self._send_error_json("missing 'incoming' project object")
        allow_bulk = bool(body.get("allowBulkDelete"))
        payload, status = _lazy_call(
            ("sync.merge3", "merge3"), "merge3", (base, mine, theirs, allow_bulk),
            {"merged": None, "conflicts": [], "auto": 0, "pending": 0,
             "deletions": [], "token": None, "truncated": None})
        if status != 200 or payload.get("stub"):
            return self._send_json(payload, status=status)
        if payload.get("truncated"):
            return self._send_json(_truncation_refusal(payload), status=409)
        if payload.get("error"):
            return self._send_error_json(payload["error"], status=400)
        out = dict(payload)
        # One name on the wire for the concurrency fingerprint.
        out["token"] = payload.get("token") or payload.get("base_sha")
        out.pop("base_sha", None)
        out.setdefault("auto", 0)
        out.setdefault("pending", len(payload.get("conflicts") or []))
        out.setdefault("deletions", [])
        out["truncated"] = None
        return self._send_json(out)

    def _api_merge3_apply(self):
        """Write the outcome of a three-way merge.

        Body {dir, merged, choices, token} -> 200 {ok, applied, snapshot}.
        ``token`` is the one /api/merge3 returned. A refused apply writes
        NOTHING -- no snapshot, no project.json, no baseline -- so every
        validation runs before anything touches the disk and the snapshot is
        taken inside apply_choices once they all pass.

        400 {"error": "merge_token_missing"} when the caller sent no token;
        409 {"error": "stale_merge", expected, actual} when the report changed
        while the user was deciding, so applying would erase that change.
        """
        body = self._read_json()
        project_dir = resolve_project_dir(body.get("dir"), create=False)
        merged = body.get("merged")
        choices = body.get("choices") or []
        token = body.get("token")
        if not isinstance(merged, dict) or "outline" not in merged:
            return self._send_error_json(
                "missing 'merged' report from /api/merge3", status=400)
        if not token:
            return self._send_json({"error": "merge_token_missing"}, status=400)
        payload, status = _lazy_call(("sync.merge3", "merge3"), "apply_choices",
                                     (project_dir, merged, choices, token),
                                     {"ok": False})
        if status != 200 or payload.get("stub"):
            return self._send_json(payload, status=status)
        err = payload.get("error")
        if err == "stale_merge":
            return self._send_json(
                {"error": "stale_merge",
                 "expected": payload.get("mergeBase") or token,
                 "actual": payload.get("localBase")}, status=409)
        if err == "merge_token_missing":
            return self._send_json({"error": "merge_token_missing"}, status=400)
        if err:
            return self._send_error_json(err, status=400)
        out = dict(payload)
        out["ok"] = True
        out.setdefault("applied", 0)
        out.setdefault("snapshot", None)
        return self._send_json(out)

    def _api_assets(self, qs):
        """The report's image pool. ?dir= -> {"assets":[...]}."""
        project_dir = resolve_project_dir((qs.get("dir") or [None])[0],
                                          create=False)
        project = _read_json_quiet(os.path.join(project_dir, "project.json")) or {}
        payload, status = _lazy_call(("store.assets_store", "assets_store"),
                                     "list_assets", (project_dir, project),
                                     {"assets": []})
        return self._send_json(payload, status=status)

    def _api_asset_tag(self):
        """Set the tags of one stored file. Body {dir, file, tags}."""
        body = self._read_json()
        project_dir = resolve_project_dir(body.get("dir"), create=False)
        file_name = body.get("file")
        if not file_name:
            return self._send_error_json("missing 'file'")
        tags = body.get("tags") or []
        payload, status = _lazy_call(("store.assets_store", "assets_store"),
                                     "set_tags", (project_dir, file_name, tags),
                                     {"ok": False, "tags": []})
        return self._send_json(payload, status=status)

    def _api_asset_rename(self):
        """Rename a stored file and repoint every block that uses it.

        Body {dir, from, to} -> {"renamed": n}. The project is snapshotted first
        because the rename rewrites project.json.
        """
        body = self._read_json()
        project_dir = resolve_project_dir(body.get("dir"), create=False)
        old = body.get("from") or body.get("old")
        new = body.get("to") or body.get("new")
        if not old or not new:
            return self._send_error_json("missing 'from' / 'to'")
        pj = os.path.join(project_dir, "project.json")
        project = _read_json_quiet(pj) or {}
        autosave_snapshot(project_dir, "prerename")
        payload, status = _lazy_call(("store.assets_store", "assets_store"),
                                     "rename", (project_dir, project, old, new),
                                     {"renamed": 0})
        if status == 200 and not payload.get("stub"):
            # the module hands back the rewritten project (or mutates the one it
            # was given); persist whichever it is, then re-count the flags.
            updated = payload.pop("project", None)
            if not isinstance(updated, dict):
                updated = project
            atomic_write(pj, json.dumps(updated, ensure_ascii=False,
                                        indent=2).encode("utf-8"))
        return self._send_json(payload, status=status)

    def _api_asset_delete(self):
        """Delete a stored file, but only when nothing in the report points at it.

        Body {dir, file} -> {"ok": true, "file": "images/<name>", "trashed": ...}.

        The refusal is the point of the endpoint, and it is decided HERE rather
        than in the browser: usage is derived from this report's project.json, so
        a client holding a different report's document -- which is exactly what a
        report switch produces for a moment -- cannot talk the server into
        removing a figure that is in use. assets_store.delete raises ValueError
        naming the sections that still use the file, which the dispatcher turns
        into a 400 carrying that sentence.

        Nothing is unlinked: the file moves to <report>/_trash/assets/, the same
        way a deleted report moves to the reports root's trash.
        """
        body = self._read_json()
        project_dir = resolve_project_dir(body.get("dir"), create=False)
        file_name = body.get("file")
        if not file_name:
            return self._send_error_json("missing 'file'")
        project = _read_json_quiet(os.path.join(project_dir, "project.json")) or {}
        payload, status = _lazy_call(("store.assets_store", "assets_store"),
                                     "delete", (project_dir, project, file_name),
                                     {"ok": False})
        return self._send_json(payload, status=status)

    def _api_open_path(self):
        """Open a file, or the folder holding it, with the OS handler.

        Body {dir, file, folder}. `file` is relative to the report folder (e.g.
        "out/report.docx"); the resolved path must stay under the reports root,
        so this can never be turned into an arbitrary "open anything" call.
        """
        body = self._read_json()
        project_dir = resolve_project_dir(body.get("dir"), create=False)
        rel = (body.get("file") or "").replace("\\", "/").lstrip("/")
        if ".." in rel.split("/"):
            return self._send_error_json("forbidden", status=403)
        target = os.path.abspath(os.path.join(project_dir, *rel.split("/"))
                                 if rel else project_dir)
        root_abs = os.path.abspath(CFG.reports_root)
        try:
            common = os.path.commonpath([root_abs, target])
        except ValueError:
            return self._send_error_json("forbidden", status=403)
        if os.path.normcase(common) != os.path.normcase(root_abs):
            return self._send_error_json("forbidden", status=403)
        if body.get("folder"):
            target = target if os.path.isdir(target) else os.path.dirname(target)
        if not os.path.exists(target):
            return self._send_error_json("not found", status=404)
        ok = _open_with_os(target)
        return self._send_json({"ok": ok, "path": target.replace("\\", "/")},
                               status=200 if ok else 500)


def _import_apply_update():
    """Import the shared apply_update module (builder/sync, shared with the CLI).

    Reload on every call so a long-running server picks up an updated
    apply_update.py after a ``git pull`` WITHOUT needing a restart (Python caches
    modules in sys.modules; a running server would otherwise keep the stale code).
    """
    import importlib
    import apply_update  # noqa: E402  (builder/sync, on sys.path via buildpath)
    return importlib.reload(apply_update)


def _rmtree_quiet(path):
    import shutil
    try:
        shutil.rmtree(path)
    except OSError:
        pass


def _guess_content_type(path):
    ext = os.path.splitext(path)[1].lower()
    return {
        ".html": "text/html; charset=utf-8",
        # ES modules must be served as JavaScript or the browser refuses them.
        ".js": "text/javascript; charset=utf-8",
        ".mjs": "text/javascript; charset=utf-8",
        ".css": "text/css; charset=utf-8",
        ".map": "application/json; charset=utf-8",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".ico": "image/x-icon",
        ".svg": "image/svg+xml",
        ".json": "application/json; charset=utf-8",
        ".txt": "text/plain; charset=utf-8",
        ".md": "text/plain; charset=utf-8",
        ".woff2": "font/woff2",
        ".woff": "font/woff",
        ".ttf": "font/ttf",
        ".otf": "font/otf",
    }.get(ext, "application/octet-stream")


# ---------------------------------------------------------------------------
# Entry point.
# ---------------------------------------------------------------------------


def make_server(port=None, root=None, config_path=None, bind="127.0.0.1"):
    """Construct (but do not start) the HTTP server. Used by smoke tests too."""
    if port is not None:
        CFG.port = port
    if root is not None:
        CFG.reports_root = os.path.abspath(root)
    if config_path is not None:
        CFG.template_config_path = os.path.abspath(config_path)
    CFG.bind = bind
    httpd = ThreadingHTTPServer((bind, CFG.port), Handler)
    return httpd


def default_config_path(local_dir=None, warn=None):
    """Best-effort default template config path.

    Prefer an explicit --config / BUILDER_TEMPLATE_CONFIG. Otherwise look in a
    sibling `local/` folder for a `template_config_*.json` and use it.
    Returns None if nothing is found (the server still starts; endpoints that
    need the config then report a clean error).

    A SINGLE match is a fact; several are a guess, and a silent guess is the
    one that costs an afternoon -- the launcher passes no --config, so every
    page would be styled and captioned by whichever file happens to sort first
    with nothing on screen to say so. When there is more than one, name them
    all, say which was taken, and say how to take another.

    ``local_dir`` and ``warn`` exist so this is testable without a real local
    folder; the defaults are the shipped behaviour.
    """
    import glob

    if local_dir is None:
        local_dir = os.path.join(HERE, "..", "..", "local")
    local_dir = os.path.abspath(local_dir)
    pattern = os.path.join(local_dir, "template_config_*.json")
    matches = sorted(glob.glob(pattern))
    if not matches:
        return None
    chosen = matches[0]
    if len(matches) > 1:
        say = warn if warn is not None else sys.stderr.write
        say("WARNING: %d template configs match %s; one had to be picked.\n"
            % (len(matches), pattern))
        for m in matches:
            say("  %s %s\n" % ("chosen ->" if m == chosen else "         ", m))
        say("Pass --config <path> or set BUILDER_TEMPLATE_CONFIG to pick "
            "another one.\n")
    return chosen


def default_reports_root():
    """Default reports_root: a sibling ``local/`` folder if it exists, else None.

    Lets ``python builder/server.py`` (or a start script) Just Work without
    ``--root`` in the common layout where reports live under ``<repo>/local``.
    """
    cand = os.path.abspath(os.path.join(HERE, "..", "..", "local"))
    return cand if os.path.isdir(cand) else None


def main(argv=None):
    parser = argparse.ArgumentParser(description="Structured document builder server")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument(
        "--root",
        default=os.environ.get("BUILDER_REPORTS_ROOT") or default_reports_root(),
        help="reports_root: project folders must stay under this path",
    )
    parser.add_argument(
        "--config",
        default=os.environ.get("BUILDER_TEMPLATE_CONFIG") or default_config_path(),
        help="path to the template config JSON",
    )
    parser.add_argument("--bind", default="127.0.0.1")
    parser.add_argument(
        "--open", dest="open_browser", action="store_true",
        help="open the app in the default browser once the server is up",
    )
    args = parser.parse_args(argv)

    CFG.reports_root = os.path.abspath(args.root) if args.root else None
    CFG.template_config_path = os.path.abspath(args.config) if args.config else None
    CFG.bind = args.bind
    CFG.port = args.port

    if args.bind != "127.0.0.1":
        sys.stderr.write(
            "WARNING: binding to %s; this server is intended for 127.0.0.1 only.\n"
            % args.bind
        )

    httpd = ThreadingHTTPServer((args.bind, args.port), Handler)
    sys.stderr.write(
        "Document builder server on http://%s:%d (root=%s, config=%s)\n"
        % (args.bind, args.port, CFG.reports_root, CFG.template_config_path)
    )
    if args.open_browser:
        import webbrowser
        import threading
        url = "http://127.0.0.1:%d/" % args.port
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        httpd.server_close()


if __name__ == "__main__":
    main()
