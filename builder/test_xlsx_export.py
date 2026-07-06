#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Regression tests for xlsx_export: a table/datatable block -> a .xlsx that
visually mirrors the Word table. Builds bytes, re-parses with openpyxl, and
asserts the replica: out-of-spec cells RED+bold, in-spec plain, category vertical
merge, 3-row header band with the config fill, sim_span horizontal merge, and free
tables (rich runs flattened + merges). Reuses the golden fixture's config/data so
the compliance structure is exercised end-to-end. Neutral / ASCII-only.
"""
import io
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import openpyxl                       # noqa: E402
import xlsx_export as X               # noqa: E402
import test_render_golden as G        # noqa: E402

_fails = []


def check(cond, name):
    print(("  PASS  " if cond else "  FAIL  ") + name)
    if not cond:
        _fails.append(name)


def _load(b):
    return openpyxl.load_workbook(io.BytesIO(b)).active


def _find(ws, val):
    for row in ws.iter_rows():
        for c in row:
            if c.value == val:
                return c
    return None


def _is_red(cell):
    col = cell.font.color
    return bool(col and getattr(col, "rgb", None) and str(col.rgb).upper().endswith("FF0000"))


def main():
    comp = G.golden_config()["compliance"]
    data = G.golden_project()["outline"][0]["blocks"][2]["data"]   # the datatable
    ws = _load(X.build_datatable_xlsx(data, comp))

    print("== compliance datatable ==")
    c758 = _find(ws, 758)   # I_total sim MAX 758 > spec MAX 500 -> flagged
    check(c758 is not None and _is_red(c758) and c758.font.bold,
          "over-spec 758 is RED + bold")
    c490 = _find(ws, 490)
    check(c490 is not None and not _is_red(c490), "in-spec 490 is not red")
    c1200 = _find(ws, 1200)   # P_static NTWC 1200 > spec_ntwc 1000 -> flagged
    check(c1200 is not None and _is_red(c1200), "over-spec NTWC 1200 is RED")

    hvals = set()
    for r in ws.iter_rows(min_row=1, max_row=3):
        for c in r:
            if isinstance(c.value, str):
                hvals.add(c.value)
    check({"Category", "Item", "Unit"} <= hvals, "header labels Category/Item/Unit present")
    check("Pilot" in hvals and "Spec" in hvals, "group titles Spec + Pilot in header band")

    cat_merges = [m for m in ws.merged_cells.ranges
                  if m.min_col == 1 and m.max_col == 1 and m.max_row > m.min_row]
    check(cat_merges and "Power" in {ws.cell(row=m.min_row, column=1).value for m in cat_merges},
          "category column has a vertical merge holding 'Power'")

    a1 = ws.cell(row=1, column=1)
    check(a1.fill and a1.fill.fgColor and
          str(a1.fill.fgColor.rgb).upper().endswith(comp["fills"]["header"].upper()),
          "header band uses the config header fill")

    print("== sim_span ==")
    span_data = {"spec_name": "Spec", "show_spec": False,
                 "sims": [{"key": "s1", "title": "S1", "stage": "Post", "axes": ["MIN", "TYP", "MAX"]}],
                 "rows": [{"cat": "C", "item": "single", "kind": "result", "unit": "mV",
                           "sim_mtm": ["12.3", "12.3", "12.3"], "sim_span": True, "limit": None}]}
    wss = _load(X.build_datatable_xlsx(span_data, comp))
    horiz = [m for m in wss.merged_cells.ranges
             if m.max_col > m.min_col and m.min_row == m.max_row and m.min_row > 3]
    check(len(horiz) >= 1, "sim_span merges the sim MIN/TYP/MAX cells horizontally")

    print("== free table ==")
    rows = [["Style", "Example"],
            ["bold+red", {"runs": [{"t": "Bold", "b": True}, {"t": " red", "color": "FF0000"}]}],
            ["plain", "hi"]]
    wf = _load(X.build_free_table_xlsx(rows, header_rows=1,
                                       merges=[{"r": 0, "c": 0, "rs": 1, "cs": 2}]))
    check(wf.cell(row=1, column=1).value == "Style" and wf.cell(row=1, column=1).font.bold,
          "header row is bold")
    check(wf.cell(row=2, column=2).value == "Bold red" and wf.cell(row=2, column=2).font.bold,
          "rich runs flattened (text + bold)")
    check(any(m.min_row == 1 and m.max_col == 2 for m in wf.merged_cells.ranges),
          "header merge applied")

    print("== dispatch / filename ==")
    check(X.filename_for({"type": "datatable", "caption": "Compliance results!"})
          == "Compliance_results_.xlsx", "filename_for slugs the caption")
    try:
        X.build_block_xlsx({"type": "para"}, {})
        check(False, "non-table block raises ValueError")
    except ValueError:
        check(True, "non-table block raises ValueError")

    print("\n== SUMMARY ==  %s" % ("ALL PASSED" if not _fails else "%d FAILED" % len(_fails)))
    return 1 if _fails else 0


if __name__ == "__main__":
    sys.exit(main())
