#!/usr/bin/env python3
"""Local HTTP server for the structured document builder.

Standard-library only (http.server / json / urllib / base64 / sqlite3).
Binds to 127.0.0.1 exclusively. Serves the single-page app and the JSON API
defined in CONTRACT.md. All domain-specific content (chapter skeleton, cover
fields, fixed body texts, style numbers, validation rules, UI strings, logo)
lives in an external template config file loaded at runtime; this module stays
neutral and domain-agnostic.

Run:
    python server.py --port 8765 --root <reports_root> --config <template.json>
"""

import argparse
import base64
import contextlib
import copy
import datetime
import io
import json
import os
import re
import shutil
import sys
import tempfile
import threading
import traceback
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs

HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path:
    sys.path.insert(0, HERE)

import templates_store as tstore  # builder/ is on sys.path via HERE  # noqa: E402

# ---------------------------------------------------------------------------
# Server configuration (populated in main()).
# ---------------------------------------------------------------------------


class Config:
    reports_root = None          # absolute path; project dirs must stay under it
    template_config_path = None  # path to the active template config JSON
    bind = "127.0.0.1"
    port = 8765
    _template_cache = None       # (path, mtime) -> parsed dict


CFG = Config()


# ---------------------------------------------------------------------------
# Helpers: template config loading.
# ---------------------------------------------------------------------------


def load_template_config(path=None):
    """Load and cache the template config JSON. Returns a dict or raises."""
    p = path or CFG.template_config_path
    if not p:
        raise FileNotFoundError("no template config path configured")
    p = os.path.abspath(p)
    if not os.path.isfile(p):
        raise FileNotFoundError("template config not found: %s" % p)
    mtime = os.path.getmtime(p)
    cache = CFG._template_cache
    if cache and cache[0] == p and cache[1] == mtime:
        return cache[2]
    with open(p, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    CFG._template_cache = (p, mtime, data)
    return data


def config_slice(template_query=None):
    """Frontend-facing slice of the template config (API #3).

    Returns only what the GUI needs; never leaks engine-only style numbers. When
    a library template id is given and exists, slice THAT template's config;
    otherwise fall back to the global template config.
    """
    tc = None
    if template_query:
        p = tstore.template_config_path(CFG.reports_root, template_query)
        if p:
            tc = load_template_config(p)
    if tc is None:
        tc = load_template_config()
    return {
        "template": tc.get("id"),
        "skeleton": tc.get("skeleton", []),
        "ui_strings": tc.get("ui_strings", {}),
        "cover": tc.get("cover", {}),
        "compliance": _compliance_defaults(tc.get("compliance", {})),
        # table_presets: company-defined starter tables (with condition rows
        # baked in). Content lives in the local template config; the engine/app
        # stay neutral -- this just forwards the list to the block picker (WS4).
        "table_presets": tc.get("table_presets", []),
    }


def _compliance_defaults(comp):
    """Expose only the compliance defaults the GUI needs for live preview."""
    return {
        "axis_labels": comp.get("axis_labels", ["MIN", "TYP", "MAX", "NTWC"]),
        "default_limit": comp.get("default_limit", {}),
        "flag_color": comp.get("flag_color", "FF0000"),
        "setting_kinds": comp.get(
            "setting_kinds", ["common_setting", "module_setting", "tb"]
        ),
    }


# ---------------------------------------------------------------------------
# Helpers: path safety.
# ---------------------------------------------------------------------------


def resolve_project_dir(dir_arg, create=False):
    """Resolve a project dir from the `dir` query param.

    `dir` may be absolute or relative to reports_root. The resolved path must
    stay under reports_root (if one is configured). Raises ValueError on any
    traversal outside the allowed root.
    """
    if not dir_arg:
        raise ValueError("missing 'dir' parameter")
    root = CFG.reports_root
    # A6: fail CLOSED when no reports root is configured. Previously a missing
    # root skipped the containment check entirely (fail-open), so any absolute /
    # relative path resolved unchecked. Mirror _api_apply_update's stance.
    if not root:
        raise ValueError("no reports root configured")
    if os.path.isabs(dir_arg):
        target = os.path.abspath(dir_arg)
    else:
        target = os.path.abspath(os.path.join(root, dir_arg))

    root_abs = os.path.abspath(root)
    # containment check that is robust to case/sep on Windows
    try:
        common = os.path.commonpath([root_abs, target])
    except ValueError:
        # different drives
        raise ValueError("path escapes reports_root")
    if os.path.normcase(common) != os.path.normcase(root_abs):
        raise ValueError("path escapes reports_root")

    if create:
        os.makedirs(target, exist_ok=True)
        os.makedirs(os.path.join(target, "images"), exist_ok=True)
    return target


def atomic_write(path, data_bytes):
    """Write bytes to path atomically (temp file + os.replace)."""
    d = os.path.dirname(path)
    os.makedirs(d, exist_ok=True)
    tmp = path + ".tmp.%d" % os.getpid()
    with open(tmp, "wb") as fh:
        fh.write(data_bytes)
        fh.flush()
        os.fsync(fh.fileno())
    os.replace(tmp, path)


# ---------------------------------------------------------------------------
# Helpers: local auto-snapshots (belt-and-suspenders backup of project.json).
#
# Every save -- and every overwrite by apply-update / paste-import -- first drops
# a timestamped copy of project.json into <project_dir>/_autosave/, kept rolling.
# This is INDEPENDENT of the apply-time _backups/ history, so an external update
# can never leave local edits unrecoverable: the last saved state is always
# sitting in _autosave/ and is one click to restore.
# ---------------------------------------------------------------------------

AUTOSAVE_DIRNAME = "_autosave"
AUTOSAVE_KEEP = 300  # rolling retention per project

# Root-level bookkeeping dirs that are never valid as a project folder name.
_RESERVED_ROOT_DIRS = {"_backups", "_updates", "_outbox", "_trash",
                       "_autosave", "__pycache__", "assets"}


def _autosave_dir(project_dir):
    return os.path.join(project_dir, AUTOSAVE_DIRNAME)


def _autosave_paths(project_dir):
    """Snapshot paths in chronological order (oldest first). Sorted by mtime
    (tie-break by name) so same-second collision suffixes can't scramble order."""
    d = _autosave_dir(project_dir)
    if not os.path.isdir(d):
        return []
    entries = []
    for n in os.listdir(d):
        if not n.endswith(".json"):
            continue
        p = os.path.join(d, n)
        try:
            mt = os.path.getmtime(p)
        except OSError:
            mt = 0.0
        entries.append((mt, n, p))
    entries.sort()
    return [p for _, _, p in entries]


def autosave_snapshot(project_dir, reason="save"):
    """Copy the current project.json into _autosave/<ts>__<reason>.json.

    Skips if identical to the newest snapshot (dedupe). Prunes to the newest
    AUTOSAVE_KEEP. Best-effort: never raises -- a snapshot must never break a
    save. Returns the snapshot filename or None.
    """
    try:
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return None
        with open(pj, "rb") as fh:
            data = fh.read()
        existing = _autosave_paths(project_dir)
        if existing:
            try:
                with open(existing[-1], "rb") as fh:
                    if fh.read() == data:
                        return None  # unchanged since last snapshot
            except OSError:
                pass
        os.makedirs(_autosave_dir(project_dir), exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        tag = _sanitize_name(reason) or "save"
        base = "%s__%s" % (ts, tag)
        name = base + ".json"
        k = 1
        while os.path.exists(os.path.join(_autosave_dir(project_dir), name)):
            name = "%s-%d.json" % (base, k)
            k += 1
        atomic_write(os.path.join(_autosave_dir(project_dir), name), data)
        for p in _autosave_paths(project_dir)[:-AUTOSAVE_KEEP]:
            try:
                os.remove(p)
            except OSError:
                pass
        return name
    except Exception:
        return None


def autosave_all(reports_root, reason="preapply"):
    """Snapshot every project under reports_root (used before an apply)."""
    made = []
    try:
        for name in sorted(os.listdir(reports_root)):
            pdir = os.path.join(reports_root, name)
            if os.path.isfile(os.path.join(pdir, "project.json")):
                snap = autosave_snapshot(pdir, reason)
                if snap:
                    made.append("%s/%s" % (name, snap))
    except OSError:
        pass
    return made


def list_autosaves(project_dir):
    out = []
    for p in _autosave_paths(project_dir):
        name = os.path.basename(p)
        try:
            st = os.stat(p)
        except OSError:
            continue
        stem = name[:-5]
        reason = stem.split("__", 1)[1].split("-")[0] if "__" in stem else ""
        out.append({"name": name, "mtime": st.st_mtime,
                    "size": st.st_size, "reason": reason})
    out.reverse()  # newest first
    return out


def restore_autosave(project_dir, name):
    """Restore _autosave/<name> to project.json (snapshotting current first)."""
    if not name or "/" in name or "\\" in name or not name.endswith(".json"):
        raise ValueError("bad snapshot name")
    src = os.path.join(_autosave_dir(project_dir), name)
    if not os.path.isfile(src):
        raise FileNotFoundError("snapshot not found: %s" % name)
    with open(src, "rb") as fh:
        data = fh.read()
    json.loads(data.decode("utf-8"))  # validate JSON before installing
    autosave_snapshot(project_dir, reason="prerestore")  # keep the restore undoable
    pj = os.path.join(project_dir, "project.json")
    atomic_write(pj, data)
    return {"ok": True, "restored": name, "saved_at": os.path.getmtime(pj)}


# ---------------------------------------------------------------------------
# Helpers: image saving.
# ---------------------------------------------------------------------------

_NAME_RE = re.compile(r"[^A-Za-z0-9_-]+")


def _sanitize_name(name):
    name = _NAME_RE.sub("_", name).strip("_")
    return name[:64] if name else ""


def next_image_path(project_dir, section, name=None):
    """Compute images/<section>_<seq>.png (or images/<section>-<seq>_<name>.png).

    <seq> is the next free sequence for that section. Returns (rel, abs).
    """
    section = _sanitize_name(str(section)) or "0"
    name = _sanitize_name(name) if name else ""
    images_dir = os.path.join(project_dir, "images")
    os.makedirs(images_dir, exist_ok=True)

    # find existing seqs for this section in either naming form
    seqs = []
    pat = re.compile(
        r"^%s[-_](\d+)(?:_.*)?\.png$" % re.escape(section), re.IGNORECASE
    )
    for fn in os.listdir(images_dir):
        m = pat.match(fn)
        if m:
            try:
                seqs.append(int(m.group(1)))
            except ValueError:
                pass
    seq = (max(seqs) + 1) if seqs else 1

    if name:
        rel = "images/%s-%d_%s.png" % (section, seq, name)
    else:
        rel = "images/%s_%d.png" % (section, seq)
    return rel, os.path.join(project_dir, rel)


# ---------------------------------------------------------------------------
# Helpers: xlsx parsing (free-table grid + compliance data).
# ---------------------------------------------------------------------------


def parse_xlsx_grid(xlsx_bytes):
    """Parse an xlsx into a free-table grid: {rows:[[..]], merges:[..]}.

    Reads the first worksheet's used range. Cell values are stringified;
    None -> "". Merged ranges are reported as {r,c,rs,cs} (0-based, top-left).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0

    rows = []
    for r in range(1, max_r + 1):
        row = []
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            row.append("" if v is None else str(v))
        rows.append(row)

    merges = []
    for rng in ws.merged_cells.ranges:
        merges.append(
            {
                "r": rng.min_row - 1,
                "c": rng.min_col - 1,
                "rs": rng.max_row - rng.min_row + 1,
                "cs": rng.max_col - rng.min_col + 1,
            }
        )
    return {"rows": rows, "merges": merges}


def _to_num_or_str(v):
    """Coerce a cell value: numbers stay numeric, blanks -> None, else str."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


def parse_xlsx_compliance(xlsx_bytes):
    """Best-effort parse of an xlsx into the compliance DATA model.

    The layout mirrors the renderer's output: a yellow 3-row header band, then
    Category / Item / [Spec] / <group axis columns> / Unit. This is a seed for
    the GUI editor, not a strict importer; the GUI lets the user fix anything.

    Returns {data: {spec_name, sims, rows}}.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append(list(row))
    if not grid:
        return {"recognized": False, "data": {"spec_name": "", "sims": [], "rows": []}}

    ncols = max(len(r) for r in grid)
    for r in grid:
        r.extend([None] * (ncols - len(r)))

    # Locate header band: the first row containing an axis label cell.
    AX = {"MIN", "TYP", "MAX", "NTWC", "NT WC", "NT-WC"}
    header_axis_row = None
    for ri, r in enumerate(grid[:10]):
        if any(isinstance(c, str) and c.strip().upper() in AX for c in r):
            header_axis_row = ri
            break
    if header_axis_row is None:
        # Fall back: treat as plain grid, no compliance structure recognized.
        return {"recognized": False, "data": {"spec_name": "", "sims": [], "rows": []}}

    # Map header columns.
    axis_cols = []  # list of (col_index, axis_label)
    cat_col = item_col = spec_col = unit_col = None
    band = grid[header_axis_row]
    for ci, c in enumerate(band):
        if not isinstance(c, str):
            continue
        cu = c.strip().upper()
        if cu in AX:
            axis_cols.append((ci, "NTWC" if cu.replace(" ", "").replace("-", "") == "NTWC" else cu))

    # Category/Item/Unit/Spec live in the merged label columns above the axes:
    # scan rows 0..header_axis_row for the literal labels.
    for ri in range(0, header_axis_row + 1):
        for ci, c in enumerate(grid[ri]):
            if not isinstance(c, str):
                continue
            cl = c.strip().lower()
            if cl == "category" and cat_col is None:
                cat_col = ci
            elif cl == "item" and item_col is None:
                item_col = ci
            elif cl == "spec" and spec_col is None:
                spec_col = ci
            elif cl == "unit" and unit_col is None:
                unit_col = ci

    # Group axis columns into runs of 3-4 consecutive axis columns.
    spec_name = ""
    sims = []
    spec_axis_cols = None
    groups = []  # (title, stage, [cols])
    if axis_cols:
        # Split into groups. A new group starts when columns are non-adjacent
        # (a spacer between groups) OR when an axis label repeats within the
        # current run (e.g. ...MAX, NTWC, MIN... => the second MIN begins a new
        # group even with no spacer column).
        runs = []
        cur = [axis_cols[0]]
        seen = {axis_cols[0][1]}
        for prev, nxt in zip(axis_cols, axis_cols[1:]):
            adjacent = (nxt[0] - prev[0] == 1)
            repeats = nxt[1] in seen
            if adjacent and not repeats:
                cur.append(nxt)
                seen.add(nxt[1])
            else:
                runs.append(cur)
                cur = [nxt]
                seen = {nxt[1]}
        runs.append(cur)

        # Group titles sit in the row above header_axis_row (the merged band top).
        title_row = grid[header_axis_row - 2] if header_axis_row >= 2 else grid[0]
        stage_row = grid[header_axis_row - 1] if header_axis_row >= 1 else grid[0]
        for gi, run in enumerate(runs):
            c0 = run[0][0]
            title = _first_str(title_row, run) or ""
            stage = _first_str(stage_row, run)
            cols = [rc[0] for rc in run]
            if gi == 0 and "spec" in (title or "").lower():
                spec_name = title
                spec_axis_cols = cols
            elif gi == 0 and stage is None and "spec" in (title or "").lower():
                spec_name = title
                spec_axis_cols = cols
            else:
                groups.append((title, stage, cols))
        # If first run wasn't recognized as spec by name, assume it is spec.
        if spec_axis_cols is None and runs:
            spec_axis_cols = [rc[0] for rc in runs[0]]
            spec_name = _first_str(title_row, runs[0]) or spec_name
            groups = []
            for gi, run in enumerate(runs[1:]):
                title = _first_str(title_row, run) or "Sim%d" % (gi + 1)
                stage = _first_str(stage_row, run)
                groups.append((title, stage, [rc[0] for rc in run]))

    for gi, (title, stage, cols) in enumerate(groups):
        sims.append(
            {
                "key": _sanitize_name(title).lower() or ("sim%d" % (gi + 1)),
                "title": title or ("Sim%d" % (gi + 1)),
                "stage": stage,
            }
        )

    # Data rows start after the header band.
    rows = []
    last_cat = ""
    for ri in range(header_axis_row + 1, len(grid)):
        r = grid[ri]
        item = _to_num_or_str(r[item_col]) if item_col is not None else None
        cat = _to_num_or_str(r[cat_col]) if cat_col is not None else None
        if cat:
            last_cat = str(cat)
        if item is None or item == "":
            continue  # skip separator / blank rows
        unit = _to_num_or_str(r[unit_col]) if unit_col is not None else None
        spec = _to_num_or_str(r[spec_col]) if spec_col is not None else None

        spec_mtm = [None, None, None]
        spec_ntwc = None
        if spec_axis_cols:
            for ai, ci in enumerate(spec_axis_cols[:4]):
                val = _to_num_or_str(r[ci]) if ci < len(r) else None
                if ai < 3:
                    spec_mtm[ai] = val
                else:
                    spec_ntwc = val

        sim_mtm = [None, None, None]
        sim_ntwc = None
        if groups:
            first_cols = groups[0][2]
            for ai, ci in enumerate(first_cols[:4]):
                val = _to_num_or_str(r[ci]) if ci < len(r) else None
                if ai < 3:
                    sim_mtm[ai] = val
                else:
                    sim_ntwc = val

        rows.append(
            {
                "cat": last_cat,
                "item": str(item),
                "unit": "" if unit is None else str(unit),
                "kind": "result",
                "spec": spec,
                "spec_mtm": spec_mtm,
                "sim_mtm": sim_mtm,
                "spec_ntwc": spec_ntwc,
                "sim_ntwc": sim_ntwc,
                "limit": None,
                "sim_span": False,
            }
        )

    return {"recognized": True, "data": {"spec_name": spec_name, "sims": sims, "rows": rows}}


def _first_str(row, run):
    """Return the first non-empty string within the given run's columns."""
    for ci, _ in run:
        if ci < len(row) and isinstance(row[ci], str) and row[ci].strip():
            return row[ci].strip()
    return None


# ---------------------------------------------------------------------------
# Helpers: compliance validation (reuses engine flag logic; never duplicated).
# ---------------------------------------------------------------------------


def _import_flag_positions():
    """Lazily import the engine's flag_positions. Raises on unavailability."""
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    try:
        from tables import flag_positions  # type: ignore
        return flag_positions
    except Exception:
        # engine module may expose it directly
        from engine import flag_positions  # type: ignore
        return flag_positions


def validate_compliance(data):
    """Run engine flag logic over compliance rows.

    Returns {"flags": {"<rowIndex>": [axisIndex,...]}, "color": flag_color}.
    Axis indices are positions within the sim group's [MIN, TYP, MAX] axes.
    """
    flag_positions = _import_flag_positions()
    rows = (data or {}).get("rows", [])
    flags = {}
    for i, row in enumerate(rows):
        positions = flag_positions(row)
        if positions:
            flags[str(i)] = sorted(int(p) for p in positions)
    try:
        tc = load_template_config()
        color = tc.get("compliance", {}).get("flag_color", "FF0000")
    except Exception:
        color = "FF0000"
    return {"flags": flags, "color": color}


# ---------------------------------------------------------------------------
# Helpers: export (delegates to the engine).
# ---------------------------------------------------------------------------


def _import_engine():
    # Reload on every export so a `git pull` of engine.py/tables.py takes effect
    # without restarting this long-running server (same rationale as the
    # apply_update reload). engine imports tables at module scope, so tables must
    # be reloaded first, otherwise engine keeps binding the stale module.
    import importlib
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import content_lint  # type: ignore  # engine imports this at module scope
    importlib.reload(content_lint)
    import tables  # type: ignore
    importlib.reload(tables)
    import engine  # type: ignore
    importlib.reload(engine)
    return engine


def _import_xlsx_export():
    # Reload so a git pull of xlsx_export.py / tables.py takes effect without a
    # restart. xlsx_export imports tables at module scope -> reload tables first.
    import importlib
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import tables  # type: ignore
    importlib.reload(tables)
    import xlsx_export  # type: ignore
    return importlib.reload(xlsx_export)


def run_export(project_dir, fmt, save_first=False, on_progress=None, on_phase=None):
    """Render the project via the engine. fmt in {docx, pdf}.

    Returns {out, abs, fmt}. For pdf: render docx then COM-export (Export then
    Close, never a method call after Close). Output lands in <dir>/out/.

    ``on_progress(done, total, label)`` (per heading/block) and ``on_phase(label)``
    ('preparing'|'rendering'|'converting') drive a live progress bar; both optional
    and best-effort (their exceptions never break the export).
    """
    def phase(label):
        if on_phase:
            try:
                on_phase(label)
            except Exception:
                pass

    phase("preparing")
    engine = _import_engine()
    out_dir = os.path.join(project_dir, "out")
    os.makedirs(out_dir, exist_ok=True)

    # Load the project document.
    with open(os.path.join(project_dir, "project.json"), "r", encoding="utf-8") as fh:
        project = json.load(fh)

    # Resolve and load the template config (engine resolves the logo path).
    # A project bound to a library template wins over the global --config.
    tid = project.get("template")
    tpl_cfg = tstore.template_config_path(CFG.reports_root, tid) if tid else None
    explicit = tpl_cfg or CFG.template_config_path
    config_path = engine._resolve_config_path(project, project_dir, explicit)
    cfg = engine._load_config(config_path)

    name = os.path.basename(project_dir.rstrip(os.sep)) or "report"
    docx_out = os.path.join(out_dir, "%s.docx" % name)
    # render_report returns {out_path, warnings, stats}; tolerate the legacy
    # bare-string return too. The render manifest (warnings + stats) is surfaced
    # to the frontend as additive keys on the success payload so the user can see
    # missing images / out-of-spec / clip risks after an export.
    phase("rendering")
    render_result = engine.render_report(project, cfg, project_dir, docx_out,
                                         on_progress=on_progress)
    out_path = engine._result_out_path(render_result) or docx_out
    docx_abs = os.path.abspath(out_path)
    warnings = render_result.get("warnings", []) if isinstance(render_result, dict) else []
    stats = render_result.get("stats", {}) if isinstance(render_result, dict) else {}

    # Merge pre-render content-lint findings into the SAME manifest so the export
    # panel surfaces structural issues (missing condition rows, never-flagging
    # rows, sim_span axis problems, bad image paths, ...) next to the render
    # warnings -- all carrying a level. no_caption is split (engine owns
    # datatable/table, lint owns image/imagegrid) so there is no double-report.
    try:
        import content_lint  # already reloaded by _import_engine
        lint_findings = content_lint.lint_project(project, cfg)
    except Exception:
        lint_findings = []
    if lint_findings:
        warnings = list(lint_findings) + list(warnings)
        stats = dict(stats)
        stats["errors"] = sum(1 for w in warnings if w.get("level") == "error")
        stats["warns"] = sum(1 for w in warnings if w.get("level") == "warn")
        stats["infos"] = sum(1 for w in warnings if w.get("level") == "info")

    if fmt == "docx":
        rel = os.path.relpath(docx_abs, project_dir).replace("\\", "/")
        return {"out": rel, "abs": docx_abs.replace("\\", "/"), "fmt": "docx",
                "warnings": warnings, "stats": stats}

    if fmt == "pdf":
        pdf_abs = os.path.splitext(docx_abs)[0] + ".pdf"
        phase("converting")
        _word_export_pdf(docx_abs, pdf_abs)
        rel = os.path.relpath(pdf_abs, project_dir).replace("\\", "/")
        return {"out": rel, "abs": pdf_abs.replace("\\", "/"), "fmt": "pdf",
                "warnings": warnings, "stats": stats}

    raise ValueError("unknown fmt: %s" % fmt)


def _word_export_pdf(docx_abs, pdf_abs):
    """Word COM: Open -> ExportAsFixedFormat(17) -> Close. Export then Close."""
    import win32com.client  # type: ignore  # pywin32, present on the work machine

    word = win32com.client.DispatchEx("Word.Application")
    word.Visible = False
    doc = None
    try:
        doc = word.Documents.Open(os.path.abspath(docx_abs))
        # Update all fields (e.g. a Table of Contents) so the PDF reflects what
        # the user sees in Word after a field refresh; otherwise TOC/PAGEREF
        # fields render as their unpopulated placeholder and pagination differs.
        try:
            for story in doc.StoryRanges:
                story.Fields.Update()
            for toc in doc.TablesOfContents:
                toc.Update()
        except Exception:
            pass  # field update is best-effort; never block the export
        # 17 = wdExportFormatPDF
        doc.ExportAsFixedFormat(os.path.abspath(pdf_abs), 17)
    finally:
        if doc is not None:
            doc.Close(False)  # never call a method on doc after this
        word.Quit()


# ---------------------------------------------------------------------------
# Helpers: docx import (templates + report seeds).
# ---------------------------------------------------------------------------


def _import_docx_module():
    """Lazily import the docx_import parser (mirrors _import_engine).

    Raises ImportError/ModuleNotFoundError if python-docx is unavailable so the
    route can degrade to a clean 503 rather than failing server import.
    """
    if HERE not in sys.path:
        sys.path.insert(0, HERE)
    import docx_import  # type: ignore
    return docx_import


@contextlib.contextmanager
def _temp_docx(docx_bytes):
    """Materialize uploaded bytes to a temp .docx (docx_import takes a path).

    The temp file is always removed on exit; any images / logo extracted out of
    it are written to their destination dirs before the context closes.
    """
    fd, path = tempfile.mkstemp(suffix=".docx", prefix="import_docx_")
    try:
        with os.fdopen(fd, "wb") as fh:
            fh.write(docx_bytes)
        yield path
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


def _rewrite_image_paths(outline):
    """Guarantee every image block's "file" is project-relative "images/<base>".

    Keeps "src". Idempotent for already-correct parser output; self-heals if a
    future parser change emits an absolute or bare-basename path.
    """
    def walk(nodes):
        for n in nodes:
            for b in n.get("blocks", []):
                if b.get("type") == "image":
                    src = b.get("file") or b.get("src") or ""
                    base = os.path.basename(src.replace("\\", "/"))
                    if base:
                        b["file"] = "images/" + base
                        b.setdefault("src", b["file"])
            walk(n.get("children", []))

    walk(outline or [])
    return outline


def _instantiate_skeleton(skeleton):
    """Clone a skeleton tree into a project outline.

    FIXED nodes carry their blocks verbatim; FILLABLE nodes get an empty blocks
    list (a clean placeholder for the editor). The ``fixed`` marker is dropped --
    a project node is just ``{title, level, blocks, children}``.
    """
    out = []
    for node in skeleton or []:
        is_fixed = bool(node.get("fixed"))
        out.append({
            "title": node.get("title", ""),
            "level": node.get("level", 1),
            "blocks": copy.deepcopy(node.get("blocks", [])) if is_fixed else [],
            "children": _instantiate_skeleton(node.get("children", [])),
        })
    return out


def _empty_meta():
    """Default empty project meta (mirrors parse_docx_report's meta dict)."""
    return {
        "title": "", "secrecy": "", "doc_no": "", "page_count": "",
        "author": "", "reviewers": [], "approver": "", "revisions": [],
    }


# ---------------------------------------------------------------------------
# Helpers: paste-import structural diff (de-LLM upstream channel, WS2).
# ---------------------------------------------------------------------------
#
# The GUI's Copy-text exports the WHOLE App.project. Re-importing it is a full
# replace, so there is no three-way merge: "user version wins" == overwrite +
# backup + a compact diff the assistant can read (structure, not full text).


def _short(v, n=60):
    s = "" if v is None else str(v)
    s = s.replace("\n", " ")
    return s if len(s) <= n else s[:n] + "..."


def _collect_nodes(outline):
    """Every outline node in document order (depth-first, children included)."""
    out = []

    def walk(nodes):
        for n in nodes or []:
            if isinstance(n, dict):
                out.append(n)
                walk(n.get("children", []))

    walk(outline)
    return out


def _block_types(node):
    from collections import Counter
    c = Counter()
    for b in (node.get("blocks") or []):
        if isinstance(b, dict):
            c[b.get("type", "?")] += 1
    return c


def _para_chars(node):
    total = 0
    for b in (node.get("blocks") or []):
        if isinstance(b, dict) and b.get("type") == "para":
            for r in (b.get("runs") or []):
                if isinstance(r, dict):
                    total += len(r.get("t", "") or "")
    return total


def _fmt_multiset_delta(old_c, new_c):
    parts = []
    for k in sorted(set(old_c) | set(new_c)):
        o, n = old_c.get(k, 0), new_c.get(k, 0)
        if o != n:
            parts.append("%s %d->%d" % (k, o, n))
    return ", ".join(parts)


def _match_nodes(old_nodes, new_nodes):
    """Pair old/new outline nodes by title (id as tie-break, robust to node-id
    drift). Returns (pairs, added, removed)."""
    from collections import defaultdict
    by_title = defaultdict(list)
    for n in old_nodes:
        by_title[n.get("title", "")].append(n)
    consumed, pairs, added = set(), [], []
    for nn in new_nodes:
        cands = by_title.get(nn.get("title", ""), [])
        pick = None
        for c in cands:  # prefer same id, not yet consumed
            if id(c) not in consumed and nn.get("id") and c.get("id") == nn.get("id"):
                pick = c
                break
        if pick is None:
            for c in cands:
                if id(c) not in consumed:
                    pick = c
                    break
        if pick is None:
            added.append(nn)
        else:
            consumed.add(id(pick))
            pairs.append((pick, nn))
    removed = [n for n in old_nodes if id(n) not in consumed]
    return pairs, added, removed


def paste_import_diff(old_project, new_project, dir_name, id_warns=None):
    """Compact structural diff between the on-disk project and the pasted one.

    No full text: meta key changes, top-level key changes (e.g. sim_checklist),
    and per-section block-type multiset + para char-length deltas. Returns a
    markdown string (~300-600 tokens) the assistant reads instead of the file."""
    old = old_project if isinstance(old_project, dict) else {}
    new = new_project if isinstance(new_project, dict) else {}
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = ["# Paste-import diff -- %s" % dir_name, "",
             "_Generated %s. Structural summary only (no full text)._" % ts, ""]
    for w in (id_warns or []):
        lines.append("> WARNING: %s" % w)
    if id_warns:
        lines.append("")

    # meta
    om, nm = old.get("meta") or {}, new.get("meta") or {}
    meta_changes = []
    for k in sorted(set(om) | set(nm)):
        ov, nv = om.get(k), nm.get(k)
        if ov == nv:
            continue
        if isinstance(ov, list) or isinstance(nv, list):
            meta_changes.append("%s %d->%d entries" % (k, len(ov or []), len(nv or [])))
        else:
            meta_changes.append("%s %r->%r" % (k, _short(ov), _short(nv)))
    lines += ["## meta",
              ("- " + "; ".join(meta_changes)) if meta_changes else "- (unchanged)", ""]

    # top-level keys besides meta/outline (template + sim_checklist etc.)
    skip = {"meta", "outline", "schema_version"}
    top_changes = []
    for k in sorted(set(old) | set(new)):
        if k in skip:
            continue
        ov, nv = old.get(k), new.get(k)
        if ov == nv:
            continue
        if isinstance(ov, list) or isinstance(nv, list):
            top_changes.append("%s %d->%d items" % (k, len(ov or []), len(nv or [])))
        else:
            top_changes.append("%s %r->%r" % (k, _short(ov), _short(nv)))
    if top_changes:
        lines += ["## top-level", "- " + "; ".join(top_changes), ""]

    # outline sections
    old_nodes = _collect_nodes(old.get("outline"))
    new_nodes = _collect_nodes(new.get("outline"))
    pairs, added, removed = _match_nodes(old_nodes, new_nodes)
    changed = []
    for on, nn in pairs:
        delta = _fmt_multiset_delta(_block_types(on), _block_types(nn))
        opl, npl = _para_chars(on), _para_chars(nn)
        if delta or opl != npl:
            desc = []
            if delta:
                desc.append("blocks " + delta)
            if opl != npl:
                desc.append("para chars %d->%d (%+d)" % (opl, npl, npl - opl))
            changed.append((nn.get("title", ""), "; ".join(desc)))
    lines += ["## sections",
              "- %d total (was %d); %d changed, %d added, %d removed"
              % (len(new_nodes), len(old_nodes), len(changed), len(added), len(removed)), ""]
    if changed:
        lines.append("### changed")
        lines += ['- "%s": %s' % (t, d) for t, d in changed]
        lines.append("")
    if added:
        lines.append("### added")
        lines += ['- "%s" [%s]' % (n.get("title", ""),
                                   _fmt_multiset_delta({}, _block_types(n)) or "empty")
                  for n in added]
        lines.append("")
    if removed:
        lines.append("### removed")
        lines += ['- "%s"' % n.get("title", "") for n in removed]
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


# ---------------------------------------------------------------------------
# HTTP handler.
# ---------------------------------------------------------------------------


class Handler(BaseHTTPRequestHandler):
    server_version = "DocBuilder/1.0"
    protocol_version = "HTTP/1.1"

    # --- low-level response helpers ---

    def _send_json(self, obj, status=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_error_json(self, msg, status=400):
        self._send_json({"error": str(msg)}, status=status)

    def _send_bytes(self, body, content_type, status=200):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return b""
        return self.rfile.read(length)

    def _read_json(self):
        raw = self._read_body()
        if not raw:
            return {}
        return json.loads(raw.decode("utf-8"))

    def log_message(self, fmt, *args):  # quieter logging
        sys.stderr.write("[server] %s - %s\n" % (self.address_string(), fmt % args))

    # --- routing ---

    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)

            if path == "/" or path == "/app.html":
                return self._serve_app_html()
            if path.startswith("/assets/"):
                return self._serve_asset(path)
            if path.startswith("/images/") or path.startswith("images/"):
                return self._serve_project_file(path, qs)
            if path == "/api/config":
                template = (qs.get("template") or [None])[0]
                return self._send_json(config_slice(template))
            if path == "/api/project":
                return self._api_project_get(qs)
            if path == "/api/projects":
                return self._api_projects_list()
            if path == "/api/autosaves":
                return self._api_autosaves(qs)
            if path == "/api/templates":
                return self._api_templates_list()
            if path == "/api/template":
                return self._api_template_get(qs)
            if path == "/api/health":
                return self._send_json({"ok": True})

            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)

    def do_PUT(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)
            if path == "/api/project":
                return self._api_project_put(qs)
            if path == "/api/template":
                return self._api_template_put(qs)
            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)

    def do_DELETE(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)
            if path == "/api/template":
                return self._api_template_delete(qs)
            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)

    def do_POST(self):
        try:
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)

            if path == "/api/image":
                return self._api_image(qs)
            if path == "/api/import-xlsx":
                return self._api_import_xlsx()
            if path == "/api/validate-compliance":
                return self._api_validate_compliance()
            if path == "/api/export":
                return self._api_export(qs)
            if path == "/api/export-stream":
                return self._api_export_stream(qs)
            if path == "/api/export-xlsx":
                return self._api_export_xlsx()
            if path == "/api/import-docx":
                return self._api_import_docx()
            if path == "/api/new-from-template":
                return self._api_new_from_template()
            if path == "/api/paste-import":
                return self._api_paste_import(qs)
            if path == "/api/copy-diff":
                return self._api_copy_diff(qs)
            if path == "/api/apply-update":
                return self._api_apply_update()
            if path == "/api/rollback":
                return self._api_rollback()
            if path == "/api/autosave-restore":
                return self._api_autosave_restore()
            if path == "/api/project-delete":
                return self._api_project_delete()
            if path == "/api/project-rename":
                return self._api_project_rename()
            if path == "/api/project-copy":
                return self._api_project_copy()

            return self._send_error_json("not found: %s" % path, status=404)
        except Exception as exc:
            self._handle_exc(exc)

    def _handle_exc(self, exc):
        # Never crash the serve loop; always return a JSON error.
        if isinstance(exc, ValueError):
            status = 400
        elif isinstance(exc, FileNotFoundError):
            status = 404
        elif isinstance(exc, json.JSONDecodeError):
            status = 400
            exc = "invalid JSON body: %s" % exc
        else:
            status = 500
            sys.stderr.write(traceback.format_exc())
        try:
            self._send_error_json(exc, status=status)
        except Exception:
            pass

    # --- endpoint implementations ---

    def _serve_app_html(self):
        path = os.path.join(HERE, "app.html")
        if not os.path.isfile(path):
            # app.html is produced by the frontend agent; until then, a stub.
            stub = (
                b"<!doctype html><meta charset=utf-8>"
                b"<title>Document Builder</title>"
                b"<h1>Document Builder</h1>"
                b"<p>app.html is not present yet.</p>"
            )
            return self._send_bytes(stub, "text/html; charset=utf-8")
        with open(path, "rb") as fh:
            body = fh.read()
        return self._send_bytes(body, "text/html; charset=utf-8")

    def _serve_asset(self, path):
        rel = path[len("/assets/"):]
        rel = rel.replace("\\", "/")
        if ".." in rel.split("/"):
            return self._send_error_json("forbidden", status=403)
        full = os.path.join(HERE, "assets", *rel.split("/"))
        if not os.path.isfile(full):
            return self._send_error_json("not found", status=404)
        ctype = _guess_content_type(full)
        with open(full, "rb") as fh:
            body = fh.read()
        return self._send_bytes(body, ctype)

    def _serve_project_file(self, path, qs):
        """Serve a project-relative file (e.g. images/<name>.png) from the
        project folder identified by the `dir` query param.

        The frontend requests image thumbnails as `images/<file>?dir=<project>`.
        Only the project's `images/` subtree is served; traversal is rejected.
        """
        dir_arg = (qs.get("dir") or [None])[0]
        rel = path.lstrip("/")  # e.g. "images/1-1_checklist.png"
        rel = rel.replace("\\", "/")
        parts = rel.split("/")
        if parts[0] != "images" or ".." in parts:
            return self._send_error_json("forbidden", status=403)
        try:
            project_dir = resolve_project_dir(dir_arg, create=False)
        except ValueError as exc:
            return self._send_error_json(exc, status=400)
        full = os.path.join(project_dir, *parts)
        # Containment guard: resolved file must stay under the project dir.
        full_abs = os.path.abspath(full)
        pdir_abs = os.path.abspath(project_dir)
        if os.path.normcase(os.path.commonpath([full_abs, pdir_abs])) != os.path.normcase(pdir_abs):
            return self._send_error_json("forbidden", status=403)
        if not os.path.isfile(full_abs):
            return self._send_error_json("not found", status=404)
        ctype = _guess_content_type(full_abs)
        with open(full_abs, "rb") as fh:
            body = fh.read()
        return self._send_bytes(body, ctype)

    def _api_projects_list(self):
        """List projects: immediate subdirs of reports_root holding project.json.

        Returns {"projects":[{"dir","title","template","mtime"}, ...]} sorted by
        most-recently-modified first. The "templates" folder is skipped. Returns
        an empty list (not an error) when no reports_root is configured.
        """
        root = CFG.reports_root
        out = []
        if root and os.path.isdir(root):
            for name in os.listdir(root):
                if name == "templates" or name.startswith("_") \
                        or name in _RESERVED_ROOT_DIRS:
                    continue
                pj = os.path.join(root, name, "project.json")
                if not os.path.isfile(pj):
                    continue
                title = ""
                template = ""
                try:
                    with open(pj, "r", encoding="utf-8") as fh:
                        proj = json.load(fh)
                    title = (proj.get("meta") or {}).get("title", "") or ""
                    template = proj.get("template", "") or ""
                except Exception:
                    pass
                out.append(
                    {
                        "dir": name,
                        "title": title,
                        "template": template,
                        "mtime": os.path.getmtime(pj),
                    }
                )
        out.sort(key=lambda p: p["mtime"], reverse=True)
        return self._send_json({"projects": out})

    def _api_project_get(self, qs):
        dir_arg = (qs.get("dir") or [None])[0]
        project_dir = resolve_project_dir(dir_arg, create=False)
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return self._send_json(
                {"project": None, "meta_info": {"exists": False, "mtime": None}}
            )
        with open(pj, "r", encoding="utf-8") as fh:
            project = json.load(fh)
        mtime = os.path.getmtime(pj)
        return self._send_json(
            {"project": project, "meta_info": {"exists": True, "mtime": mtime}}
        )

    def _api_project_put(self, qs):
        dir_arg = (qs.get("dir") or [None])[0]
        saved_at = (qs.get("saved_at") or [None])[0]
        project_dir = resolve_project_dir(dir_arg, create=True)
        project = self._read_json()
        if not isinstance(project, dict):
            return self._send_error_json("body must be a JSON object")
        sv = project.get("schema_version")
        if sv is not None and sv != 1:
            return self._send_error_json("unsupported schema_version: %r" % sv)
        pj = os.path.join(project_dir, "project.json")
        # A4 optimistic concurrency: if the client last saw mtime `saved_at` but the
        # file changed on disk since (a second tab, or an applied update bundle),
        # refuse with 409 so a stale autosave cannot clobber / revert it. The client
        # then reloads or explicitly overwrites (by re-PUTting without saved_at).
        if saved_at and os.path.isfile(pj):
            try:
                if abs(os.path.getmtime(pj) - float(saved_at)) > 1e-6:
                    return self._send_json(
                        {"error": "conflict: project.json changed on disk",
                         "conflict": True, "saved_at": os.path.getmtime(pj)},
                        status=409)
            except (ValueError, OSError):
                pass
        body = json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")
        atomic_write(pj, body)
        autosave_snapshot(project_dir, "save")  # capture every saved state
        return self._send_json(
            {
                "ok": True,
                "saved_at": os.path.getmtime(pj),
                "path": pj.replace("\\", "/"),
            }
        )

    def _api_image(self, qs):
        dir_arg = (qs.get("dir") or [None])[0]
        section = (qs.get("section") or [None])[0]
        name = (qs.get("name") or [None])[0]

        ctype = (self.headers.get("Content-Type") or "").lower()
        if ctype.startswith("image/png"):
            png_bytes = self._read_body()
        else:
            payload = self._read_json()
            b64 = payload.get("png_b64")
            if not b64:
                return self._send_error_json("missing png bytes / png_b64")
            if "," in b64 and b64.lstrip().lower().startswith("data:"):
                b64 = b64.split(",", 1)[1]
            png_bytes = base64.b64decode(b64)
            section = payload.get("section", section)
            name = payload.get("name", name)

        if section is None:
            return self._send_error_json("missing 'section'")
        if not png_bytes:
            return self._send_error_json("empty image payload")
        if png_bytes[:8] != b"\x89PNG\r\n\x1a\n":
            return self._send_error_json("payload is not a PNG")

        project_dir = resolve_project_dir(dir_arg, create=True)
        rel, abs_path = next_image_path(project_dir, section, name)
        atomic_write(abs_path, png_bytes)
        return self._send_json({"file": rel})

    def _api_import_xlsx(self):
        payload = self._read_json()
        b64 = payload.get("xlsx_b64")
        mode = payload.get("mode", "grid")
        if not b64:
            return self._send_error_json("missing 'xlsx_b64'")
        if "," in b64 and b64.lstrip().lower().startswith("data:"):
            b64 = b64.split(",", 1)[1]
        xlsx_bytes = base64.b64decode(b64)
        if mode == "grid":
            return self._send_json(parse_xlsx_grid(xlsx_bytes))
        if mode == "compliance":
            return self._send_json(parse_xlsx_compliance(xlsx_bytes))
        return self._send_error_json("unknown mode: %r" % mode)

    def _api_validate_compliance(self):
        payload = self._read_json()
        data = payload.get("data")
        if data is None:
            return self._send_error_json("missing 'data'")
        try:
            result = validate_compliance(data)
        except (ImportError, ModuleNotFoundError) as exc:
            return self._send_error_json(
                "engine not available: %s" % exc, status=503
            )
        return self._send_json(result)

    def _api_export(self, qs):
        dir_arg = (qs.get("dir") or [None])[0]
        fmt = (qs.get("fmt") or ["docx"])[0]
        if fmt not in ("docx", "pdf"):
            return self._send_error_json("fmt must be docx|pdf")
        project_dir = resolve_project_dir(dir_arg, create=False)
        if not os.path.isfile(os.path.join(project_dir, "project.json")):
            return self._send_error_json("no project.json in dir", status=404)
        payload = {}
        if int(self.headers.get("Content-Length") or 0) > 0:
            payload = self._read_json()
        save_first = bool(payload.get("save_first"))
        try:
            result = run_export(project_dir, fmt, save_first=save_first)
        except (ImportError, ModuleNotFoundError) as exc:
            return self._send_error_json(
                "engine not available: %s" % exc, status=503
            )
        return self._send_json(result)

    def _api_export_stream(self, qs):
        """Export with a LIVE progress feed as newline-delimited JSON. Emits
        {type:'phase',label}, {type:'progress',done,total,label} (throttled to
        ~50 lines), then {type:'done',result} or {type:'error',error}. The body is
        close-delimited (Connection: close, no Content-Length); the GUI reads the
        stream to drive a progress bar. /api/export stays as the plain fallback."""
        dir_arg = (qs.get("dir") or [None])[0]
        fmt = (qs.get("fmt") or ["docx"])[0]
        if fmt not in ("docx", "pdf"):
            return self._send_error_json("fmt must be docx|pdf")
        try:
            project_dir = resolve_project_dir(dir_arg, create=False)
        except ValueError as exc:
            return self._send_error_json(str(exc))
        if not os.path.isfile(os.path.join(project_dir, "project.json")):
            return self._send_error_json("no project.json in dir", status=404)
        payload = {}
        if int(self.headers.get("Content-Length") or 0) > 0:
            payload = self._read_json()
        save_first = bool(payload.get("save_first"))

        self.send_response(200)
        self.send_header("Content-Type", "application/x-ndjson; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("X-Accel-Buffering", "no")   # discourage any proxy buffering
        self.send_header("Connection", "close")
        self.end_headers()
        self.close_connection = True

        def emit(obj):
            try:
                self.wfile.write((json.dumps(obj, ensure_ascii=False) + "\n").encode("utf-8"))
                self.wfile.flush()
            except Exception:
                pass   # client gone: writes are ignored, the render still finishes

        last = {"done": -1}

        def on_progress(done, total, label):
            step = max(1, total // 50)        # ~50 lines max; always send the last
            if done - last["done"] >= step or done >= total:
                last["done"] = done
                emit({"type": "progress", "done": done, "total": total, "label": label or ""})

        try:
            result = run_export(project_dir, fmt, save_first=save_first,
                                on_progress=on_progress,
                                on_phase=lambda label: emit({"type": "phase", "label": label}))
            emit({"type": "done", "result": result})
        except (ImportError, ModuleNotFoundError) as exc:
            emit({"type": "error", "error": "engine not available: %s" % exc})
        except Exception as exc:
            sys.stderr.write(traceback.format_exc())
            emit({"type": "error", "error": str(exc)})

    def _api_export_xlsx(self):
        """Export ONE table/datatable block to a .xlsx that visually mirrors the
        Word table. Body: {dir, block}. The project's template config supplies the
        compliance fills/flags (datatable) or the header fill (free table). Returns
        {ok, filename, xlsx_b64} -- the client turns the base64 into a download."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        block = payload.get("block")
        if not isinstance(block, dict):
            return self._send_error_json("missing 'block'")
        project_dir = resolve_project_dir(payload.get("dir"), create=False)
        pj = os.path.join(project_dir, "project.json")
        if not os.path.isfile(pj):
            return self._send_error_json("no project.json in dir", status=404)
        with open(pj, "r", encoding="utf-8") as fh:
            project = json.load(fh)
        # resolve the template config exactly like run_export (compliance section)
        engine = _import_engine()
        tid = project.get("template")
        tpl_cfg = tstore.template_config_path(CFG.reports_root, tid) if tid else None
        explicit = tpl_cfg or CFG.template_config_path
        cfg = engine._load_config(engine._resolve_config_path(project, project_dir, explicit))
        xe = _import_xlsx_export()
        try:
            data = xe.build_block_xlsx(block, cfg)
        except ValueError as ex:
            return self._send_error_json(str(ex))
        return self._send_json({
            "ok": True,
            "filename": xe.filename_for(block),
            "xlsx_b64": base64.b64encode(data).decode("ascii"),
        })

    # --- template library ---

    def _api_templates_list(self):
        return self._send_json(
            {"templates": tstore.list_templates(CFG.reports_root)}
        )

    def _api_template_get(self, qs):
        tid = (qs.get("id") or [None])[0]
        return self._send_json(tstore.get_template(CFG.reports_root, tid))

    def _api_template_put(self, qs):
        tid = (qs.get("id") or [None])[0]
        body = self._read_json()
        if not isinstance(body, dict):
            return self._send_error_json("body must be a JSON object")
        config = body.get("config")
        if not isinstance(config, dict):
            return self._send_error_json("missing 'config'")
        name = body.get("name") or tid
        skeleton = body.get("skeleton", [])
        rid = tstore.save_template(
            CFG.reports_root, tid, name, config, skeleton, atomic_write
        )
        return self._send_json({"ok": True, "id": rid})

    def _api_template_delete(self, qs):
        tid = (qs.get("id") or [None])[0]
        tstore.delete_template(CFG.reports_root, tid)
        return self._send_json({"ok": True})

    def _api_import_docx(self):
        payload = self._read_json()
        b64 = payload.get("docx_b64")
        mode = payload.get("mode", "report")
        if not b64:
            return self._send_error_json("missing 'docx_b64'")
        if "," in b64 and b64.lstrip().lower().startswith("data:"):
            b64 = b64.split(",", 1)[1]
        docx_bytes = base64.b64decode(b64)
        try:
            di = _import_docx_module()
        except (ImportError, ModuleNotFoundError) as exc:
            return self._send_error_json(
                "docx parser not available: %s" % exc, status=503
            )
        with _temp_docx(docx_bytes) as tmp_path:
            if mode == "template":
                return self._import_docx_template(di, tmp_path, payload)
            if mode == "report":
                return self._import_docx_report(di, tmp_path, payload)
        return self._send_error_json("unknown mode: %r" % mode)

    def _import_docx_template(self, di, tmp_path, payload):
        warnings = []
        logo_dir = tempfile.mkdtemp(prefix="docx_logo_")
        try:
            derived = di.derive_template(
                tmp_path, logo_dir=logo_dir, warn=warnings.append
            )
            tid, rname = tstore.save_derived_template(
                CFG.reports_root, payload.get("name"), derived,
                atomic_write, _sanitize_name,
                logo_dir=logo_dir, warn=warnings.append,
            )
        finally:
            _rmtree_quiet(logo_dir)
        return self._send_json(
            {"id": tid, "name": rname, "warnings": warnings or None}
        )

    def _import_docx_report(self, di, tmp_path, payload):
        warnings = []
        dir_arg = payload.get("dir")
        if not dir_arg:
            return self._send_error_json("missing 'dir' for report import")
        project_dir = resolve_project_dir(dir_arg, create=True)
        images_dir = os.path.join(project_dir, "images")
        parsed = di.parse_docx_report(
            tmp_path, images_dir=images_dir, warn=warnings.append
        )
        meta = parsed.get("meta", {})
        outline = parsed.get("outline", [])
        _rewrite_image_paths(outline)

        # Determine the bound template id -- import MUST NOT dead-end.
        tid = payload.get("template")
        if not (tid and tstore.template_config_path(CFG.reports_root, tid)):
            logo_dir = tempfile.mkdtemp(prefix="docx_logo_")
            try:
                derived = di.derive_template(
                    tmp_path, logo_dir=logo_dir, warn=warnings.append
                )
                name = meta.get("title") or os.path.basename(
                    project_dir.rstrip(os.sep)
                )
                tid, _ = tstore.save_derived_template(
                    CFG.reports_root, name, derived,
                    atomic_write, _sanitize_name,
                    logo_dir=logo_dir, warn=warnings.append,
                )
            finally:
                _rmtree_quiet(logo_dir)

        project_seed = {
            "schema_version": 1, "template": tid,
            "meta": meta, "outline": outline,
        }
        return self._send_json(
            {
                "project": project_seed,
                "template_id": tid,
                "warnings": warnings or None,
            }
        )

    def _api_new_from_template(self):
        payload = self._read_json()
        dir_arg = payload.get("dir")
        tid = payload.get("template")
        force = bool(payload.get("force"))
        if not dir_arg:
            return self._send_error_json("missing 'dir'")
        tpl = tstore.get_template(CFG.reports_root, tid)  # 404 if absent
        outline = _instantiate_skeleton(tpl.get("skeleton", []))
        project = {
            "schema_version": 1, "template": tid,
            "meta": _empty_meta(), "outline": outline,
        }
        project_dir = resolve_project_dir(dir_arg, create=True)
        pj = os.path.join(project_dir, "project.json")
        # Refuse to clobber an existing project unless the caller explicitly
        # opted in; even then keep a .bak so an accidental overwrite is
        # recoverable. (This is the most destructive path in the tool.)
        if os.path.isfile(pj):
            if not force:
                return self._send_error_json(
                    "a project already exists in this folder", status=409
                )
            try:
                os.replace(pj, pj + ".bak")
            except OSError:
                pass
        atomic_write(
            pj, json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")
        )
        return self._send_json({"ok": True, "project": project})

    def _api_paste_import(self, qs):
        """Full-project replace from pasted text (de-LLM upstream channel, WS2).

        Body: the raw pasted project.json text (bytes go browser -> HTTP -> disk,
        never through a model). Validates, backs up the existing project into the
        shared rollback history, writes it, and emits a compact structural diff to
        <project>/_paste_diff.md. A truncated/mangled paste fails to parse and is
        rejected -- the on-disk project is left untouched. Returns
        {ok, backup, warn, diff, diff_file}."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        dir_arg = (qs.get("dir") or [None])[0]
        raw = self._read_body()
        if not raw:
            return self._send_error_json("empty paste body -- nothing changed")
        try:
            project = json.loads(raw.decode("utf-8"))
        except Exception as exc:
            return self._send_error_json(
                "could not parse pasted JSON (truncated?): %s -- nothing changed" % exc)
        if not isinstance(project, dict):
            return self._send_error_json(
                "pasted JSON must be an object -- nothing changed")
        sv = project.get("schema_version")
        if sv is not None and sv != 1:
            return self._send_error_json(
                "unsupported schema_version: %r -- nothing changed" % sv)

        project_dir = resolve_project_dir(dir_arg, create=True)
        pj = os.path.join(project_dir, "project.json")
        old_project = None
        if os.path.isfile(pj):
            try:
                with open(pj, "r", encoding="utf-8") as fh:
                    old_project = json.load(fh)
            except Exception:
                old_project = None

        # identity check: guard against pasting into the wrong project folder.
        warns = []
        if isinstance(old_project, dict):
            ot = (old_project.get("meta") or {}).get("title")
            nt = (project.get("meta") or {}).get("title")
            if ot and nt and ot != nt:
                warns.append('pasted title "%s" != existing "%s"' % (nt, ot))
            otpl, ntpl = old_project.get("template"), project.get("template")
            if otpl and ntpl and otpl != ntpl:
                warns.append('pasted template "%s" != existing "%s"' % (ntpl, otpl))

        name = os.path.basename(project_dir.rstrip(os.sep)) or "project"
        diff_md = paste_import_diff(old_project, project, name, warns)

        body = json.dumps(project, ensure_ascii=False, indent=2).encode("utf-8")
        autosave_snapshot(project_dir, "prepaste")  # keep pre-paste state recoverable
        apply_update = _import_apply_update()
        rel = os.path.relpath(pj, CFG.reports_root)
        rec = apply_update.record_replace(CFG.reports_root, rel, body)

        diff_path = os.path.join(project_dir, "_paste_diff.md")
        atomic_write(diff_path, diff_md.encode("utf-8"))

        return self._send_json({
            "ok": True,
            "backup": (rec.get("backup") or "").replace("\\", "/"),
            "warn": warns,
            "diff": diff_md,
            "diff_file": diff_path.replace("\\", "/"),
        })

    def _api_copy_diff(self, qs):
        """Compute the upstream 'Copy diff' delta for a project (WS: incremental
        upstream). Body: the CURRENT project JSON (the editor's in-memory state,
        same payload as Copy text). The server diffs it against the last synced
        <project>/_baseline.json and returns a COMPACT op-diff the user pastes to
        the assistant instead of the whole report -- only edited sections travel.

        Returns {ok, no_baseline, empty, diff_text, summary, diff_chars,
        full_chars}. no_baseline=True when the project has never been synced from
        the assistant yet (the UI falls back to Copy text)."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        dir_arg = (qs.get("dir") or [None])[0]
        raw = self._read_body()
        if not raw:
            return self._send_error_json("empty body -- nothing to diff")
        try:
            current = json.loads(raw.decode("utf-8"))
        except Exception as exc:
            return self._send_error_json("could not parse current project JSON: %s" % exc)
        if not isinstance(current, dict):
            return self._send_error_json("current project must be a JSON object")

        project_dir = resolve_project_dir(dir_arg)
        name = os.path.basename(project_dir.rstrip(os.sep)) or "project"
        baseline_path = os.path.join(project_dir, "_baseline.json")
        if not os.path.isfile(baseline_path):
            return self._send_json({
                "ok": True, "no_baseline": True,
                "hint": ("No sync baseline yet for this project -- Copy diff starts "
                         "working after the next 'Apply update' from the assistant. "
                         "Use Copy text this once."),
            })
        try:
            with open(baseline_path, "r", encoding="utf-8") as fh:
                baseline = json.load(fh)
        except Exception as exc:
            return self._send_error_json("baseline unreadable: %s" % exc)

        apply_update = _import_apply_update()
        diff = apply_update.make_text_diff(baseline, current, name)
        diff_text = json.dumps(diff, ensure_ascii=False, separators=(",", ":"))
        full_text = json.dumps(current, ensure_ascii=False, separators=(",", ":"))
        return self._send_json({
            "ok": True,
            "no_baseline": False,
            "empty": apply_update.diff_is_empty(diff),
            # Edits spread across a small report (or a datatable-cell change, which
            # resends its whole section) can make the op-diff bigger than the full
            # text; surface that so the UI never silently hands over MORE to paste.
            "smaller": len(diff_text) < len(full_text),
            "diff_text": diff_text,
            "summary": apply_update.diff_summary(diff),
            "diff_chars": len(diff_text),
            "full_chars": len(full_text),
        })

    def _api_apply_update(self):
        """Apply an uploaded update bundle (.zip) to the reports root.

        Body: {"name": <filename>, "zip_b64": <base64 zip>}. The zip is stored
        under reports_root/_updates/ and applied via the shared apply_update
        module (full backup of every overwritten file first). Returns the
        apply summary {note, actions:[{verb,rel,warn}], backup, logs}.
        """
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        b64 = payload.get("zip_b64") or ""
        if not b64:
            return self._send_error_json("missing 'zip_b64'")
        try:
            raw = base64.b64decode(b64)
        except Exception:
            return self._send_error_json("invalid base64 zip")
        base = payload.get("name") or "update"
        if base.lower().endswith(".zip"):
            base = base[:-4]
        name = (_sanitize_name(base) or "update") + ".zip"
        updates_dir = os.path.join(CFG.reports_root, "_updates")
        os.makedirs(updates_dir, exist_ok=True)
        dest = os.path.join(updates_dir, name)
        atomic_write(dest, raw)
        # snapshot every project's current state BEFORE applying, so an update
        # bundle can never silently overwrite local edits beyond recovery.
        autosaved = autosave_all(CFG.reports_root, "preapply")
        apply_update = _import_apply_update()
        summary = apply_update.apply_bundle(CFG.reports_root, dest, dry=False)
        return self._send_json({"ok": True, "autosaved": autosaved, **summary})

    def _api_rollback(self):
        """Undo the most recent apply / paste-import by restoring the newest
        backup (shared history). Non-interactive counterpart of the CLI
        --rollback. Returns {ok, restored, from, pre} or an error."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        apply_update = _import_apply_update()
        result = apply_update.rollback_last(CFG.reports_root)
        if not result.get("ok"):
            return self._send_error_json(result.get("error", "rollback failed"),
                                         status=400)
        return self._send_json(result)

    def _api_autosaves(self, qs):
        """List the local auto-snapshots of a project's project.json (newest first)."""
        dir_arg = (qs.get("dir") or [None])[0]
        project_dir = resolve_project_dir(dir_arg)
        return self._send_json({"autosaves": list_autosaves(project_dir)})

    def _api_autosave_restore(self):
        """Restore a named auto-snapshot to project.json (current is snapshotted
        first, so the restore is itself undoable). Body: {dir, name}."""
        payload = self._read_json()
        dir_arg = payload.get("dir")
        name = payload.get("name")
        if not name:
            return self._send_error_json("missing 'name'")
        project_dir = resolve_project_dir(dir_arg)
        return self._send_json(restore_autosave(project_dir, name))

    def _new_project_target(self, name):
        """Validate a NEW project folder name -> (segment, abs path). Single
        sanitized segment, contained under reports_root, not reserved, not '_'."""
        if not name or not isinstance(name, str):
            raise ValueError("missing new project name")
        seg = _sanitize_name(name.strip())
        if not seg or seg.startswith("_") or seg in _RESERVED_ROOT_DIRS:
            raise ValueError("invalid project name: %r" % name)
        return seg, resolve_project_dir(seg)  # resolve_project_dir enforces containment

    def _api_project_delete(self):
        """Move a project to reports_root/_trash/<name>-<ts>/ (recoverable, never
        a hard delete). Body: {dir}."""
        if not CFG.reports_root:
            return self._send_error_json("no reports root configured", status=400)
        payload = self._read_json()
        project_dir = resolve_project_dir(payload.get("dir"))
        if not os.path.isdir(project_dir):
            return self._send_error_json("no such project", status=404)
        name = os.path.basename(project_dir.rstrip(os.sep)) or "project"
        trash = os.path.join(CFG.reports_root, "_trash")
        os.makedirs(trash, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        dest = os.path.join(trash, "%s-%s" % (name, ts))
        shutil.move(project_dir, dest)
        return self._send_json({"ok": True, "trashed_to": dest.replace("\\", "/")})

    def _api_project_rename(self):
        """Rename a project folder (and optionally its meta.title).
        Body: {dir, new_name, title?}."""
        payload = self._read_json()
        src = resolve_project_dir(payload.get("dir"))
        if not os.path.isdir(src):
            return self._send_error_json("no such project", status=404)
        seg, dest = self._new_project_target(payload.get("new_name"))
        if os.path.normcase(dest) == os.path.normcase(src):
            return self._send_json({"ok": True, "dir": seg})  # no-op
        if os.path.exists(dest):
            return self._send_error_json("a project named '%s' already exists" % seg,
                                         status=409)
        shutil.move(src, dest)
        title = payload.get("title")
        if title is not None:
            pj = os.path.join(dest, "project.json")
            if os.path.isfile(pj):
                try:
                    with open(pj, encoding="utf-8") as fh:
                        d = json.load(fh)
                    d.setdefault("meta", {})["title"] = title
                    atomic_write(pj, json.dumps(d, ensure_ascii=False,
                                                indent=2).encode("utf-8"))
                except Exception:
                    pass
        return self._send_json({"ok": True, "dir": seg})

    def _api_project_copy(self):
        """Duplicate a project (project.json + images/ only, not the heavy
        _backups/_autosave/_updates) under a new name. Body: {dir, new_name}."""
        payload = self._read_json()
        src = resolve_project_dir(payload.get("dir"))
        if not os.path.isfile(os.path.join(src, "project.json")):
            return self._send_error_json("no such project", status=404)
        seg, dest = self._new_project_target(payload.get("new_name"))
        if os.path.exists(dest):
            return self._send_error_json("a project named '%s' already exists" % seg,
                                         status=409)
        os.makedirs(dest)
        shutil.copy2(os.path.join(src, "project.json"),
                     os.path.join(dest, "project.json"))
        img = os.path.join(src, "images")
        if os.path.isdir(img):
            shutil.copytree(img, os.path.join(dest, "images"))
        return self._send_json({"ok": True, "dir": seg})


def _import_apply_update():
    """Import the repo-root apply_update module (shared with the CLI).

    Reload on every call so a long-running server picks up an updated
    apply_update.py after a ``git pull`` WITHOUT needing a restart (Python caches
    modules in sys.modules; a running server would otherwise keep the stale code).
    """
    import importlib
    repo_root = os.path.abspath(os.path.join(HERE, ".."))
    if repo_root not in sys.path:
        sys.path.insert(0, repo_root)
    import apply_update  # noqa: E402  (repo-root sibling of builder/)
    return importlib.reload(apply_update)


def _rmtree_quiet(path):
    import shutil
    try:
        shutil.rmtree(path)
    except OSError:
        pass


def _guess_content_type(path):
    ext = os.path.splitext(path)[1].lower()
    return {
        ".html": "text/html; charset=utf-8",
        ".js": "application/javascript; charset=utf-8",
        ".css": "text/css; charset=utf-8",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".svg": "image/svg+xml",
        ".json": "application/json; charset=utf-8",
        ".woff2": "font/woff2",
        ".woff": "font/woff",
    }.get(ext, "application/octet-stream")


# ---------------------------------------------------------------------------
# Entry point.
# ---------------------------------------------------------------------------


def make_server(port=None, root=None, config_path=None, bind="127.0.0.1"):
    """Construct (but do not start) the HTTP server. Used by smoke tests too."""
    if port is not None:
        CFG.port = port
    if root is not None:
        CFG.reports_root = os.path.abspath(root)
    if config_path is not None:
        CFG.template_config_path = os.path.abspath(config_path)
    CFG.bind = bind
    httpd = ThreadingHTTPServer((bind, CFG.port), Handler)
    return httpd


def default_config_path():
    """Best-effort default template config path.

    Prefer an explicit --config / BUILDER_TEMPLATE_CONFIG. Otherwise look in a
    sibling `local/` folder for a single `template_config_*.json` and use it.
    Returns None if nothing is found (the server still starts; endpoints that
    need the config then report a clean error).
    """
    import glob

    local_dir = os.path.abspath(os.path.join(HERE, "..", "local"))
    matches = sorted(glob.glob(os.path.join(local_dir, "template_config_*.json")))
    return matches[0] if matches else None


def default_reports_root():
    """Default reports_root: a sibling ``local/`` folder if it exists, else None.

    Lets ``python builder/server.py`` (or a start script) Just Work without
    ``--root`` in the common layout where reports live under ``<repo>/local``.
    """
    cand = os.path.abspath(os.path.join(HERE, "..", "local"))
    return cand if os.path.isdir(cand) else None


def main(argv=None):
    parser = argparse.ArgumentParser(description="Structured document builder server")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument(
        "--root",
        default=os.environ.get("BUILDER_REPORTS_ROOT") or default_reports_root(),
        help="reports_root: project folders must stay under this path",
    )
    parser.add_argument(
        "--config",
        default=os.environ.get("BUILDER_TEMPLATE_CONFIG") or default_config_path(),
        help="path to the template config JSON",
    )
    parser.add_argument("--bind", default="127.0.0.1")
    parser.add_argument(
        "--open", dest="open_browser", action="store_true",
        help="open the app in the default browser once the server is up",
    )
    args = parser.parse_args(argv)

    CFG.reports_root = os.path.abspath(args.root) if args.root else None
    CFG.template_config_path = os.path.abspath(args.config) if args.config else None
    CFG.bind = args.bind
    CFG.port = args.port

    if args.bind != "127.0.0.1":
        sys.stderr.write(
            "WARNING: binding to %s; this server is intended for 127.0.0.1 only.\n"
            % args.bind
        )

    httpd = ThreadingHTTPServer((args.bind, args.port), Handler)
    sys.stderr.write(
        "Document builder server on http://%s:%d (root=%s, config=%s)\n"
        % (args.bind, args.port, CFG.reports_root, CFG.template_config_path)
    )
    if args.open_browser:
        import webbrowser
        import threading
        url = "http://127.0.0.1:%d/app.html" % args.port
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        httpd.server_close()


if __name__ == "__main__":
    main()
