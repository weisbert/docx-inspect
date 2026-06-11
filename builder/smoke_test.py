#!/usr/bin/env python3
"""Smoke test for server.py.

Starts the server on a test port (bound to 127.0.0.1) against a temporary
reports_root and a minimal self-contained template config, then exercises every
endpoint from CONTRACT.md using only stdlib urllib. Asserts HTTP status and the
shape of each response.

Endpoints that require the rendering engine (engine.py / tables.py) are probed
too: if those modules are not importable yet, the endpoint is expected to return
a clean JSON error (503), which is recorded as "stubbed (engine)" rather than a
failure. Once the engine lands, those become full passes with no code change.

Run:
    .venv\\Scripts\\python.exe builder\\smoke_test.py
"""

import base64
import io
import json
import os
import sys
import tempfile
import threading
import time
from urllib.request import Request, urlopen
from urllib.error import HTTPError

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import server  # noqa: E402

PORT = 8799
BASE = "http://127.0.0.1:%d" % PORT

RESULTS = []  # (name, status_str)


def record(name, status):
    RESULTS.append((name, status))
    print("  [%-16s] %s" % (status, name))


# ---------------------------------------------------------------------------
# Fixtures.
# ---------------------------------------------------------------------------


MINIMAL_CONFIG = {
    "id": "test_tpl_v1",
    "skeleton": [
        {"title": "Overview", "fixed_body": "sec_1_1", "children": []},
        {"title": "Compliance Table", "datatable": "compliance"},
    ],
    "cover": {
        "company_line": "ACME",
        "secrecy_default": "Internal",
        "page_count_field": True,
        "fields": [
            {"key": "title", "label": "Title", "table": "info", "required": True},
            {"key": "author", "label": "Author", "table": "signature", "required": False},
        ],
        "tables": {},
    },
    "fixed_bodies": {
        "sec_1_1": {"style": "body", "paragraphs": []}
    },
    "styles": {"colors": {"red": "FF0000"}},
    "compliance": {
        "axis_labels": ["MIN", "TYP", "MAX", "NTWC"],
        "default_limit": {"le": "<= upper", "ge": ">= target", "range": "within"},
        "flag_color": "FF0000",
        "setting_kinds": ["common_setting", "module_setting", "tb"],
    },
    "logo": "logo.png",
    "ui_strings": {"toolbar.save": "Save"},
}


def make_compliance_xlsx_bytes():
    """Build a tiny compliance-shaped xlsx mirroring the renderer layout."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    # Row1: group titles  Row2: stage  Row3: axis labels
    # Cols: A=Category B=Item C..F=Spec axes G..J=Sim axes K=Unit
    ws["C1"] = "ACME Spec"
    ws["G1"] = "Sim_pilot"
    ws["G2"] = "Pre"
    ws["A3"] = "Category"
    ws["B3"] = "Item"
    for i, ax in enumerate(["MIN", "TYP", "MAX", "NTWC"]):
        ws.cell(row=3, column=3 + i, value=ax)   # spec axes C..F
        ws.cell(row=3, column=7 + i, value=ax)   # sim axes  G..J
    ws["K3"] = "Unit"
    # data row
    ws["A4"] = "Supply"
    ws["B4"] = "I_total"
    ws["E4"] = 500          # spec MAX
    ws["G4"] = 279
    ws["H4"] = 490
    ws["I4"] = 758
    ws["K4"] = "uA"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_plain_xlsx_bytes():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "H1"
    ws["B1"] = "H2"
    ws["A2"] = "a"
    ws["B2"] = 2
    ws.merge_cells("A1:B1")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# HTTP helpers.
# ---------------------------------------------------------------------------


def call(method, path, body=None, headers=None, raw=None):
    """Return (status, parsed_json_or_bytes)."""
    url = BASE + path
    data = None
    hdrs = dict(headers or {})
    if raw is not None:
        data = raw
    elif body is not None:
        data = json.dumps(body).encode("utf-8")
        hdrs.setdefault("Content-Type", "application/json")
    req = Request(url, data=data, method=method, headers=hdrs)
    try:
        with urlopen(req) as resp:
            status = resp.status
            payload = resp.read()
            ctype = resp.headers.get("Content-Type", "")
    except HTTPError as e:
        status = e.code
        payload = e.read()
        ctype = e.headers.get("Content-Type", "")
    if "application/json" in ctype:
        try:
            return status, json.loads(payload.decode("utf-8"))
        except Exception:
            return status, payload
    return status, payload


# ---------------------------------------------------------------------------
# Test body.
# ---------------------------------------------------------------------------


def run():
    tmp = tempfile.mkdtemp(prefix="builder_smoke_")
    root = os.path.join(tmp, "reports")
    os.makedirs(root, exist_ok=True)
    config_path = os.path.join(tmp, "template_config_test.json")
    with open(config_path, "w", encoding="utf-8") as fh:
        json.dump(MINIMAL_CONFIG, fh, ensure_ascii=False)

    httpd = server.make_server(
        port=PORT, root=root, config_path=config_path, bind="127.0.0.1"
    )
    t = threading.Thread(target=httpd.serve_forever, daemon=True)
    t.start()
    # wait until listening
    for _ in range(50):
        try:
            s, _ = call("GET", "/api/health")
            if s == 200:
                break
        except Exception:
            time.sleep(0.05)

    failures = 0
    try:
        failures += check_app_html()
        failures += check_assets()
        failures += check_config()
        failures += check_project_roundtrip(root)
        failures += check_image(root)
        failures += check_import_xlsx()
        failures += check_validate_compliance()
        failures += check_export(root)
        failures += check_errors(root)
    finally:
        httpd.shutdown()
        httpd.server_close()

    print("\n=== SMOKE SUMMARY ===")
    for name, status in RESULTS:
        print("  %-16s %s" % (status, name))
    print("=====================")
    return failures


def expect(cond, name, detail=""):
    if cond:
        record(name, "PASS")
        return 0
    record(name, "FAIL")
    if detail:
        print("      -> %s" % detail)
    return 1


def check_app_html():
    s, body = call("GET", "/app.html")
    f = expect(s == 200 and isinstance(body, (bytes, bytearray)),
               "GET /app.html", "status=%s" % s)
    s2, _ = call("GET", "/")
    f += expect(s2 == 200, "GET / (app.html)", "status=%s" % s2)
    return f


def check_assets():
    # No asset exists -> 404 JSON error (not a crash).
    s, body = call("GET", "/assets/does-not-exist.js")
    return expect(s == 404 and isinstance(body, dict) and "error" in body,
                  "GET /assets/* 404", "status=%s body=%r" % (s, body))


def check_config():
    s, body = call("GET", "/api/config")
    ok = (
        s == 200
        and isinstance(body, dict)
        and body.get("template") == "test_tpl_v1"
        and "skeleton" in body
        and "ui_strings" in body
        and "cover" in body
        and "compliance" in body
        and "axis_labels" in body["compliance"]
    )
    return expect(ok, "GET /api/config", "status=%s body=%r" % (s, body))


def check_project_roundtrip(root):
    dirname = "proj_a"
    # GET before exists -> exists:false
    s, body = call("GET", "/api/project?dir=%s" % dirname)
    f = expect(
        s == 200 and body.get("meta_info", {}).get("exists") is False,
        "GET /api/project (absent)", "status=%s body=%r" % (s, body),
    )
    # PUT
    project = {
        "schema_version": 1,
        "template": "test_tpl_v1",
        "meta": {"title": "T", "reviewers": [], "revisions": []},
        "outline": [],
    }
    s, body = call("PUT", "/api/project?dir=%s" % dirname, body=project)
    f += expect(
        s == 200 and body.get("ok") is True and "path" in body,
        "PUT /api/project (save)", "status=%s body=%r" % (s, body),
    )
    # images/ folder created
    f += expect(
        os.path.isdir(os.path.join(root, dirname, "images")),
        "PUT creates images/", "missing images dir",
    )
    # GET after exists
    s, body = call("GET", "/api/project?dir=%s" % dirname)
    f += expect(
        s == 200
        and body.get("meta_info", {}).get("exists") is True
        and body.get("project", {}).get("template") == "test_tpl_v1",
        "GET /api/project (present)", "status=%s body=%r" % (s, body),
    )
    # reject bad schema_version
    bad = dict(project)
    bad["schema_version"] = 2
    s, body = call("PUT", "/api/project?dir=%s" % dirname, body=bad)
    f += expect(
        s == 400 and "error" in body,
        "PUT /api/project rejects schema!=1", "status=%s" % s,
    )
    return f


def check_image(root):
    dirname = "proj_a"
    png = base64.b64decode(
        # 1x1 transparent PNG
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+M9QDwADhgGAWjR9"
        "awAAAABJRU5ErkJggg=="
    )
    # raw png bytes form
    s, body = call(
        "POST",
        "/api/image?dir=%s&section=4&name=diagram" % dirname,
        raw=png,
        headers={"Content-Type": "image/png"},
    )
    f = expect(
        s == 200 and body.get("file", "").startswith("images/4-1_diagram"),
        "POST /api/image (raw png)", "status=%s body=%r" % (s, body),
    )
    saved = os.path.join(root, dirname, body.get("file", "")) if isinstance(body, dict) else ""
    f += expect(saved and os.path.isfile(saved), "image file written", saved)

    # GET the saved image back through the project-file serving route
    # (frontend thumbnails request `images/<f>?dir=<project>`).
    rel = body.get("file", "") if isinstance(body, dict) else ""
    s, got = call("GET", "/%s?dir=%s" % (rel, dirname))
    f += expect(
        s == 200 and isinstance(got, (bytes, bytearray)) and got[:8] == b"\x89PNG\r\n\x1a\n",
        "GET /images/<f> (project image served)", "status=%s" % s,
    )
    # traversal out of images/ must be rejected
    s, _ = call("GET", "/images/../project.json?dir=%s" % dirname)
    f += expect(s in (400, 403, 404), "GET image traversal blocked", "status=%s" % s)

    # base64 json form, sequence increments
    s, body = call(
        "POST",
        "/api/image?dir=%s" % dirname,
        body={"png_b64": base64.b64encode(png).decode(), "section": "4"},
    )
    f += expect(
        s == 200 and body.get("file") == "images/4_2.png",
        "POST /api/image (b64 json, seq++)", "status=%s body=%r" % (s, body),
    )

    # reject non-png
    s, body = call(
        "POST",
        "/api/image?dir=%s&section=4" % dirname,
        raw=b"not a png",
        headers={"Content-Type": "image/png"},
    )
    f += expect(s == 400 and "error" in body, "POST /api/image rejects non-PNG", "status=%s" % s)
    return f


def check_import_xlsx():
    # grid mode
    xb = base64.b64encode(make_plain_xlsx_bytes()).decode()
    s, body = call("POST", "/api/import-xlsx", body={"xlsx_b64": xb, "mode": "grid"})
    ok = (
        s == 200
        and isinstance(body.get("rows"), list)
        and body["rows"][0][0] == "H1"
        and isinstance(body.get("merges"), list)
        and len(body["merges"]) == 1
    )
    f = expect(ok, "POST /api/import-xlsx (grid)", "status=%s body=%r" % (s, body))

    # compliance mode
    cb = base64.b64encode(make_compliance_xlsx_bytes()).decode()
    s, body = call("POST", "/api/import-xlsx", body={"xlsx_b64": cb, "mode": "compliance"})
    data = body.get("data", {}) if isinstance(body, dict) else {}
    ok = (
        s == 200
        and isinstance(data.get("rows"), list)
        and len(data["rows"]) >= 1
        and data["rows"][0]["item"] == "I_total"
        and data["rows"][0]["sim_mtm"][2] == 758
        and len(data.get("sims", [])) >= 1
    )
    f += expect(ok, "POST /api/import-xlsx (compliance)", "status=%s body=%r" % (s, body))

    # bad mode
    s, body = call("POST", "/api/import-xlsx", body={"xlsx_b64": xb, "mode": "nope"})
    f += expect(s == 400 and "error" in body, "import-xlsx rejects bad mode", "status=%s" % s)
    return f


COMPLIANCE_DATA = {
    "spec_name": "Spec",
    "sims": [{"key": "pre", "title": "Pilot", "stage": "Pre"}],
    "rows": [
        {
            "cat": "Supply", "item": "I_total", "unit": "uA", "kind": "result",
            "spec": 500, "spec_mtm": [None, 500, None],
            "sim_mtm": [279, 490, 758], "spec_ntwc": None, "sim_ntwc": None,
            "limit": "le", "sim_span": False,
        }
    ],
}


def check_validate_compliance():
    s, body = call("POST", "/api/validate-compliance", body={"data": COMPLIANCE_DATA})
    if s == 200 and isinstance(body, dict) and "flags" in body:
        # engine present: I_total MAX (index 2 = 758) exceeds 500 -> flagged.
        ok = body["flags"].get("0") == [2] and body.get("color")
        return expect(ok, "POST /api/validate-compliance", "body=%r" % body)
    if s == 503 and isinstance(body, dict) and "error" in body:
        record("POST /api/validate-compliance", "STUB (engine)")
        return 0
    record("POST /api/validate-compliance", "FAIL")
    print("      -> status=%s body=%r" % (s, body))
    return 1


def check_export(root):
    """Verify the export endpoint's server wiring.

    A full docx render requires a complete template config (page geometry,
    headings, cover tables, etc.). The smoke harness uses an intentionally
    minimal config, so a *successful* render is bonus; what this test owns is
    the server contract: route resolution, project loading, config resolution,
    and that any engine error is surfaced as a clean JSON {error} (never a
    crashed loop). The server-only validation paths (bad fmt, missing project)
    are asserted strictly.
    """
    dirname = "proj_a"  # has a project.json from the roundtrip test
    s, body = call("POST", "/api/export?dir=%s&fmt=docx" % dirname, body={})
    if s == 200 and isinstance(body, dict) and body.get("fmt") == "docx" and "abs" in body:
        f = expect(True, "POST /api/export (docx)", "")
    elif s == 503 and isinstance(body, dict) and "error" in body:
        record("POST /api/export (docx)", "STUB (engine)")
        f = 0
    elif s == 500 and isinstance(body, dict) and "error" in body:
        # Engine present but the minimal smoke config is not fully renderable.
        # Server correctly surfaced the engine error as JSON without crashing.
        record("POST /api/export (docx)", "ENGINE-CFG (minimal config)")
        f = 0
    else:
        record("POST /api/export (docx)", "FAIL")
        print("      -> status=%s body=%r" % (s, body))
        f = 1
    # server-only validation paths (independent of the engine):
    s2, b2 = call("POST", "/api/export?dir=%s&fmt=bogus" % dirname, body={})
    f += expect(s2 == 400 and "error" in b2, "export rejects bad fmt", "status=%s" % s2)
    s3, b3 = call("POST", "/api/export?dir=missing_proj&fmt=docx", body={})
    f += expect(s3 == 404 and "error" in b3, "export 404 when no project.json", "status=%s" % s3)
    return f


def check_errors(root):
    # invalid JSON body
    s, body = call(
        "PUT", "/api/project?dir=proj_b",
        raw=b"{not json", headers={"Content-Type": "application/json"},
    )
    f = expect(s == 400 and "error" in body, "invalid JSON -> 400 error", "status=%s" % s)
    # missing dir param
    s, body = call("GET", "/api/project")
    f += expect(s == 400 and "error" in body, "missing dir -> 400 error", "status=%s" % s)
    # path traversal outside root
    s, body = call("GET", "/api/project?dir=" + "..%2F..%2Fetc")
    f += expect(s in (400, 404) and "error" in body, "path traversal blocked", "status=%s body=%r" % (s, body))
    # unknown route
    s, body = call("GET", "/api/nope")
    f += expect(s == 404 and "error" in body, "unknown route -> 404 error", "status=%s" % s)
    return f


def check_real_config():
    """End-to-end checks against the real company template config + demo project,
    if they are present on this machine. Skipped cleanly otherwise (so the smoke
    test stays self-contained on a clean checkout).

    Asserts the two engine-dependent guarantees that the minimal config cannot:
      - validate-compliance over the demo's real datatable yields exactly 4 flags;
      - export?fmt=docx through the server actually renders a docx file.
    """
    import glob

    local_dir = os.path.abspath(os.path.join(HERE, "..", "local"))
    cfgs = sorted(glob.glob(os.path.join(local_dir, "template_config_*.json")))
    demo = os.path.join(local_dir, "demo_project")
    if not cfgs or not os.path.isfile(os.path.join(demo, "project.json")):
        record("real-config e2e (validate + export)", "SKIP (no local config)")
        return 0

    port = PORT + 1
    base = "http://127.0.0.1:%d" % port
    httpd = server.make_server(
        port=port, root=local_dir, config_path=cfgs[0], bind="127.0.0.1"
    )
    t = threading.Thread(target=httpd.serve_forever, daemon=True)
    t.start()

    def rcall(method, path, body=None):
        data = json.dumps(body).encode("utf-8") if body is not None else None
        hdrs = {"Content-Type": "application/json"} if data else {}
        req = Request(base + path, data=data, method=method, headers=hdrs)
        try:
            with urlopen(req) as resp:
                return resp.status, json.loads(resp.read().decode("utf-8"))
        except HTTPError as e:
            try:
                return e.code, json.loads(e.read().decode("utf-8"))
            except Exception:
                return e.code, {}

    f = 0
    try:
        for _ in range(50):
            try:
                s, _ = rcall("GET", "/api/health")
                if s == 200:
                    break
            except Exception:
                time.sleep(0.05)

        # find the demo datatable
        s, pj = rcall("GET", "/api/project?dir=demo_project")
        dt = None

        def _find(nodes):
            for n in nodes:
                for blk in n.get("blocks", []):
                    if blk.get("type") == "datatable":
                        return blk
                r = _find(n.get("children", []))
                if r:
                    return r
            return None

        if s == 200 and pj.get("project"):
            dt = _find(pj["project"].get("outline", []))

        if dt:
            s, vc = rcall("POST", "/api/validate-compliance", body={"data": dt["data"]})
            nflags = sum(len(v) for v in vc.get("flags", {}).values())
            f += expect(
                s == 200 and nflags == 4,
                "real validate-compliance == 4 flags",
                "flags=%r" % vc.get("flags"),
            )
        else:
            record("real validate-compliance == 4 flags", "SKIP (no datatable)")

        s, ex = rcall("POST", "/api/export?dir=demo_project&fmt=docx", body={})
        ok = (
            s == 200
            and ex.get("fmt") == "docx"
            and os.path.isfile(ex.get("abs", "").replace("/", os.sep))
        )
        f += expect(ok, "real export docx (server->engine)", "body=%r" % ex)
    finally:
        httpd.shutdown()
        httpd.server_close()
    return f


if __name__ == "__main__":
    failures = run()
    print("\n=== REAL-CONFIG E2E ===")
    failures += check_real_config()
    print("=======================")
    if failures:
        print("\nFAILURES: %d" % failures)
        sys.exit(1)
    print("\nALL CORE ENDPOINTS PASSED")
    sys.exit(0)
