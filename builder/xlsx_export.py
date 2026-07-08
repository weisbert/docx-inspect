#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export a single report table block to a .xlsx that visually mirrors the Word
table (yellow header band, category vertical merge, sim_span horizontal merge,
condition-row shading, RED+bold out-of-spec cells).

The compliance layout REUSES tables.py's structural helpers (make_groups,
_plan_columns, _axis_value, flag_positions, _fmt_val) so the spreadsheet can never
drift from the docx renderer -- the column plan, per-axis values and flag logic
are computed by exactly the same code the engine uses. Only the cell painting
differs (openpyxl instead of python-docx). ASCII-only / no company data.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import tables  # sibling module: the single source of truth for table structure

DEFAULT_SIZE = 10                       # readable in Excel (Word uses a denser 7pt)
_THIN = Side(style="thin", color="FF000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)


# ---------------------------------------------------------------------------
# Low-level cell helpers.
# ---------------------------------------------------------------------------
def _argb(hex6):
    """Normalize a hex colour to openpyxl ARGB ('FFRRGGBB'), or None for auto/blank."""
    if not hex6:
        return None
    s = str(hex6).lstrip("#").strip()
    if not s or s.lower() == "auto":
        return None
    if len(s) == 6:
        return "FF" + s.upper()
    if len(s) == 8:
        return s.upper()
    return None


def _fill(hex6):
    a = _argb(hex6)
    return PatternFill(fill_type="solid", fgColor=a, bgColor=a) if a else None


def _cm_to_width(cm):
    """Approximate cm -> Excel column-width units (~5.4 chars/cm at the default
    font), floored so narrow gutter columns stay visible."""
    try:
        return round(max(1.2, float(cm) * 4.9), 1)
    except (TypeError, ValueError):
        return 8.0


def _coerce(val):
    """Numbers stay numeric so Excel does not flag "number stored as text"; a value
    is only coerced when it round-trips to the SAME string (so a padded "0.50" or a
    "value(CORNER)" token keeps its exact display). None -> ""."""
    if val is None:
        return ""
    if isinstance(val, (int, float, bool)):
        return val
    s = str(val)
    st = s.strip()
    if st == "":
        return ""
    try:
        f = float(st)
        n = int(f) if f.is_integer() else f
        if str(n) == st:
            return n
    except ValueError:
        pass
    return s


def _paint(cell, fill):
    cell.border = BORDER
    cell.alignment = CENTER
    if fill:
        cell.fill = fill


def _text(cell, val, bold=False, color=None):
    cell.value = _coerce(val)
    cell.font = Font(bold=bold, color=_argb(color), size=DEFAULT_SIZE)


def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Compliance datatable -> xlsx (mirrors tables.render_datatable).
# ---------------------------------------------------------------------------
def build_datatable_xlsx(data, comp_cfg):
    """``data`` = block['data']; ``comp_cfg`` = template config's 'compliance' section
    (col_w_cm, fills, setting_kinds, flag_color, axis_labels). Returns xlsx bytes."""
    w = comp_cfg["col_w_cm"]
    fills = comp_cfg["fills"]
    setting_kinds = set(comp_cfg.get("setting_kinds", ["common_setting", "module_setting", "tb"]))
    flag_color = comp_cfg.get("flag_color", "FF0000")

    groups = tables.make_groups(data, comp_cfg)
    show_spec_col = data.get("show_spec", True) and not any(g["role"] == "spec" for g in groups)
    plan = tables._plan_columns(groups, show_spec_col, w)
    ncols = len(plan)
    rows = data.get("rows", []) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    for idx, p in enumerate(plan):
        ws.column_dimensions[get_column_letter(idx + 1)].width = _cm_to_width(p["w"])

    col_of = {}
    group_axis_cols = {g["key"]: [] for g in groups}
    for idx, p in enumerate(plan):
        if p["kind"] in ("cat", "item", "spec", "unit"):
            col_of[p["kind"]] = idx
        elif p["kind"] == "axis":
            col_of[("axis", p["group"], p["axis"])] = idx
            group_axis_cols[p["group"]].append(idx)

    header_fill = _fill(fills["header"])

    # ---- header band: 3 rows ----
    for r in range(3):
        for c in range(ncols):
            _paint(ws.cell(row=r + 1, column=c + 1), header_fill)
    for key, label in (("cat", "Category"), ("item", "Item"), ("spec", "Spec"), ("unit", "Unit")):
        if key in col_of:
            c = col_of[key] + 1
            ws.merge_cells(start_row=1, start_column=c, end_row=3, end_column=c)
            _text(ws.cell(row=1, column=c), label, bold=True)
    for g in groups:
        cc = group_axis_cols[g["key"]]
        ws.merge_cells(start_row=1, start_column=cc[0] + 1, end_row=1, end_column=cc[-1] + 1)
        _text(ws.cell(row=1, column=cc[0] + 1), g["title"], bold=True)
        if g.get("stage"):
            ws.merge_cells(start_row=2, start_column=cc[0] + 1, end_row=2, end_column=cc[-1] + 1)
            _text(ws.cell(row=2, column=cc[0] + 1), g["stage"], bold=True)
        for ai, ax in enumerate(g["axes"]):
            _text(ws.cell(row=3, column=col_of[("axis", g["key"], ai)] + 1), ax, bold=True)

    # ---- data rows: category runs (mirror render_datatable) ----
    start = 3                                   # 0-based count of header rows
    sim_groups = [g for g in groups if g["role"] == "sim"]
    first_sim = sim_groups[0] if sim_groups else None

    catg, i = [], 0
    while i < len(rows):
        j = i
        while j + 1 < len(rows) and rows[j + 1].get("cat") == rows[i].get("cat"):
            j += 1
        catg.append((i, j))
        i = j + 1

    for (g0, g1) in catg:
        for gi in range(g0, g1 + 1):
            row = rows[gi]
            xr = start + gi + 1                 # openpyxl row (1-based)
            band = _fill(fills["setting"] if row.get("kind") in setting_kinds else fills["result"])
            flags = tables.flag_positions(row)
            # full-span: a sim_span value merges across ALL sim groups' axis columns
            # into one wide cell (mirror render_datatable). Show the first non-null
            # axis value; blank the covered cells so the merge is clean.
            span = bool(row.get("sim_span"))
            span_cols = sorted(c for g in sim_groups
                               for c in group_axis_cols[g["key"]]) if span else []
            span_first_col = span_cols[0] if span_cols else None
            span_val = None
            if span:
                for g in sim_groups:
                    for ai in range(len(g["axes"])):
                        vv = tables._axis_value(row, g["key"], ai)
                        if vv is not None and str(vv) != "":
                            span_val = vv
                            break
                    if span_val is not None:
                        break
            for idx, p in enumerate(plan):
                cell = ws.cell(row=xr, column=idx + 1)
                _paint(cell, band)
                if p["kind"] == "cat":
                    continue                    # value + vertical merge handled per group
                if p["kind"] == "item":
                    _text(cell, row.get("item", ""))
                elif p["kind"] == "unit":
                    _text(cell, row.get("unit", ""))
                elif p["kind"] == "spec":
                    _text(cell, row.get("spec"))
                elif p["kind"] == "axis":
                    if span and p["role"] == "sim":
                        if idx == span_first_col:
                            _text(cell, tables._fmt_val(span_val))
                        continue                # covered cells blanked; merged below
                    v = tables._axis_value(row, p["group"], p["axis"])
                    red = (p["role"] == "sim" and p["axis"] in flags)
                    _text(cell, tables._fmt_val(v), bold=red, color=(flag_color if red else None))
                # spacer: painted, left blank
            if span and span_cols:
                lo, hi = span_cols[0], span_cols[-1]
                if hi > lo:
                    ws.merge_cells(start_row=xr, start_column=lo + 1,
                                   end_row=xr, end_column=hi + 1)
        # vertical merge of the category column across the run (first row's band)
        cc = col_of["cat"]
        xr0, xr1 = start + g0 + 1, start + g1 + 1
        band0 = _fill(fills["setting"] if rows[g0].get("kind") in setting_kinds else fills["result"])
        for xr in range(xr0, xr1 + 1):
            _paint(ws.cell(row=xr, column=cc + 1), band0)
        if xr1 > xr0:
            ws.merge_cells(start_row=xr0, start_column=cc + 1, end_row=xr1, end_column=cc + 1)
        top = ws.cell(row=xr0, column=cc + 1)
        _text(top, rows[g0].get("cat", ""), bold=True)

    ws.freeze_panes = "A4"                       # keep the 3 header rows on scroll
    return _save(wb)


# ---------------------------------------------------------------------------
# Free table -> xlsx (mirrors tables.render_free_table).
# ---------------------------------------------------------------------------
def build_free_table_xlsx(rows, header_rows=1, merges=None, col_w=None,
                          row_fills=None, header_fill="D9D9D9"):
    rows = rows or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"
    if not rows:
        return _save(wb)

    ncols = max(len(r) for r in rows)
    rfills = {int(k): v for k, v in (row_fills or {}).items()}
    hfill = _fill(header_fill)

    for r, rowvals in enumerate(rows):
        band = hfill if r < header_rows else (_fill(rfills[r]) if r in rfills else None)
        for c in range(ncols):
            val = rowvals[c] if c < len(rowvals) else ""
            cell = ws.cell(row=r + 1, column=c + 1)
            _paint(cell, band)
            runs = val.get("runs") if isinstance(val, dict) else None
            if isinstance(runs, list):
                # openpyxl has one font per cell: flatten rich runs (concatenate
                # text; bold/italic/colour if ANY run carries it).
                text = "".join(str(rn.get("t", "")) for rn in runs)
                bold = (r < header_rows) or any(rn.get("b") for rn in runs)
                italic = any(rn.get("i") for rn in runs)
                color = next((rn.get("color") for rn in runs if rn.get("color")), None)
                cell.value = _coerce(text)
                cell.font = Font(bold=bold, italic=italic, color=_argb(color), size=DEFAULT_SIZE)
            else:
                _text(cell, val, bold=(r < header_rows))

    if col_w:
        for idx, cw in enumerate(col_w):
            if idx < ncols:
                ws.column_dimensions[get_column_letter(idx + 1)].width = _cm_to_width(cw)
    for m in (merges or []):
        r, c = m["r"], m["c"]
        rs, cs = m.get("rs", 1), m.get("cs", 1)
        if rs > 1 or cs > 1:
            ws.merge_cells(start_row=r + 1, start_column=c + 1,
                           end_row=r + rs, end_column=c + cs)
    if header_rows:
        ws.freeze_panes = "A%d" % (header_rows + 1)
    return _save(wb)


# ---------------------------------------------------------------------------
# Dispatch + filename.
# ---------------------------------------------------------------------------
def _slug(s):
    out = "".join(ch if (ch.isalnum() or ch in " -_") else "_" for ch in str(s or "")).strip()
    return (out or "table").replace(" ", "_")[:60]


def filename_for(block):
    cap = block.get("caption") or ("compliance" if block.get("type") == "datatable" else "table")
    return _slug(cap) + ".xlsx"


def build_block_xlsx(block, cfg):
    """Build xlsx bytes for a table/datatable block. ``cfg`` = full template config
    (needs its 'compliance' / 'free_table' sections). Raises ValueError for a
    non-table block or a datatable with no compliance config."""
    btype = block.get("type")
    if btype == "datatable":
        comp = (cfg or {}).get("compliance")
        if not comp:
            raise ValueError("datatable export needs a template 'compliance' config")
        return build_datatable_xlsx(block.get("data") or {}, comp)
    if btype == "table":
        free = (cfg or {}).get("free_table") or {}
        return build_free_table_xlsx(
            block.get("rows") or [], block.get("header_rows", 1),
            block.get("merges"), block.get("col_w"), block.get("row_fills"),
            free.get("header_fill", "D9D9D9"))
    raise ValueError("block type %r is not an exportable table" % btype)
