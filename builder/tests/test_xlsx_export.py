#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Regression tests for xlsx_export: a table/datatable block -> a .xlsx that
visually mirrors the Word table. Builds bytes, re-parses with openpyxl, and
asserts the replica: out-of-spec cells RED+bold, in-spec plain, category vertical
merge, 3-row header band with the config fill, sim_span horizontal merge, and free
tables (rich runs flattened + merges). Reuses the golden fixture's config/data so
the compliance structure is exercised end-to-end. Neutral / ASCII-only.
"""
import io
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))  # builder/tests
BUILDER = os.path.dirname(HERE)                  # builder/
if BUILDER not in sys.path:
    sys.path.insert(0, BUILDER)
import buildpath  # noqa: E402,F401  (registers core/docx_io/store/sync/web)
import openpyxl                       # noqa: E402
import xlsx_export as X               # noqa: E402
import test_render_golden as G        # noqa: E402

_fails = []


def check(cond, name):
    print(("  PASS  " if cond else "  FAIL  ") + name)
    if not cond:
        _fails.append(name)


def _load(b):
    return openpyxl.load_workbook(io.BytesIO(b)).active


def _find(ws, val):
    for row in ws.iter_rows():
        for c in row:
            if c.value == val:
                return c
    return None


def _is_red(cell):
    col = cell.font.color
    return bool(col and getattr(col, "rgb", None) and str(col.rgb).upper().endswith("FF0000"))


def main():
    comp = G.golden_config()["compliance"]
    data = G.golden_project()["outline"][0]["blocks"][2]["data"]   # the datatable
    ws = _load(X.build_datatable_xlsx(data, comp))

    print("== compliance datatable ==")
    c758 = _find(ws, 758)   # I_total sim MAX 758 > spec MAX 500 -> flagged
    check(c758 is not None and _is_red(c758) and c758.font.bold,
          "over-spec 758 is RED + bold")
    c490 = _find(ws, 490)
    check(c490 is not None and not _is_red(c490), "in-spec 490 is not red")
    c1200 = _find(ws, 1200)   # P_static NTWC 1200 > spec_ntwc 1000 -> flagged
    check(c1200 is not None and _is_red(c1200), "over-spec NTWC 1200 is RED")

    hvals = set()
    for r in ws.iter_rows(min_row=1, max_row=3):
        for c in r:
            if isinstance(c.value, str):
                hvals.add(c.value)
    check({"Category", "Item", "Unit"} <= hvals, "header labels Category/Item/Unit present")
    check("Pilot" in hvals and "Spec" in hvals, "group titles Spec + Pilot in header band")

    cat_merges = [m for m in ws.merged_cells.ranges
                  if m.min_col == 1 and m.max_col == 1 and m.max_row > m.min_row]
    check(cat_merges and "Power" in {ws.cell(row=m.min_row, column=1).value for m in cat_merges},
          "category column has a vertical merge holding 'Power'")

    a1 = ws.cell(row=1, column=1)
    check(a1.fill and a1.fill.fgColor and
          str(a1.fill.fgColor.rgb).upper().endswith(comp["fills"]["header"].upper()),
          "header band uses the config header fill")

    print("== sim_span ==")
    span_data = {"spec_name": "Spec", "show_spec": False,
                 "sims": [{"key": "s1", "title": "S1", "stage": "Post", "axes": ["MIN", "TYP", "MAX"]}],
                 "rows": [{"cat": "C", "item": "single", "kind": "result", "unit": "mV",
                           "sim_mtm": ["12.3", "12.3", "12.3"], "sim_span": True, "limit": None}]}
    wss = _load(X.build_datatable_xlsx(span_data, comp))
    horiz = [m for m in wss.merged_cells.ranges
             if m.max_col > m.min_col and m.min_row == m.max_row and m.min_row > 3]
    check(len(horiz) >= 1, "sim_span merges the sim MIN/TYP/MAX cells horizontally")

    print("== multi-sim-group flags (per-group, not the flat-schema helper) ==")
    # row_a: two sim groups, ONLY the second violates its own limit. A stale flat
    # sim_mtm also sits on the row (real rows can carry one left over) with a
    # value that WOULD flag under the flat-schema helper (tables.flag_positions) --
    # the fix must ignore it and flag purely from each group's own
    # row["sims"][key] values, so it can never bleed into the other group.
    # row_b: filled ONLY via row["sims"] (no flat sim_mtm key at all) and
    # violates -- the flat-schema helper sees no sim_mtm and would flag nothing.
    multi_data = {
        "spec_name": "Spec", "show_spec": True,
        "sims": [{"key": "g1", "title": "G1", "stage": "Pre"},
                 {"key": "g2", "title": "G2", "stage": "Post"}],
        "rows": [
            {"cat": "X", "item": "row_a", "unit": "mV", "kind": "result",
             "spec": 100, "spec_mtm": [None, 100, None],
             "sim_mtm": [150, 20, 30],                  # stale flat: flags index 0
             "sims": {"g1": {"mtm": [11, 21, 31]},       # in spec for g1
                      "g2": {"mtm": [41, 51, 121]}},     # MAX 121 > 100 -> only g2
             "limit": "le", "sim_span": False},
            {"cat": "X", "item": "row_b", "unit": "mV", "kind": "result",
             "spec": 60, "spec_mtm": [None, 60, None],
             "sims": {"g1": {"mtm": [5, 6, 777]},        # MAX 777 > 60 -> violates
                      "g2": {"mtm": [1, 2, 3]}},
             "limit": "le", "sim_span": False},
        ],
    }
    wm = _load(X.build_datatable_xlsx(multi_data, comp))

    c121 = _find(wm, 121)
    check(c121 is not None and _is_red(c121) and c121.font.bold,
          "row_a: g2 MAX 121 (over its own spec) is RED + bold")
    c41 = _find(wm, 41)
    check(c41 is not None and not _is_red(c41), "row_a: g2 MIN 41 (in spec) is not red")
    c51 = _find(wm, 51)
    check(c51 is not None and not _is_red(c51), "row_a: g2 TYP 51 (in spec) is not red")
    c11 = _find(wm, 11)
    check(c11 is not None and not _is_red(c11),
          "row_a: g1 MIN 11 is not red (stale flat sim_mtm[0]=150 must not leak into g1)")
    c31 = _find(wm, 31)
    check(c31 is not None and not _is_red(c31), "row_a: g1 MAX 31 (in spec for g1) is not red")

    c777 = _find(wm, 777)
    check(c777 is not None and _is_red(c777) and c777.font.bold,
          "row_b: g1 MAX 777 (over spec, row has NO flat sim_mtm at all) is RED + bold")
    c3 = _find(wm, 3)
    check(c3 is not None and not _is_red(c3), "row_b: g2 MAX 3 (in spec) is not red")

    print("== free table ==")
    rows = [["Style", "Example"],
            ["bold+red", {"runs": [{"t": "Bold", "b": True}, {"t": " red", "color": "FF0000"}]}],
            ["plain", "hi"]]
    wf = _load(X.build_free_table_xlsx(rows, header_rows=1,
                                       merges=[{"r": 0, "c": 0, "rs": 1, "cs": 2}]))
    check(wf.cell(row=1, column=1).value == "Style" and wf.cell(row=1, column=1).font.bold,
          "header row is bold")
    check(wf.cell(row=2, column=2).value == "Bold red" and wf.cell(row=2, column=2).font.bold,
          "rich runs flattened (text + bold)")
    check(any(m.min_row == 1 and m.max_col == 2 for m in wf.merged_cells.ranges),
          "header merge applied")

    print("== dispatch / filename ==")
    check(X.filename_for({"type": "datatable", "caption": "Compliance results!"})
          == "Compliance_results_.xlsx", "filename_for slugs the caption")
    try:
        X.build_block_xlsx({"type": "para"}, {})
        check(False, "non-table block raises ValueError")
    except ValueError:
        check(True, "non-table block raises ValueError")

    print("\n== SUMMARY ==  %s" % ("ALL PASSED" if not _fails else "%d FAILED" % len(_fails)))
    return 1 if _fails else 0


if __name__ == "__main__":
    sys.exit(main())
